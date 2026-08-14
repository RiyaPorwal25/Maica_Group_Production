from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
# from django.contrib.auth.models import User
from django.core.paginator import Paginator 
from .models import Product, Production, Stock, Machine, CustomUser, Color, Thickness, Density, Size, Length, Height, Width, format_float, MachineOperator, RolePermission, RawMaterial, RawMaterialStock, Formulation, FormulationItem, PurchaseInvoice, Supplier, Planning, WeightSheet, WeightSheetRow, UnfinishedProduction, ScrapLog, DensityName, GroupOrder, DeliveryChallan, Party, Dashboard, UserSession, BatchAddHistory, FormulationBatch
from django.db.models import FloatField, Max, Sum, Case, When, IntegerField, Value, F, Q, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.conf import settings
from pathlib import Path
from .forms import CustomUserCreateForm, CustomUserUpdateForm, ProductForm, MachineForm
from .decorators import role_required, admin_only, admin_manager_only, module_required
import openpyxl
import csv
from django.db import IntegrityError, transaction
import pandas as pd
import json
from decimal import Decimal
import sumy
from django.utils.timezone import now
from datetime import datetime, timedelta
from django.forms import modelformset_factory
from django.utils.dateparse import parse_date
from collections import OrderedDict, defaultdict
from datetime import date
from .stock_alerts import queue_low_raw_material_alert


# ==================== LOGGING SETUP ====================
import logging
errors_logger = logging.getLogger('errors')
audit_logger = logging.getLogger('audit')
business_logger = logging.getLogger('business')





# Login page
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            audit_logger.info(f"User {user.username} logged in successfully")
            # Create UserSession record
            today = timezone.localdate()
            UserSession.objects.create(
                user=user,
                login_time=timezone.now(),
                date=today
            )
            if user.role in ['Admin', 'Manager']:
                queue_low_raw_material_alert(request)
            if user.role == 'Operator':
                return redirect('production_2_0_door')
            return redirect('dashboard_v2')
        else:
            error = "Invalid username or password"
            return render(request, 'index.html', {'error': error})
    return render(request, 'index.html')


def logout_view(request):
    if request.user.is_authenticated:
        user = request.user
        # Update the latest open session with logout time
        UserSession.objects.filter(user=user, logout_time__isnull=True).update(logout_time=timezone.now())
        logout(request)
        audit_logger.info(f"User {user.username} logged out successfully")
    return redirect('login')


#product page 
@module_required('master')
def product_list(request):
    category = request.GET.get('category')
    products = Product.objects.all().order_by('id')

    #  APPLY CATEGORY FILTER
    if category == "Frame":
        products = products.filter(category="Frame")
    elif category == "DoorSheet":
        products = products.filter(category__in=["Door", "Sheet"])

    #  APPLY OTHER FILTERS
    thickness_id = request.GET.get('thickness')
    length_id = request.GET.get('length')
    color_id = request.GET.get('color')
    stamp = request.GET.get('stamp')

    if thickness_id:
        products = products.filter(thickness_id=thickness_id)
    if length_id:
        products = products.filter(length_id=length_id)
    if color_id:
        products = products.filter(color_id=color_id)
    if stamp:
        products = products.filter(stamp=stamp)

    # Handle Add Product
    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                return redirect('product_list')
            except IntegrityError:
                messages.error(request, "A product with this configuration already exists.")
        # If error, fall through to render with form and messages
    else:
        form = ProductForm()

    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    start_serial = (page_obj.number - 1) * paginator.per_page

    return render(request, 'product.html', {
        'page_obj': page_obj,
        'form': form,
        'start_serial': start_serial,
        'selected_category': category,
        'all_colors': Color.objects.filter(is_active=True).order_by('color'),
        'all_thicknesses': Thickness.objects.filter(is_active=True).order_by('thickness'),
        'all_lengths': Length.objects.filter(is_active=True).order_by('length'),
        'all_stamps': Product.STAMP_CHOICES,
        'selected_thickness': thickness_id,
        'selected_length': length_id,
        'selected_color': color_id,
        'selected_stamp': stamp,

    })

@login_required
def delete_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    
    # Delete related stock entries explicitly (optional, cascade might handle it)
    product.stock_movements.all().delete()

    # Delete the product itself
    product.delete()

    messages.success(request, f"Product '{product.product_name}' and related stock removed successfully.")
    return redirect('product_list')  # Replace with your product list URL name


# products stock page 
def _handle_product_dispatch(request, redirect_name):
    dispatch_data = request.POST.get("dispatch_data")

    if not dispatch_data:
        messages.error(request, "No products selected")
        return redirect(redirect_name)

    try:
        items = json.loads(dispatch_data)

        for item in items:
            product_id = item["id"]
            quantity = int(item["qty"])

            product = Product.objects.get(id=product_id)

            stock_data = Stock.objects.filter(product=product).aggregate(
                total_in=Sum(
                    Case(
                        When(movement_type='IN', then=F('quantity')),
                        output_field=IntegerField()
                    )
                ),
                total_out=Sum(
                    Case(
                        When(movement_type='OUT', then=F('quantity')),
                        output_field=IntegerField()
                    )
                )
            )

            current_stock = (stock_data['total_in'] or 0) - (stock_data['total_out'] or 0)

            if quantity > current_stock:
                messages.error(request, f"Not enough stock for {product.product_name}")
                return redirect(redirect_name)

            Stock.objects.create(
                product=product,
                quantity=quantity,
                movement_type='OUT',
                operator=request.user
            )

        messages.success(request, "Dispatch completed successfully")

    except Exception as e:
        messages.error(request, f"Error: {str(e)}")

    return redirect(redirect_name)


def _products_with_stock():
    return Product.objects.annotate(
        total_in=Coalesce(
            Sum(
                Case(
                    When(stock_movements__movement_type='IN', then='stock_movements__quantity'),
                    output_field=IntegerField()
                )
            ),
            Value(0)
        ),
        total_out=Coalesce(
            Sum(
                Case(
                    When(stock_movements__movement_type='OUT', then='stock_movements__quantity'),
                    output_field=IntegerField()
                )
            ),
            Value(0)
        )
    ).annotate(
        stock=F('total_in') - F('total_out')
    ).order_by('id')


@module_required('stock')
def stock_view(request):
    # Dispatch only sees products and can dispatch, Manager and Admin see everything
    query = request.GET.get('q', '')
    category = request.GET.get('category')
    products = Product.objects.all().order_by('id')

    #  APPLY CATEGORY FILTER
    if category == "Frame":
        products = products.filter(category="Frame")
    elif category == "DoorSheet":
        products = products.filter(category__in=["Door", "Sheet"])

    #  APPLY OTHER FILTERS
    thickness_id = request.GET.get('thickness')
    length_id = request.GET.get('length')
    color_id = request.GET.get('color')
    stamp = request.GET.get('stamp')

    if thickness_id:
        products = products.filter(thickness_id=thickness_id)
    if length_id:
        products = products.filter(length_id=length_id)
    if color_id:
        products = products.filter(color_id=color_id)
    if stamp:
        products = products.filter(stamp=stamp)

    
    if request.method == "POST" and request.POST.get("form_type") == "dispatch":

        dispatch_data = request.POST.get("dispatch_data")

        if not dispatch_data:
            messages.error(request, "No products selected")
            return redirect('stock')

        try:
            items = json.loads(dispatch_data)

            for item in items:

                product_id = item["id"]
                quantity = int(item["qty"])

                product = Product.objects.get(id=product_id)

                stock_data = Stock.objects.filter(product=product).aggregate(
                    total_in=Sum(
                        Case(
                            When(movement_type='IN', then=F('quantity')),
                            output_field=IntegerField()
                        )
                    ),
                    total_out=Sum(
                        Case(
                            When(movement_type='OUT', then=F('quantity')),
                            output_field=IntegerField()
                        )
                    )
                )

                current_stock = (stock_data['total_in'] or 0) - (stock_data['total_out'] or 0)

                if quantity > current_stock:
                    messages.error(request, f"Not enough stock for {product.product_name}")
                    return redirect('stock')

                Stock.objects.create(
                    product=product,
                    quantity=quantity,
                    movement_type='OUT',
                    operator=request.user
                )

            messages.success(request, "Dispatch completed successfully")

        except Exception as e:
            errors_logger.error(f"Dispatch failed in stock_view: {str(e)}", exc_info=True)
            messages.error(request, f"Error: {str(e)}")

        return redirect('stock')


    products = products.annotate(
        total_in=Coalesce(
            Sum(
                Case(
                    When(stock_movements__movement_type='IN', then='stock_movements__quantity'),
                    output_field=IntegerField()
                )
            ),
            Value(0)
        ),
        total_out=Coalesce(
            Sum(
                Case(
                    When(stock_movements__movement_type='OUT', then='stock_movements__quantity'),
                    output_field=IntegerField()
                )
            ),
            Value(0)
        )
    ).annotate(
        stock=F('total_in') - F('total_out')
    ).order_by('id')

     # ---- Apply Search Filter ----
    if query:
        products = products.filter(
            Q(product_name__icontains=query) |
            Q(category__icontains=query)
        )

    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'stock.html', {
        'page_obj': page_obj,
        'machines': Machine.objects.all(),
        # 'products': Product.objects.only('id', 'product_name'),
        'products': products,
        'query': query,
        'current_category': category,
        'all_colors': Color.objects.filter(is_active=True).order_by('color'),
        'all_thicknesses': Thickness.objects.filter(is_active=True).order_by('thickness'),
        'all_lengths': Length.objects.filter(is_active=True).order_by('length'),
        'all_stamps': Product.STAMP_CHOICES,
        'selected_thickness': thickness_id,
        'selected_length': length_id,
        'selected_color': color_id,
        'selected_stamp': stamp,
    })

#home page
@login_required
def dashboard(request):

    today = timezone.now().date()

    # =========================
    # TODAY SUMMARY
    # =========================
    today_production_qs = Production.objects.filter(created_at__date=today)
    
    # Filter for operators
    if request.user.role == 'Operator':
        today_production_qs = today_production_qs.filter(operator=request.user)
    
    # Door
    door_agg = today_production_qs.filter(product__category='Door').aggregate(
        total_weight=Sum(F('quantity') * F('weight_per_piece')),
        total_quantity=Sum('quantity')
    )
    door_total = door_agg['total_weight'] or 0
    door_quantity = door_agg['total_quantity'] or 0

    # Frame
    frame_agg = today_production_qs.filter(product__category='Frame').aggregate(
        total_weight=Sum(F('quantity') * F('weight_per_piece')),
        total_quantity=Sum('quantity')
    )
    frame_total = frame_agg['total_weight'] or 0
    frame_quantity = frame_agg['total_quantity'] or 0

    # Sheet
    sheet_agg = today_production_qs.filter(product__category='Sheet').aggregate(
        total_weight=Sum(F('quantity') * F('weight_per_piece')),
        total_quantity=Sum('quantity')
    )
    sheet_total = sheet_agg['total_weight'] or 0
    sheet_quantity = sheet_agg['total_quantity'] or 0

    today_operators = today_production_qs.values('operator').distinct().count()
    pending_count_qs = Production.objects.filter(status='Pending')
    if request.user.role == 'Operator':
        pending_count_qs = pending_count_qs.filter(operator=request.user)
    pending_count = pending_count_qs.count()

   
    production_qs = Production.objects.all()
    if request.user.role == 'Operator':
        production_qs = production_qs.filter(operator=request.user)
    dispatch_qs = Stock.objects.filter(movement_type='OUT')
    if request.user.role == 'Operator':
        dispatch_qs = dispatch_qs.filter(operator=request.user)

    operators = CustomUser.objects.all()
    machines = Machine.objects.all()

    # yestaerdays data
    yesterday = timezone.now().date() - timedelta(days=1)

    y_qs = Production.objects.filter(created_at__date=yesterday)

    if request.user.role == 'Operator':
        y_qs = y_qs.filter(operator=request.user)

    y_door = y_qs.filter(product__category='Door').aggregate(
        total_weight=Sum(F('quantity') * F('weight_per_piece')),
        total_quantity=Sum('quantity')
    )

    y_frame = y_qs.filter(product__category='Frame').aggregate(
        total_weight=Sum(F('quantity') * F('weight_per_piece')),
        total_quantity=Sum('quantity')
    )

    y_sheet = y_qs.filter(product__category='Sheet').aggregate(
        total_weight=Sum(F('quantity') * F('weight_per_piece')),
        total_quantity=Sum('quantity')
    )
    y_door_total = y_door['total_weight'] or 0
    y_door_quantity = y_door['total_quantity'] or 0

    y_frame_total = y_frame['total_weight'] or 0
    y_frame_quantity = y_frame['total_quantity'] or 0

    y_sheet_total = y_sheet['total_weight'] or 0
    y_sheet_quantity = y_sheet['total_quantity'] or 0

    # ---- Dashboard report filter ----
    report_date = request.GET.get("report_date", "")
    report_rows = []
    summary_data = {}
    raw_report_rows = []
    if report_date:
        try:
            report_date_obj = timezone.datetime.strptime(report_date, "%Y-%m-%d").date()
            report_qs = Production.objects.filter(created_at__date=report_date_obj)
            if request.user.role == 'Operator':
                report_qs = report_qs.filter(operator=request.user)
            raw_report_rows = report_qs

            report_map = {}
            for p in report_qs:
                key = (p.machine.name if p.machine else "-", p.product.category if p.product else "-", p.shift ,p.operator.username if p.operator else "-")
                row = report_map.setdefault(key, {
                    "machine": key[0],
                    "category": key[1],
                    "shift": key[2],
                    "operator": key[3], 
                    "total_production": 0,
                    "actual_production": 0,
                    "line_setting": 0,
                    "side_patti": 0,
                    "qty": 0,
                    "sizes": set(),
                    "items": set(),
                })
                row["total_production"] += p.overall_production
                row["actual_production"] += p.actual_production
                row["line_setting"] += p.linesetting
                row["side_patti"] += p.sidepatti
                row["qty"] += p.quantity
                if p.size:
                    row["sizes"].add(str(p.size))
                if p.product:
                    row["items"].add(str(p.product))

            report_rows = [
                {
                    **row,
                    "sizes": ", ".join(sorted(row["sizes"])),
                    "items": ", ".join(sorted(row["items"]))
                }
                for row in report_map.values()
            ]

            summary = defaultdict(lambda: {
                "day_prod": 0, "day_qty": 0, "day_weight": 0,"day_line_setting":0,"day_side_patti":0,
                "night_prod": 0, "night_qty": 0, "night_weight": 0,"night_line_setting":0,"night_side_patti":0,
            })

            for p in report_qs:
                machine = p.machine.name if p.machine else "-"

                if p.shift == "Day":
                    summary[machine]["day_prod"] += p.overall_production
                    summary[machine]["day_qty"] += p.quantity
                    summary[machine]["day_weight"] += p.actual_production
                    summary[machine]["day_line_setting"] += p.linesetting
                    summary[machine]["day_side_patti"] += p.sidepatti

                else:
                    summary[machine]["night_prod"] += p.overall_production
                    summary[machine]["night_qty"] += p.quantity
                    summary[machine]["night_weight"] += p.actual_production
                    summary[machine]["night_line_setting"] += p.linesetting
                    summary[machine]["night_side_patti"] += p.sidepatti

            # summary_data = dict(summary)
            ordered_summary = {}

            machines = Machine.objects.all().order_by('name')  # or 'id'

            for m in machines:
                machine_name = m.name
                if machine_name in summary:
                    ordered_summary[machine_name] = summary[machine_name]

            summary_data = ordered_summary
        except ValueError:
            report_date = ""

    return render(request, 'dashboard.html', {
        'door_total': door_total,
        'door_quantity': door_quantity,
        'frame_total': frame_total,
        'frame_quantity': frame_quantity,
        'sheet_total': sheet_total,
        'sheet_quantity': sheet_quantity,
        'today_operators': today_operators,
        'pending_count': pending_count,
        'operators': operators,
        'machines': machines,
        'production_qs': production_qs,
        'dispatch_qs': dispatch_qs,
        'report_date': report_date,
        'report_rows': report_rows,
        'summary_data': summary_data,
        'raw_report_rows': raw_report_rows,
        'y_door_total': y_door_total,
        'y_door_quantity': y_door_quantity,
        'y_frame_total': y_frame_total,
        'y_frame_quantity': y_frame_quantity,
        'y_sheet_total': y_sheet_total,
        'y_sheet_quantity': y_sheet_quantity,
    })


def export_dashboard_report(request):
    report_date = request.GET.get("report_date")

    wb = openpyxl.Workbook()

    # ================= SUMMARY SHEET =================
    ws1 = wb.active
    ws1.title = "Summary"

    ws1.append([
        "Machine", "Shift", "Total Qty", "Total Weight",
        "Side Patti", "Line Setting", "Overall Production"
    ])

    if report_date:
        report_qs = Production.objects.filter(created_at__date=report_date)

        summary = {}

        for p in report_qs:
            machine = p.machine.name if p.machine else "-"
            shift = p.shift

            key = (machine, shift)

            if key not in summary:
                summary[key] = {
                    "qty": 0,
                    "weight": 0,
                    "side": 0,
                    "line": 0,
                    "overall": 0,
                }

            summary[key]["qty"] += p.quantity
            summary[key]["weight"] += p.actual_production
            summary[key]["side"] += p.sidepatti
            summary[key]["line"] += p.linesetting
            summary[key]["overall"] += p.overall_production

        for (machine, shift), data in summary.items():
            ws1.append([
                machine,
                shift,
                data["qty"],
                float(data["weight"]),
                float(data["side"]),
                float(data["line"]),
                float(data["overall"]),
            ])

    # ================= DETAIL SHEET =================
    ws2 = wb.create_sheet(title="Details")

    ws2.append([
        "Machine", "Shift", "Operator", "Product",
        "Qty", "Size", "Actual Production",
        "Line Setting", "Side Patti", "Overall Production"
    ])

    if report_date:
        report_qs = Production.objects.filter(created_at__date=report_date)

        for p in report_qs:
            ws2.append([
                p.machine.name if p.machine else "",
                p.shift,
                p.operator.full_name if p.operator else "",
                p.product.product_name if p.product else "",
                p.quantity,
                str(p.size) if p.size else "",
                float(p.actual_production),
                float(p.linesetting),
                float(p.sidepatti),
                float(p.overall_production),
            ])

    # ================= RESPONSE =================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="dashboard_report_{report_date}.xlsx"'

    wb.save(response)
    return response


@login_required
@module_required('master')
def user_list(request):
    users = CustomUser.objects.all()

    create_form = CustomUserCreateForm()
    update_form = None

    # ADD USER
    if request.method == "POST" and request.POST.get("action") == "add":
        create_form = CustomUserCreateForm(request.POST)
        
        if create_form.is_valid():
            user = create_form.save()
            audit_logger.info(f"User created: {user.username} (ID: {user.id}) by admin {request.user.username}")
            return redirect("user_list")
        else:
            errors_logger.error(f"User creation failed: {create_form.errors}")
            print(create_form.errors)  # debug
       
    # EDIT USER
    if request.method == "POST" and request.POST.get("action") == "edit":
        user_id = request.POST.get("user_id")
        user_obj = get_object_or_404(CustomUser, id=user_id)
        update_form = CustomUserUpdateForm(request.POST, instance=user_obj)
    
        if update_form.is_valid():
            user = update_form.save(commit=False)

            new_password = request.POST.get("new_password")

            # If admin entered a new password
            if new_password:
                user.set_password(new_password)

            user.save()
            audit_logger.info(f"User edited: {user.username} (ID: {user.id}) by admin {request.user.username}")

            return redirect("user_list")


    # DELETE USER
    if request.method == "POST" and request.POST.get("action") == "delete":
        user_id = request.POST.get("user_id")
        user_obj = get_object_or_404(CustomUser, id=user_id)
        username = user_obj.username
        user_obj.delete()
        audit_logger.info(f"User deleted: {username} (ID: {user_id}) by admin {request.user.username}")
        return redirect("user_list")

    context = {
        "users": users,
        "create_form": create_form,
        "roles": CustomUser.ROLE_CHOICES,

    }
    return render(request, "user.html", context)


# download stock of products
def export_stock_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="stock_mangalmurti.csv"'

    writer = csv.writer(response)
    writer.writerow(['Product ID', 'Product Name', 'Category', 'Total Stock'])

    query = request.GET.get('q', '')
    category = request.GET.get('category')
    products = Product.objects.all().order_by('id')

    # Apply filters same as stock_view
    if category == "Frame":
        products = products.filter(category="Frame")
    elif category == "DoorSheet":
        products = products.filter(category__in=["Door", "Sheet"])

    products = products.annotate(
        stock_in=Sum(
            Case(
                When(stock_movements__movement_type='IN',
                     then='stock_movements__quantity'),
                default=0,
                output_field=IntegerField()
            )
        ),
        stock_out=Sum(
            Case(
                When(stock_movements__movement_type='OUT',
                     then='stock_movements__quantity'),
                default=0,
                output_field=IntegerField()
            )
        )
    )

    # Apply search
    if query:
        products = products.filter(
            Q(product_name__icontains=query) |
            Q(category__icontains=query)
        )

    for product in products:
        total_stock = (product.stock_in or 0) - (product.stock_out or 0)

        writer.writerow([
            product.id,
            product.product_name,
            product.category,
            total_stock
        ])
    return response



# for both door and frame production
@login_required
@module_required('production')
def production_view(request):
    """Production page for Frame & Door"""
    User = get_user_model()

    #  Get category from URL (default = Frame)
    category = request.GET.get("category", "Frame")

    # ---- Add Production (operator) ----
    if request.method == "POST":
        # Collect form fields
        date_str = request.POST.get("date")
        date = timezone.datetime.strptime(date_str, "%Y-%m-%d") if date_str else timezone.now()

        machine_id = request.POST.get("machine")
        shift = request.POST.get("shift", "Day")
        scrap = float(request.POST.get("scrap", 0))
        operator_id = request.POST.get("operator")

        size_ids = request.POST.getlist("size")
        color_ids = request.POST.getlist("color")
        length_ids = request.POST.getlist("length")
        quantities = request.POST.getlist("quantity")
        weights = request.POST.getlist("weight")

        thickness_ids = request.POST.getlist("thickness")

        height_vals = request.POST.getlist("height")
        width_vals = request.POST.getlist("width")
        density_vals = request.POST.getlist("density")
        density_name_vals = request.POST.getlist("density_name")
        stamp_vals = request.POST.getlist("stamp")
        sidepattis = request.POST.getlist("sidepatti")
        line_settings = request.POST.getlist("line_setting")
        remark = request.POST.get("remark", "")


        # machine = get_object_or_404(Machine, id=machine_id)
        machine = get_object_or_404(Machine, id=machine_id, category=category)
        if request.user.role == 'Operator':
            operator = request.user
        else:
            operator = User.objects.get(id=operator_id) if operator_id else request.user

        #  Fix category for POST
        category = request.POST.get("category") or request.GET.get("category") or machine.category

        # Determine number of rows based on category
        if category == "Frame":
            num_rows = len(size_ids)
        else:
            num_rows = len(thickness_ids)

        for idx in range(num_rows):
            size = Size.objects.get(id=size_ids[idx]) if idx < len(size_ids) and size_ids[idx] else None
            color = Color.objects.get(id=color_ids[idx]) if idx < len(color_ids) and color_ids[idx] else None

            length_obj = None
            length_unit = request.POST.getlist('length_unit')[idx] if idx < len(request.POST.getlist('length_unit')) else 'ft'
            length_mm = request.POST.getlist('length_mm')[idx] if idx < len(request.POST.getlist('length_mm')) else ''

            if length_unit == 'mm' and length_mm:
                try:
                    mm_value = float(length_mm)
                    ft_value = mm_value / 304.8
                    length_obj, _ = Length.objects.get_or_create(length=ft_value, defaults={'unit': 'mm', 'original_value': mm_value})
                except ValueError:
                    length_obj = None
            else:
                length_obj = Length.objects.get(id=length_ids[idx]) if idx < len(length_ids) and length_ids[idx] else None

            length = length_obj

            quantity = int(quantities[idx]) if idx < len(quantities) and quantities[idx] else 0
            weight_per_piece = float(weights[idx]) if idx < len(weights) and weights[idx] else 0

            #  HANDLE BASED ON CATEGORY
            if category == "Door":

                thickness = Thickness.objects.get(id=thickness_ids[idx]) if idx < len(thickness_ids) and thickness_ids[idx] else None

                height_obj = None
                height_unit = request.POST.getlist('height_unit')[idx] if idx < len(request.POST.getlist('height_unit')) else 'inch'
                height_mm = request.POST.getlist('height_mm')[idx] if idx < len(request.POST.getlist('height_mm')) else ''

                if height_unit == 'mm' and height_mm:
                    try:
                        mm_value = float(height_mm)
                        inch_value = mm_value / 25.4
                        height_obj, _ = Height.objects.get_or_create(height=inch_value, defaults={'unit': 'mm', 'original_value': mm_value})
                    except ValueError:
                        height_obj = None
                else:
                    height_obj = Height.objects.get(id=height_vals[idx]) if idx < len(height_vals) and height_vals[idx] else None

                height = height_obj

                width_obj = None
                width_unit = request.POST.getlist('width_unit')[idx] if idx < len(request.POST.getlist('width_unit')) else 'inch'
                width_mm = request.POST.getlist('width_mm')[idx] if idx < len(request.POST.getlist('width_mm')) else ''

                if width_unit == 'mm' and width_mm:
                    try:
                        mm_value = float(width_mm)
                        inch_value = mm_value / 25.4
                        width_obj, _ = Width.objects.get_or_create(width=inch_value, defaults={'unit': 'mm', 'original_value': mm_value})
                    except ValueError:
                        width_obj = None
                else:
                    width_obj = Width.objects.get(id=width_vals[idx]) if idx < len(width_vals) and width_vals[idx] else None

                width = width_obj

                density = Density.objects.get(id=density_vals[idx]) if idx < len(density_vals) and density_vals[idx] else None
                density_name = DensityName.objects.get(id=density_name_vals[idx]) if idx < len(density_name_vals) and density_name_vals[idx] else None
                line_setting = float(line_settings[idx]) if idx < len(line_settings) and line_settings[idx] else 0
                sidepatti = float(sidepattis[idx]) if idx < len(sidepattis) and sidepattis[idx] else 0
                stamp = stamp_vals[idx] if idx < len(stamp_vals) and stamp_vals[idx] else ""

            else:  #  FRAME
                stamp = stamp_vals[idx] if idx < len(stamp_vals) and stamp_vals[idx] else ""

                thickness = None
                height = None
                width = None
                density = None
                density_name = None
                line_setting = scrap if idx == 0 else 0
                sidepatti = None 


            #  VALIDATION (VERY IMPORTANT)
            if quantity <= 0:
                continue  # Skip productions with zero or negative quantity

            if category == "Door" and not (thickness and width and height and density and color):
                continue

            if category == "Frame" and not (size and length and color):
                continue

            #  CREATE PRODUCT BASED ON CATEGORY
            if category == "Door":
                product, _ = Product.objects.get_or_create(
                    category=category,
                    color=color,
                    thickness=thickness,
                    density=density,
                    density_name=density_name,
                    width=width,
                    height=height,
                    stamp=stamp,
                    defaults={
                        "size": None,
                        "length": None,
                    }
                )

            elif category == "Frame":
                product, _ = Product.objects.get_or_create(
                    category=category,
                    color=color,
                    size=size,
                    length=length,
                    stamp=stamp,
                    defaults={
                        "thickness": None,
                        "density": None,
                        "width": None,
                        "height": None,
                    }
                )

            Production.objects.create(
                product=product,
                category=category,
                machine=machine,
                operator=operator,
                size=size,
                length=length,
                color=color,
                thickness=thickness,
                #  NEW
                height=height,
                width=width,
                density=density,
                density_name=density_name,
                linesetting=line_setting if idx == 0 else 0,
                sidepatti=sidepatti if category == "Door" and idx == 0 else 0,

                quantity=quantity,
                weight_per_piece=weight_per_piece,
                shift=shift,
                status="Pending",
                created_at=date,
                remark=remark
            )

        audit_logger.info(f"Production entry added: {category} - Qty: {quantities} by user {request.user.username}")
        business_logger.info(f"New {category} production created - Machine: {machine.name}, Shift: {shift}, Operator: {operator.username}")
        #  Dynamic success message + redirect
        messages.success(request, f"{category} production added successfully.")
        return redirect(f"/production/?category={category}")

    # ---- Fetch Productions ----
    # Only show pending productions (hide approved ones)
    productions = Production.objects.filter(
        category=category,
        status="Pending"
    ).select_related(
        "operator", "product", "machine", "size", "color", "length",
        "thickness", "height", "width", "density"
    ).order_by("id")
    
    # Filter for operators
    if request.user.role == 'Operator':
        productions = productions.filter(operator=request.user)
    
    # ---- Apply Date Filter ----
    filter_date_str = request.GET.get("filter_date")
    filter_date = None

    if filter_date_str:
        try:
            filter_date = timezone.datetime.strptime(filter_date_str, "%Y-%m-%d").date()
            productions = productions.filter(created_at__date=filter_date)
        except ValueError:
            filter_date = None  # important

    # ---- Apply Machine Filter ----
    filter_machine_id = request.GET.get("filter_machine")
    filter_machine = None
    if filter_machine_id:
        try:
            filter_machine = Machine.objects.get(id=filter_machine_id)
            productions = productions.filter(machine=filter_machine)
        except (Machine.DoesNotExist, ValueError):
            filter_machine = None

    # ---- Apply Shift Filter ----
    filter_shift = request.GET.get("filter_shift", "")
    if filter_shift:
        productions = productions.filter(shift=filter_shift)

    # ---- Production report date filter ----
    report_date = request.GET.get("report_date", "")
    report_rows = []
    if report_date:
        try:
            report_date_obj = timezone.datetime.strptime(report_date, "%Y-%m-%d").date()
            report_qs = Production.objects.filter(created_at__date=report_date_obj)
            if request.user.role == 'Operator':
                report_qs = report_qs.filter(operator=request.user)

            # Group by machine, category, shift
            report_map = {}
            for p in report_qs:
                key = (p.machine.name if p.machine else "-", p.product.category, p.shift)
                row = report_map.setdefault(key, {
                    "machine": key[0],
                    "category": key[1],
                    "shift": key[2],
                    "total_production": 0,
                    "actual_production": 0,
                    "line_setting": 0,
                    "side_patti": 0,
                    "qty": 0,
                    "sizes": set(),
                    "items": set(),
                })
                row["total_production"] += p.total_weight
                row["actual_production"] += p.actual_production
                row["line_setting"] += p.linesetting
                row["side_patti"] += p.sidepatti or 0
                row["qty"] += p.quantity
                if p.size:
                    row["sizes"].add(str(p.size))
                if p.product:
                    row["items"].add(str(p.product))

            report_rows = [
                {
                    **row,
                    "sizes": ", ".join(sorted(row["sizes"])),
                    "items": ", ".join(sorted(row["items"]))
                }
                for row in report_map.values()
            ]
        except ValueError:
            report_date = ""

    # Pagination
    paginator = Paginator(productions, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "productions": page_obj,
        # "machines": Machine.objects.all(),
        "machines": Machine.objects.filter(category=category, is_active=True),
        "colors": Color.objects.filter(is_active=True),
        "sizes": Size.objects.filter(is_active=True),
        "lengths": Length.objects.filter(is_active=True).order_by('length'),
        "today": timezone.now().date(),
        "filter_date": filter_date,
        "filter_machine": filter_machine,
        "filter_shift": filter_shift,
        "report_date": report_date,
        "report_rows": report_rows,

        "thicknesses": Thickness.objects.filter(is_active=True, thickness__lte=36).order_by('thickness'),
        "heights": Height.objects.filter(is_active=True, height__gte=48).order_by('height'),
        "widths": Width.objects.filter(is_active=True, width__lte=49).order_by('width'),
        "densities": Density.objects.filter(is_active=True).order_by('density'),
        "density_names": DensityName.objects.all(),
        # Dropdown data
        "shifts": [choice[0] for choice in Production.SHIFT_CHOICES],
        "categories": [choice[0] for choice in Product.CATEGORY_CHOICES],
        "operators": User.objects.filter(role="Operator").order_by("username"),

        #  Pass category to template
        "current_category": category,
    }

    #  Dynamic template rendering
    if category == "Door":
        return render(request, "door_production.html", context)
    else:
        return render(request, "production.html", context)
    



@login_required
@module_required('production')
def edit_production(request, production_id):
    User = get_user_model()
    try:
        prod = Production.objects.get(id=production_id)
    except Production.DoesNotExist:
        messages.error(request, "Production not found. It may have been deleted or the ID is invalid.")
        return redirect('production')
    
    # Check if operator can edit this production
    if request.user.role == 'Operator' and prod.operator != request.user:
        messages.error(request, "You can only edit your own productions.")
        return redirect('production')
    
    # Find all productions in the same "batch" (same date, machine, shift, operator)
    related_productions = Production.objects.filter(
        created_at__date=prod.created_at.date(),
        machine=prod.machine,
        shift=prod.shift,
        operator=prod.operator,
        category=prod.category
    ).order_by('id')
    
    if request.method == "POST":
        # Get common fields
        date_str = request.POST.get("date")
        date = timezone.datetime.strptime(date_str, "%Y-%m-%d") if date_str else prod.created_at
        machine_id = request.POST.get("machine")
        shift = request.POST.get("shift", prod.shift)
        operator_id = request.POST.get("operator")
        
        machine = get_object_or_404(Machine, id=machine_id)
        operator = User.objects.get(id=operator_id) if operator_id else request.user
        category = request.POST.get("category") or prod.category
        
        # Get all row data
        if category == "Frame":
            sizes = request.POST.getlist("size")
            colors = request.POST.getlist("color")
            lengths = request.POST.getlist("length")
            quantities = request.POST.getlist("quantity")
            weights = request.POST.getlist("weight")
            scrap = float(request.POST.get("scrap", 0))
            rejections = request.POST.getlist("rejection")
            scrap_qtys = request.POST.getlist("scrap_qty")
            stamp_vals = request.POST.getlist("stamp")
            
            num_rows = len(sizes)
            productions_list = list(related_productions)
            remark = request.POST.get("remark", "")
            
            # Update common fields for all existing
            for p in productions_list:
                p.created_at = date
                p.machine = machine
                p.shift = shift
                p.operator = operator
                p.category = category
                p.remark = remark
                p.edited = True
                p.save()
            
            # Update or create productions for each row
            for idx in range(num_rows):
                if idx < len(productions_list):
                    prod = productions_list[idx]
                else:
                    prod = Production(
                        machine=machine,
                        operator=operator,
                        shift=shift,
                        category=category,
                        created_at=date,
                        status="Pending",
                        edited=True,
                        remark=remark
                    )
                
                size = Size.objects.get(id=sizes[idx]) if sizes[idx] else None
                color = Color.objects.get(id=colors[idx]) if colors[idx] else None
                
                length_obj = None
                length_unit = request.POST.getlist('length_unit')[idx] if idx < len(request.POST.getlist('length_unit')) else 'ft'
                length_mm = request.POST.getlist('length_mm')[idx] if idx < len(request.POST.getlist('length_mm')) else ''
                
                if length_unit == 'mm' and length_mm:
                    try:
                        mm_value = float(length_mm)
                        ft_value = mm_value / 304.8
                        length_obj, _ = Length.objects.get_or_create(length=ft_value, defaults={'unit': 'mm', 'original_value': mm_value})
                    except ValueError:
                        length_obj = None
                else:
                    length_obj = Length.objects.get(id=lengths[idx]) if lengths[idx] else None
                
                quantity = int(quantities[idx]) if quantities[idx] else 0
                weight_per_piece = float(weights[idx]) if weights[idx] else 0
                
                # 🔥 REJECTION HANDLING FOR FRAME
                rejection = rejections[idx] if idx < len(rejections) else "ok"
                scrap_qty = float(scrap_qtys[idx]) if idx < len(scrap_qtys) and scrap_qtys[idx] else 0
                
                product, _ = Product.objects.get_or_create(
                    category=category,
                    size=size,
                    color=color,
                    length=length_obj,
                    stamp=stamp_vals[idx] if idx < len(stamp_vals) and stamp_vals[idx] else "",
                    defaults={
                        "thickness": None,
                        "density": None,
                        "width": None,
                        "height": None,
                    }
                )
                
                prod.product = product
                prod.size = size
                prod.color = color
                prod.length = length_obj
                prod.quantity = quantity
                prod.weight_per_piece = weight_per_piece
                prod.linesetting = scrap if idx == 0 else 0
                prod.sidepatti = 0
                
                # 🔥 STORE REJECTION STATUS
                if rejection == "ok":
                    prod.rejection_status = "OK"
                    prod.rejected_quantity = 0
                elif rejection == "scrap" and scrap_qty > 0:
                    prod.rejection_status = "Scrap"
                    prod.rejected_quantity = scrap_qty
                    # ADD SCRAP WEIGHT TO LINE SETTING
                    scrap_weight = scrap_qty * weight_per_piece
                    prod.linesetting += scrap_weight
                else:
                    prod.rejection_status = "OK"
                    prod.rejected_quantity = 0
                    
                prod.save()
                
                # 🔥 SCRAP LOG
                if rejection == "scrap" and scrap_qty > 0:
                    scrap_weight = scrap_qty * weight_per_piece
                    ScrapLog.objects.create(
                        production=prod,
                        quantity=scrap_qty,
                        weight=scrap_weight,
                        date=date.date() if hasattr(date, 'date') else date
                    )
            
            # Delete extra productions if fewer rows
            if len(productions_list) > num_rows:
                for extra in productions_list[num_rows:]:
                    extra.delete()
                
        elif category == "Door":
            thicknesses = request.POST.getlist("thickness")
            heights = request.POST.getlist("height")
            widths = request.POST.getlist("width")
            densities = request.POST.getlist("density")
            colors = request.POST.getlist("color")
            quantities = request.POST.getlist("quantity")
            weights = request.POST.getlist("weight")
            sidepattis = request.POST.getlist("sidepatti")
            line_settings = request.POST.getlist("linesetting")
            rejections = request.POST.getlist("rejection")
            unfinished_qtys = request.POST.getlist("unfinished_qty")
            scrap_qtys = request.POST.getlist("scrap_qty")
            stamp_vals = request.POST.getlist("stamp")
            # Single header-level stamp applies to all rows (matching add production behavior)
            stamp_value = stamp_vals[0] if stamp_vals and stamp_vals[0] else ""
            
            num_rows = len(thicknesses)
            productions_list = list(related_productions)
            remark = request.POST.get("remark", "")
            
            # Update common fields for all existing
            for p in productions_list:
                p.created_at = date
                p.machine = machine
                p.shift = shift
                p.operator = operator
                p.category = category
                p.remark = remark
                p.edited = True
                p.save()
            
            # Update or create productions for each row
            for idx in range(num_rows):
                if idx < len(productions_list):
                    prod = productions_list[idx]
                else:
                    prod = Production(
                        machine=machine,
                        operator=operator,
                        shift=shift,
                        category=category,
                        created_at=date,
                        status="Pending",
                        edited=True,
                        remark=remark
                    )
                
                thickness = Thickness.objects.get(id=thicknesses[idx]) if thicknesses[idx] else None
                density = Density.objects.get(id=densities[idx]) if densities[idx] else None
                color = Color.objects.get(id=colors[idx]) if colors[idx] else None
                
                height_obj = None
                height_unit = request.POST.getlist('height_unit')[idx] if idx < len(request.POST.getlist('height_unit')) else 'inch'
                height_mm = request.POST.getlist('height_mm')[idx] if idx < len(request.POST.getlist('height_mm')) else ''
                
                if height_unit == 'mm' and height_mm:
                    try:
                        mm_value = float(height_mm)
                        inch_value = mm_value / 25.4
                        height_obj, _ = Height.objects.get_or_create(height=inch_value, defaults={'unit': 'mm', 'original_value': mm_value})
                    except ValueError:
                        height_obj = None
                else:
                    height_obj = Height.objects.get(id=heights[idx]) if heights[idx] else None
                
                width_obj = None
                width_unit = request.POST.getlist('width_unit')[idx] if idx < len(request.POST.getlist('width_unit')) else 'inch'
                width_mm = request.POST.getlist('width_mm')[idx] if idx < len(request.POST.getlist('width_mm')) else ''
                
                if width_unit == 'mm' and width_mm:
                    try:
                        mm_value = float(width_mm)
                        inch_value = mm_value / 25.4
                        width_obj, _ = Width.objects.get_or_create(width=inch_value, defaults={'unit': 'mm', 'original_value': mm_value})
                    except ValueError:
                        width_obj = None
                else:
                    width_obj = Width.objects.get(id=widths[idx]) if widths[idx] else None
                
                quantity = int(quantities[idx]) if quantities[idx] else 0
                weight_per_piece = float(weights[idx]) if weights[idx] else 0
               
                line_setting = float(line_settings[idx]) if idx < len(line_settings) and line_settings[idx] else 0
                sidepatti = float(sidepattis[idx]) if idx < len(sidepattis) and sidepattis[idx] else 0
                
                # 🔥 REJECTION HANDLING
                rejection = rejections[idx] if idx < len(rejections) else "ok"
                unfinished_qty = float(unfinished_qtys[idx]) if idx < len(unfinished_qtys) and unfinished_qtys[idx] else 0
                scrap_qty = float(scrap_qtys[idx]) if idx < len(scrap_qtys) and scrap_qtys[idx] else 0
                
                product, _ = Product.objects.get_or_create(
                    category=category,
                    color=color,
                    thickness=thickness,
                    density=density,
                    width=width_obj,
                    height=height_obj,
                    stamp=stamp_value,
                    defaults={
                        "size": None,
                        "length": None,
                    }
                )
                
                prod.product = product
                prod.thickness = thickness
                prod.height = height_obj
                prod.width = width_obj
                prod.density = density
                prod.color = color
                prod.quantity = quantity
                prod.weight_per_piece = weight_per_piece
                prod.sidepatti = sidepatti if idx == 0 else 0
                prod.linesetting = line_setting if idx == 0 else 0
                
                # 🔥 STORE REJECTION STATUS
                if rejection == "ok":
                    prod.rejection_status = "OK"
                    prod.rejected_quantity = 0
                elif rejection == "unfinished" and unfinished_qty > 0:
                    prod.rejection_status = "Unfinished"
                    prod.rejected_quantity = unfinished_qty
                elif rejection == "scrap" and scrap_qty > 0:
                    prod.rejection_status = "Scrap"
                    prod.rejected_quantity = scrap_qty
                else:
                    prod.rejection_status = "OK"
                    prod.rejected_quantity = 0
                    
                prod.save()
                
                # 🔥 CREATE UNFINISHED PRODUCTION LOG
                if rejection == "unfinished" and unfinished_qty > 0:
                    UnfinishedProduction.objects.create(
                        product=prod.product,
                        production=prod,
                        quantity=unfinished_qty,
                        date=date.date() if hasattr(date, 'date') else date
                    )
                
                # 🔥 SCRAP LOG + ADD TO LINE SETTING
                if rejection == "scrap" and scrap_qty > 0:
                    scrap_weight = scrap_qty * weight_per_piece
                    prod.linesetting += scrap_weight
                    prod.save()
                    
                    ScrapLog.objects.create(
                        production=prod,
                        quantity=scrap_qty,
                        weight=scrap_weight,
                        date=date.date() if hasattr(date, 'date') else date
                    )
            
            # Delete extra productions if fewer rows
            if len(productions_list) > num_rows:
                for extra in productions_list[num_rows:]:
                    extra.delete()
        
        messages.success(request, f"{category} production batch updated successfully.")
        return redirect(f"/production/?category={category}")
    
    # For GET request, prepare data for editing
    context = {
        'production_batch': related_productions,
        'machines': Machine.objects.all(),
        'shifts': [choice[0] for choice in Production.SHIFT_CHOICES],
        'categories': [choice[0] for choice in Product.CATEGORY_CHOICES],
        'operators': User.objects.filter(role="Operator").order_by("username"),
        'today': prod.created_at.date(),
        'current_category': prod.category,
    }
    
    # Add category-specific data
    if prod.category == "Frame":
        context.update({
            'sizes': Size.objects.filter(is_active=True),
            'colors': Color.objects.filter(is_active=True),
            'lengths': Length.objects.filter(is_active=True).order_by('length'),
        })
    else:  # Door
        sidepatti_value = sum(p.sidepatti for p in related_productions)
        linesetting_value = sum(p.linesetting for p in related_productions)

        context.update({
            'thicknesses': Thickness.objects.filter(is_active=True, thickness__lte=36).order_by('thickness'),
            'heights': Height.objects.filter(Q(is_active=True) | Q(id=prod.height_id)).filter(height__gte=48).order_by('height'),
            'widths': Width.objects.filter(Q(is_active=True) | Q(id=prod.width_id)).filter(width__lte=49).order_by('width'),
            'densities': Density.objects.filter(is_active=True),
            'density_names': DensityName.objects.all(),
            'colors': Color.objects.filter(is_active=True),
            'sidepatti_value': sidepatti_value,
            'linesetting_value': linesetting_value,
        })
    
    # Dynamic template rendering
    if prod.category == "Door":
        return render(request, "edit_door_production.html", context)
    else:
        return render(request, "edit_production.html", context)
    


@login_required
@admin_only
def approve_production(request, production_id):
    prod = get_object_or_404(Production, id=production_id)

    #  Prevent double approval
    if prod.status == "Approved":
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Already approved.'})
        messages.warning(request, "Already approved.")
        return redirect("production")

    try:
        #  Add to Stock
        Stock.objects.create(
            product=prod.product,
            quantity=prod.quantity,
            movement_type='IN',
            operator=prod.operator,
            production=prod   #  link to production (important)
        )

        #  Update Production
        prod.status = "Approved"
        prod.approved_at = timezone.now()
        prod.save()

        audit_logger.info(f"Production {prod.id} approved by user {request.user.username} - Product: {prod.product.product_name}, Qty: {prod.quantity}")
        business_logger.info(f"Production {prod.id} added to stock - Product: {prod.product.product_name}, Qty: {prod.quantity}")
        messages.success(request, f"Production {prod.id} approved and added to stock.")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'message': f'Production {prod.id} approved and added to stock.'})
    except Exception as e:
        errors_logger.error(f"Error approving production {prod.id}: {str(e)}", exc_info=True)
        messages.error(request, f"Error approving production: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': f'Error approving production: {str(e)}'})
        return redirect(f"/production/?category={prod.category}")
    return redirect(f"/production/?category={prod.category}")



@module_required('master')
def machine_list(request):
    machines = Machine.objects.all().order_by('id')
    paginator = Paginator(machines, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    form = MachineForm()

    if request.method == "POST":
        machine_id = request.POST.get('machine_id')

        if machine_id:  # EDIT
            machine = get_object_or_404(Machine, id=machine_id)
            form = MachineForm(request.POST, instance=machine)
        else:  # ADD
            form = MachineForm(request.POST)

        if form.is_valid():
            try:
                form.save()
                return redirect('machine_list')
            except IntegrityError:
                messages.error(request, f"Machine '{form.cleaned_data['name']}' already exists.")

    return render(request, 'machine.html', {
        'page_obj': page_obj,
        'form': form
    })


@login_required
@module_required('master')
def delete_machine(request, pk):
    machine = get_object_or_404(Machine, pk=pk)
    machine.delete()
    return redirect('machine_list')


@login_required
def get_operators_for_machine(request):
    """API endpoint to get operators for a selected machine (category-based)
    
    Admin users see all operators. Other users see only operators assigned 
    to the selected machine's category via MachineOperator.
    """

    machine_id = request.GET.get('machine_id')
    if not machine_id:
        return JsonResponse({'operators': []})

    try:
        machine = Machine.objects.get(id=machine_id)

        if request.user.role in ['Admin', 'Manager']:
            # Admin and Manager see all operators
            operators_qs = CustomUser.objects.filter(role='Operator').order_by('username')
            operators = [
                {'id': op.id, 'full_name': op.full_name}
                for op in operators_qs
            ]
        else:
            # Other users see only category-assigned operators
            operator_assignments = MachineOperator.objects.filter(category=machine.category).select_related('operator')
            operators = [
                {'id': mo.operator.id, 'full_name': mo.operator.full_name}
                for mo in operator_assignments
            ]

        return JsonResponse({'operators': operators})
    except Machine.DoesNotExist:
        return JsonResponse({'operators': []})


@module_required('master')
def machine_operator_list(request):
    """Manage operator-category assignments"""

    assignments = MachineOperator.objects.all().order_by('category', 'operator__username')
    categories = Machine.CATEGORY_CHOICES
    operators = CustomUser.objects.filter(role='Operator').order_by('username')

    # Handle form submission
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add":
            category_selected = request.POST.get("category")
            operator_id = request.POST.get("operator")

            if category_selected and operator_id:
                operator = get_object_or_404(CustomUser, id=operator_id)

                # Check if already assigned for this category
                if not MachineOperator.objects.filter(category=category_selected, operator=operator).exists():
                    MachineOperator.objects.create(category=category_selected, operator=operator)
                    messages.success(request, f"✓ {operator.full_name} assigned to {category_selected} machines")
                else:
                    messages.warning(request, f"Already assigned: {operator.full_name} → {category_selected} machines")

        elif action == "delete":
            assignment_id = request.POST.get("assignment_id")
            assignment = get_object_or_404(MachineOperator, id=assignment_id)
            category_name = assignment.category
            operator_name = assignment.operator.full_name
            assignment.delete()
            messages.success(request, f"✓ {operator_name} removed from {category_name} machines")

        return redirect("machine_operator_list")

    paginator = Paginator(assignments, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'machine_operator.html', {
        'page_obj': page_obj,
        'categories': categories,
        'operators': operators,
        'total_assignments': assignments.count(),
    })

@login_required
@module_required('master')
def color_list(request):
    colors = Color.objects.all().order_by('id')

    paginator = Paginator(colors, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        value = request.POST.get("value")

        if obj_id:
            obj = get_object_or_404(Color, id=obj_id)
            obj.color = value
            try:
                obj.save()
            except ValidationError as e:
                if hasattr(e, 'message_dict') and 'color' in e.message_dict:
                    messages.error(request, f"Color '{value}' already exists.")
                else:
                    messages.error(request, str(e))
        else:
            try:
                Color.objects.create(color=value)
            except (ValidationError, IntegrityError):
                messages.error(request, f"Color '{value}' already exists.")

        return redirect("color_list")

    return render(request, "master.html", {
        "page_obj": page_obj,
        "title": "Color",
        "field": "color",
        "delete_url": "delete_color"
    })


@module_required('master')
def delete_color(request, pk):
    obj = get_object_or_404(Color, pk=pk)
    obj.delete()
    return redirect("color_list")

@module_required('master')
def density_list(request):
    items = Density.objects.all().order_by('id')

    paginator = Paginator(items, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        value = request.POST.get("value")

        if obj_id:
            obj = get_object_or_404(Density, id=obj_id)
            obj.density = value
            try:
                obj.save()
            except ValidationError as e:
                if hasattr(e, 'message_dict') and 'density' in e.message_dict:
                    messages.error(request, f"Density '{value}' already exists.")
                else:
                    messages.error(request, str(e))
        else:
            try:
                Density.objects.create(density=value)
            except (ValidationError, IntegrityError):
                messages.error(request, f"Density '{value}' already exists.")

        return redirect("density_list")

    return render(request, "master.html", {
        "page_obj": page_obj,
        "title": "Density",
        "delete_url": "delete_density"
    })


@module_required('master')
def delete_density(request, pk):
    obj = get_object_or_404(Density, pk=pk)
    obj.delete()
    return redirect("density_list")


@module_required('master')
def density_name_list(request):
    items = DensityName.objects.all().order_by('name')

    paginator = Paginator(items, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        value = request.POST.get("value")

        if obj_id:
            obj = get_object_or_404(DensityName, id=obj_id)
            obj.name = value
            try:
                obj.save()
            except ValidationError as e:
                if hasattr(e, 'message_dict') and 'name' in e.message_dict:
                    messages.error(request, f"Density '{value}' already exists.")
                else:
                    messages.error(request, str(e))
        else:
            try:
                DensityName.objects.create(name=value)
            except (ValidationError, IntegrityError):
                messages.error(request, f"Density '{value}' already exists.")

        return redirect("density_name_list")

    return render(request, "master.html", {
        "page_obj": page_obj,
        "title": "Density Name",
        "delete_url": "delete_density_name"
    })


@module_required('master')
def delete_density_name(request, pk):
    obj = get_object_or_404(DensityName, pk=pk)
    obj.delete()
    return redirect("density_name_list")



@module_required('master')
def thickness_list(request):
    items = Thickness.objects.all().order_by('thickness')

    paginator = Paginator(items, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        value = request.POST.get("value")

        try:
            thickness_float = float(value)
        except (TypeError, ValueError):
            messages.error(request, "Invalid number. Thickness must be numeric.")
            return redirect("thickness_list")

        if thickness_float > 36:
            messages.error(request, "Invalid number. Thickness cannot be greater than 36.")
            return redirect("thickness_list")

        if obj_id:
            obj = get_object_or_404(Thickness, id=obj_id)
            obj.thickness = thickness_float
            try:
                obj.save()
            except ValidationError as e:
                if hasattr(e, 'message_dict') and 'thickness' in e.message_dict:
                    messages.error(request, f"Thickness '{value}' already exists.")
                else:
                    messages.error(request, str(e))
        else:
            try:
                Thickness.objects.create(thickness=thickness_float)
            except (ValidationError, IntegrityError):
                messages.error(request, f"Thickness '{value}' already exists.")

        return redirect("thickness_list")

    return render(request, "master.html", {
        "page_obj": page_obj,
        "title": "Thickness",
        "delete_url": "delete_thickness"
    })


@module_required('master')
def delete_thickness(request, pk):
    obj = get_object_or_404(Thickness, pk=pk)
    obj.delete()
    return redirect("thickness_list")

@module_required('master')
def size_list(request):
    items = Size.objects.all().order_by('id')

    paginator = Paginator(items, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        value = request.POST.get("value")
        standard_weight = request.POST.get("standard_weight")
        rate = request.POST.get("rate")

        if obj_id:
            obj = get_object_or_404(Size, id=obj_id)
            obj.size = value
            obj.standard_weight = standard_weight
            obj.rate = rate if rate else None
            try:
                obj.save()
            except ValidationError as e:
                if hasattr(e, 'message_dict') and 'size' in e.message_dict:
                    messages.error(request, f"Size '{value}' already exists.")
                else:
                    messages.error(request, str(e))
        else:
            try:
                Size.objects.create(size=value, standard_weight=standard_weight, rate=rate if rate else None)
            except (ValidationError, IntegrityError):
                messages.error(request, f"Size '{value}' already exists.")

        return redirect("size_list")

    return render(request, "master.html", {
        "page_obj": page_obj,
        "title": "Size",
        "delete_url": "delete_size"
    })


@module_required('master')
def delete_size(request, pk):
    obj = get_object_or_404(Size, pk=pk)
    obj.delete()
    return redirect("size_list")


@module_required('master')
def party_list(request):
    items = Party.objects.all().order_by('id')

    paginator = Paginator(items, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        c_name = request.POST.get("c_name")
        address = request.POST.get("address")
        gst = request.POST.get("gst")
        state = request.POST.get("state")
        psft = request.POST.get("psft")

        if obj_id:
            obj = get_object_or_404(Party, id=obj_id)
            obj.c_name = c_name
            obj.address = address
            obj.gst = gst
            obj.state = state
            obj.psft = psft
            try:
                obj.save()
            except (ValidationError, IntegrityError):
                messages.error(request, f"Party '{c_name}' already exists.")
        else:
            try:
                Party.objects.create(c_name=c_name, address=address, gst=gst, state=state, psft=psft)
            except (ValidationError, IntegrityError):
                messages.error(request, f"Party '{c_name}' already exists.")

        return redirect("party_list")

    return render(request, "master.html", {
        "page_obj": page_obj,
        "title": "Party",
        "delete_url": "delete_party"
    })


@module_required('master')
def delete_party(request, pk):
    obj = get_object_or_404(Party, pk=pk)
    obj.delete()
    return redirect("party_list")


@module_required('master')

def length_list(request):
    items = Length.objects.all().order_by('id')

    paginator = Paginator(items, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        value = request.POST.get("value")
        unit = request.POST.get("unit", "ft")

        if not value:
            return redirect("length_list")

        try:
            value_float = float(value)
        except ValueError:
            return redirect("length_list")

        if unit == 'mm':
            length_ft = value_float / 304.8
        else:
            unit = 'ft'
            length_ft = value_float

        if obj_id:
            obj = get_object_or_404(Length, id=obj_id)
            obj.length = length_ft
            obj.unit = unit
            obj.original_value = value_float
            try:
                obj.save()
            except ValidationError as e:
                if hasattr(e, 'message_dict') and 'length' in e.message_dict:
                    messages.error(request, f"Length '{value} {unit}' already exists.")
                else:
                    messages.error(request, str(e))
        else:
            try:
                Length.objects.create(length=length_ft, unit=unit, original_value=value_float)
            except (ValidationError, IntegrityError):
                messages.error(request, f"Length '{value} {unit}' already exists.")

        return redirect("length_list")

    return render(request, "master.html", {
        "page_obj": page_obj,
        "title": "Length",
        "delete_url": "delete_length"
    })


@module_required('master')
def delete_length(request, pk):
    obj = get_object_or_404(Length, pk=pk)
    obj.delete()
    return redirect("length_list")


@module_required('master')
def height_list(request):
    items = Height.objects.all().order_by('id')

    paginator = Paginator(items, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        value = request.POST.get("value")
        unit = request.POST.get("unit", "inch")

        if not value:
            return redirect("height_list")

        try:
            value_float = float(value)
        except ValueError:
            messages.error(request, "Invalid number. Height must be numeric.")
            return redirect("height_list")

        if unit == 'mm':
            height_inch = value_float / 25.4
        else:
            unit = 'inch'
            height_inch = value_float

        if height_inch < 48:
            messages.error(request, "Invalid number. Height cannot be less than 48 inches.")
            return redirect("height_list")

        if obj_id:
            obj = get_object_or_404(Height, id=obj_id)
            obj.height = height_inch
            obj.unit = unit
            obj.original_value = value_float
            try:
                obj.save()
            except ValidationError as e:
                if hasattr(e, 'message_dict') and 'height' in e.message_dict:
                    messages.error(request, f"Height '{value} {unit}' already exists.")
                else:
                    messages.error(request, str(e))
            except IntegrityError:
                messages.error(request, f"Height '{value} {unit}' already exists.")
        else:
            try:
                Height.objects.create(height=height_inch, unit=unit, original_value=value_float)
            except ValidationError as e:
                if hasattr(e, 'message_dict') and 'height' in e.message_dict:
                    messages.error(request, f"Height '{value} {unit}' already exists.")
                else:
                    messages.error(request, str(e))
            except IntegrityError:
                messages.error(request, f"Height '{value} {unit}' already exists.")

        return redirect("height_list")

    return render(request, "master.html", {
        "page_obj": page_obj,
        "title": "Height",
        "delete_url": "delete_height"
    })


@module_required('master')
def delete_height(request, pk):
    obj = get_object_or_404(Height, pk=pk)
    obj.delete()
    return redirect("height_list")

@module_required('master')
def width_list(request):
    items = Width.objects.all().order_by('id')

    paginator = Paginator(items, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == "POST":
        obj_id = request.POST.get("obj_id")
        value = request.POST.get("value")
        unit = request.POST.get("unit", "inch")

        if not value:
            return redirect("width_list")

        try:
            value_float = float(value)
        except ValueError:
            messages.error(request, "Invalid number. Width must be numeric.")
            return redirect("width_list")

        if unit == 'mm':
            width_inch = value_float / 25.4
        else:
            unit = 'inch'
            width_inch = value_float

        if width_inch > 49:
            messages.error(request, "Invalid number. Width cannot be greater than 49 inches.")
            return redirect("width_list")

        if obj_id:
            obj = get_object_or_404(Width, id=obj_id)
            obj.width = width_inch
            obj.unit = unit
            obj.original_value = value_float
            try:
                obj.save()
            except ValidationError as e:
                if hasattr(e, 'message_dict') and 'width' in e.message_dict:
                    messages.error(request, f"Width '{value} {unit}' already exists.")
                else:
                    messages.error(request, str(e))
        else:
            try:
                Width.objects.create(width=width_inch, unit=unit, original_value=value_float)
            except (ValidationError, IntegrityError):
                messages.error(request, f"Width '{value} {unit}' already exists.")

        return redirect("width_list")

    return render(request, "master.html", {
        "page_obj": page_obj,
        "title": "Width",
        "delete_url": "delete_width"
    })


@module_required('master')
def delete_width(request, pk):
    obj = get_object_or_404(Width, pk=pk)
    obj.delete()
    return redirect("width_list")



@module_required('reports')
def reports(request):
    User = get_user_model()

    report_type = request.GET.get("type", "production")
    mode = request.GET.get("mode")
    if report_type == "production" and not mode:
        mode = "by_operator"

    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    machines_filter = request.GET.getlist("machine")  # for "by_machine"
    shift_filter = request.GET.get("shift")           # optional

    context = {
        "report_type": report_type,
        "mode": mode,
        "machines": Machine.objects.all(),
        "from_date": from_date or "",
        "to_date": to_date or "",
        "summary": [],
        "productions": [],
    }

    if report_type == "production":
        data = Production.objects.all()
        if request.user.role == 'Operator':
            data = data.filter(operator=request.user)

        if mode == "by_date":
            date_only = request.GET.get("date")  # this is the single date input from template
            if date_only:
                data = data.filter(created_at__date=date_only)
            else:
                data = data.none()  # if no date selected yet, show empty

        # ----------------- OTHER MODES -----------------
        else:

            if from_date and to_date:
                data = data.filter(created_at__date__range=[from_date, to_date])
            if machines_filter:
                data = data.filter(machine_id__in=machines_filter)
            if shift_filter:
                data = data.filter(shift=shift_filter)

        # These are @property methods on Production, not DB columns — annotate for .values() / JSON
        data = data.annotate(total_weight=F("quantity") * F("weight_per_piece"))
        data = data.annotate(actual_production=F("total_weight"))
        data = data.annotate(
            overall_production=F("total_weight") + F("sidepatti") + F("linesetting")
        )

        prod_values = list(data.values(
            "id",
            "created_at",
            "operator__id",
            "operator__full_name",
            "machine__id",
            "machine__name",
            "shift",
            "category",
            "product__product_name",
            "sidepatti",
            "linesetting",
            "quantity",
            "total_weight",
            "actual_production",
            "overall_production",
        ))
        for row in prod_values:
            ts = row.get("created_at")
            if ts is not None:
                row["created_at"] = ts.isoformat()

        context["productions"] = prod_values

        if mode == "by_machine":
            # Summary grouped by machine -> shift -> category
            summary = []
            machines = data.values("machine__id", "machine__name").distinct()
            for m in machines:
                machine_id = m["machine__id"]
                machine_name = m["machine__name"]

                shifts = data.filter(machine_id=machine_id).values("shift", "category").distinct()
                shifts_list = []
                for s in shifts:
                    shift_name = s["shift"]
                    category = s["category"]
                    agg = data.filter(machine_id=machine_id, shift=shift_name, category=category).aggregate(
                        total_qty=Sum("quantity"),
                        total_weight=Sum(F("quantity") * F("weight_per_piece"))
                    )
                    shifts_list.append({
                        "name": shift_name,
                        "category": category,
                        "total_qty": agg["total_qty"] or 0,
                        "total_weight": agg["total_weight"] or 0,
                    })
                summary.append({
                    "id": machine_id,
                    "name": machine_name,
                    "shifts": shifts_list
                })
            context["summary"] = summary

        elif mode == "by_operator":
            # Summary grouped by operator -> shift
            summary = []
            operators = data.values("operator__id", "operator__full_name").distinct()
            for op in operators:
                operator_id = op["operator__id"]
                operator_name = op["operator__full_name"]
                shifts = data.filter(operator_id=operator_id).values("shift").distinct()
                shifts_list = []
                for s in shifts:
                    shift_name = s["shift"]
                    agg = data.filter(operator_id=operator_id, shift=shift_name).aggregate(
                        total_qty=Sum("quantity"),
                        total_weight=Sum(F("quantity") * F("weight_per_piece"))
                    )
                    shifts_list.append({
                        "name": shift_name,
                        "total_qty": agg["total_qty"] or 0,
                        "total_weight": agg["total_weight"] or 0,
                    })
                summary.append({
                    "id": operator_id,
                    "name": operator_name,
                    "shifts": shifts_list
                })
            context["summary"] = summary

        elif mode == "by_date":
            summary = []

            shifts = data.values("shift").distinct()
            shifts_list = []

            for s in shifts:
                shift_name = s["shift"]
                agg = data.filter(shift=shift_name).aggregate(
                    total_qty=Sum("quantity"),
                    total_weight=Sum(F("quantity") * F("weight_per_piece"))
                )

                shifts_list.append({
                    "name": shift_name,
                    "total_qty": agg["total_qty"] or 0,
                    "total_weight": agg["total_weight"] or 0,
                })

            # Use selected date as single row
            date_only = request.GET.get("date")

            if date_only:
                summary.append({
                    "id": date_only,
                    "name": date_only,
                    "shifts": shifts_list
                })

            context["summary"] = summary

    elif report_type == "dispatch":
        from_date = request.GET.get("from_date")
        to_date = request.GET.get("to_date")
        
        dispatches = DeliveryChallan.objects.all()
        if from_date and to_date:
            dispatches = dispatches.filter(challan_date__range=[from_date, to_date])
        
        dispatches = dispatches.prefetch_related("items__product").order_by("-challan_date", "-id")
        
        # Annotate each challan with total quantity
        dispatches = dispatches.annotate(
            total_quantity=Sum('items__quantity')
        )
        
        context["dispatches"] = dispatches


    elif report_type == "consumption":
        from_date = request.GET.get("from_date")
        to_date = request.GET.get("to_date")

        data = []
        machine_cols = []

        if from_date and to_date:
            
            start_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(to_date, "%Y-%m-%d").date()

            # -----------------------------
            # MACHINES
            # -----------------------------
            all_machines = list(
                Machine.objects.filter(is_active=True).order_by("name")
            )

            machine_cols = []
            for m in all_machines:
                machine_cols.extend([
                    f"{m.name} Day",
                    f"{m.name} Night"
                ])

            # -----------------------------
            # MATERIALS USED IN RANGE ONLY (from BatchAddHistory via +Add button)
            # -----------------------------
            materials = RawMaterial.objects.filter(
                formulationbatch__batch_history__date__range=[start_date, end_date]
            ).distinct()

            material_map = {x.id: x for x in materials}

            # -----------------------------
            # RECEIVED STOCK (1 query)
            # -----------------------------
            received_rows = RawMaterialStock.objects.filter(
                movement_type="IN",
                created_at__date__range=[start_date, end_date],
                raw_material__in=materials
            ).values(
                "raw_material_id",
                "created_at__date"
            ).annotate(
                qty=Sum("quantity")
            )

            received_map = {}
            for r in received_rows:
                key = (r["raw_material_id"], str(r["created_at__date"]))
                received_map[key] = r["qty"] or 0

            # -----------------------------
            # CONSUMPTION (1 query only) - from BatchAddHistory via +Add button
            # -----------------------------
            usage_rows = FormulationBatch.objects.filter(
                raw_material__in=materials,
                batch_history__date__range=[start_date, end_date]
            ).values(
                "raw_material_id",
                "batch_history__date",
                "batch_history__machine__name",
                "batch_history__shift",
                "batch_history__formulation__is_loss",
            ).annotate(
                per_batch_weight=Sum(F("weight")),
                total_qty=Sum(F("weight") * F("batch_history__batch_count")),
                batches=Sum("batch_history__batch_count")
            )

            usage_map = {}

            for row in usage_rows:
                raw_material_id = row["raw_material_id"]
                date_key = str(row["batch_history__date"])
                is_loss = row["batch_history__formulation__is_loss"]

                if is_loss:
                    key = (raw_material_id, date_key, "Production Loss")
                else:
                    machine_name = row["batch_history__machine__name"]
                    shift = row["batch_history__shift"]

                    col = f"{machine_name} {shift}"

                    key = (raw_material_id, date_key, col)

                existing = usage_map.get(key, {"per_batch_weight": 0, "total_qty": 0, "batches": 0})
                usage_map[key] = {
                    "per_batch_weight": existing["per_batch_weight"] + (row["per_batch_weight"] or 0),
                    "total_qty": existing["total_qty"] + (row["total_qty"] or 0),
                    "batches": existing["batches"] + (row["batches"] or 0),
                }

            # -----------------------------
            # BUILD REPORT
            # -----------------------------
            current_date = start_date

            while current_date <= end_date:
                date_str = current_date.strftime("%Y-%m-%d")

                for rm in materials:
                    opening = rm.current_stock or 0
                    received = received_map.get((rm.id, date_str), 0)

                    loss_by_combo = {}
                    total_loss = 0

                    for col in machine_cols:
                        entry = usage_map.get((rm.id, date_str, col), {"per_batch_weight": 0, "total_qty": 0, "batches": 0})
                        loss_by_combo[col] = entry
                        total_loss += entry["total_qty"]

                    production_loss_entry = usage_map.get(
                        (rm.id, date_str, "Production Loss"),
                        {"per_batch_weight": 0, "total_qty": 0, "batches": 0}
                    )
                    total_loss += production_loss_entry["total_qty"]

                    # Skip zero rows
                    if received == 0 and total_loss == 0:
                        continue

                    total = received + total_loss
                    balance = opening + received - total_loss

                    row = {
                        "date": date_str,
                        "name": rm.name,
                        "opening": opening,
                        "received": received,
                        "production_loss_qty": production_loss_entry["per_batch_weight"],
                        "production_loss_batches": production_loss_entry["batches"],
                        "production_loss_total": production_loss_entry["total_qty"],
                        "total_loss": total_loss,
                        "total": total,
                        "balance": balance,
                        "total_consumed": total_loss,
                    }

                    for col in machine_cols:
                        entry = loss_by_combo[col]
                        row[f"{col}_qty"] = entry["per_batch_weight"]
                        row[f"{col}_batches"] = entry["batches"]
                        row[f"{col}_total"] = entry["total_qty"]

                    data.append(row)

                current_date += timedelta(days=1)

            # Calculate column totals for the footer row
            column_totals = {
                "opening": sum(row["opening"] for row in data),
                "received": sum(row["received"] for row in data),
                "production_loss_qty": sum(row["production_loss_qty"] for row in data),
                "production_loss_batches": sum(row["production_loss_batches"] for row in data),
                "production_loss_total": sum(row["production_loss_total"] for row in data),
                "total_consumed": sum(row["total_consumed"] for row in data),
                "total": sum(row["total"] for row in data),
                "balance": sum(row["balance"] for row in data),
            }

            # Add machine column totals
            for col in machine_cols:
                column_totals[f"{col}_qty"] = sum(row.get(f"{col}_qty", 0) for row in data)
                column_totals[f"{col}_batches"] = sum(row.get(f"{col}_batches", 0) for row in data)
                column_totals[f"{col}_total"] = sum(row.get(f"{col}_total", 0) for row in data)

            context["consumption_totals"] = {
                "total_loss": sum(row["total_loss"] for row in data),
                "total_consumption": sum(row["total"] for row in data),
            }

            context["column_totals"] = column_totals

        context["data"] = data
        context["machine_cols"] = machine_cols

    elif report_type == "consumption_vs_production":

        from_date = request.GET.get("from_date")
        to_date = request.GET.get("to_date")

        data = []

        if from_date and to_date:

            start_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(to_date, "%Y-%m-%d").date()

            # ==========================
            # TOTAL SUMMARY (WHOLE RANGE)
            # ==========================
            total_consumption = FormulationItem.objects.filter(
                formulation__status="Approved",
                formulation__date__range=[start_date, end_date]
            ).aggregate(
                total=Sum(F("weight") * F("formulation__batches"))
            )["total"] or 0

            total_production = Production.objects.filter(
                created_at__date__range=[start_date, end_date]
            ).aggregate(
                total=Sum(F("quantity") * F("weight_per_piece") + F("sidepatti"))
            )["total"] or 0

            total_production_loss = FormulationItem.objects.filter(
                formulation__is_loss=True,
                formulation__date__range=[start_date, end_date]
            ).aggregate(
                total=Sum(F("weight") * F("formulation__batches"))
            )["total"] or 0

            total_received = RawMaterialStock.objects.filter(
                movement_type="IN",
                created_at__date__range=[start_date, end_date]
            ).aggregate(
                total=Sum("quantity")
            )["total"] or 0

            difference = total_consumption - total_production - total_production_loss
            loss_percent = round(
                (difference / total_consumption) * 100, 2
            ) if total_consumption else 0

            # ==========================
            # DAILY BREAKDOWN (OPTIONAL)
            # ==========================
            current_date = start_date

            running_balance = 0

            while current_date <= end_date:

                cons = FormulationItem.objects.filter(
                    formulation__status="Approved",
                    formulation__date=current_date
                ).aggregate(
                    total=Sum(F("weight") * F("formulation__batches"))
                )["total"] or 0

                prod = Production.objects.filter(
                    created_at__date=current_date
                ).aggregate(
                    total=Sum(F("quantity") * F("weight_per_piece")+ F("sidepatti") )
                )["total"] or 0

                # pro_loss = Production.objects.filter(
                #     created_at_date=current_date
                # ).aggregate(total=Sum("is_loss")  # adjust field name
                # )["total"] or 0

                rec = RawMaterialStock.objects.filter(
                    movement_type="IN",
                    created_at__date=current_date
                ).aggregate(
                    total=Sum("quantity")
                )["total"] or 0

                daily_production_loss = FormulationItem.objects.filter(
                    formulation__is_loss=True,
                    formulation__date=current_date
                ).aggregate(
                    total=Sum(F("weight") * F("formulation__batches"))
                )["total"] or 0

                diff = cons - prod - daily_production_loss
                running_balance += (rec + prod - cons)

                data.append({
                    "date": current_date,
                    "consumption": cons,
                    "production": prod,
                    "received": rec,
                    "difference": diff,
                    "running_balance": running_balance,
                })

                current_date += timedelta(days=1)

            context["summary"] = {
                "total_consumption": total_consumption,
                "total_production": total_production,
                "total_production_loss": total_production_loss,
                "total_received": total_received,
                "difference": difference,
                "loss_percent": loss_percent,
            }

            context["consumption_vs_production"] = data

    elif report_type == "raw_material":
        supplier_filter = request.GET.get("supplier", "")
        receiving_from = request.GET.get("receiving_from", "")
        receiving_to = request.GET.get("receiving_to", "")

        data = RawMaterialStock.objects.filter(
            movement_type='IN',
            purchase_invoice__isnull=False
        ).select_related(
            'raw_material',
            'purchase_invoice',
            'purchase_invoice__supplier'
        ).order_by('-purchase_invoice__receiving_date')

        if supplier_filter:
            data = data.filter(purchase_invoice__supplier__name__icontains=supplier_filter)
        
        if receiving_from:
            data = data.filter(purchase_invoice__receiving_date__gte=receiving_from)
        
        if receiving_to:
            data = data.filter(purchase_invoice__receiving_date__lte=receiving_to)

        raw_material_stocks = []
        for stock in data:
            supplier = stock.purchase_invoice.supplier if stock.purchase_invoice else None
            supplier_name = supplier.name if supplier else "-"
            invoice_no = stock.purchase_invoice.invoice_no if stock.purchase_invoice else "-"
            raw_material_stocks.append({
                "id": stock.id,
                "group_key": f"{supplier_name}||{invoice_no}",
                "raw_material_name": stock.raw_material.name,
                "raw_material_category": stock.raw_material.category,
                "supplier_name": supplier_name,
                "supplier_contact": supplier.contact_person if supplier else "",
                "supplier_phone": supplier.phone if supplier else "",
                "invoice_no": invoice_no,
                "invoice_date": stock.purchase_invoice.invoice_date if stock.purchase_invoice else "-",
                "receiving_date": stock.purchase_invoice.receiving_date if stock.purchase_invoice else "-",
                "quantity": float(stock.quantity),
                "rate_per_kg": float(stock.rate_per_kg) if stock.rate_per_kg else 0,
                "transportation_rate": float(stock.transportation_rate) if stock.transportation_rate else 0,
                "actual_rate": float(stock.actual_rate) if stock.actual_rate else 0,
                "total_rate": float(stock.total_rate) if stock.total_rate else 0,
                "created_at": stock.created_at,
            })
        context["raw_material_stocks"] = raw_material_stocks
        context["suppliers"] = Supplier.objects.all().order_by('name')
        context["supplier_filter"] = supplier_filter
        context["receiving_from"] = receiving_from
        context["receiving_to"] = receiving_to
                
    elif report_type == "scrap":

        selected_date = request.GET.get("date")

        productions = (
            Production.objects
            .select_related(
                "product",
                "machine",
                "operator"
            )
            .prefetch_related("scrap_logs")
            .order_by("-created_at")
        )

        if selected_date:
            productions = productions.filter(created_at__date=selected_date)

        productions = productions.annotate(
            scrap_qty=Sum("scrap_logs__quantity"),
            scrap_weight=Sum("scrap_logs__weight")
        )

        context["productions"] = productions
        context["selected_date"] = selected_date

        # JSON data for JavaScript
        context["productions"] = [
            {
                "id": p.id,
                "created_at": p.created_at.strftime("%Y-%m-%d"),
                "shift": p.shift,
                "category": p.category,
                "sidepatti": p.sidepatti,
                "linesetting": p.linesetting,
                "rejected_quantity": p.rejected_quantity,
                "scrap_qty": p.scrap_qty or 0,
                "scrap_weight": p.scrap_weight or 0,
            }
            for p in productions
        ]

    elif report_type == "planning":
        planning_date = request.GET.get("date")
        planning_category = request.GET.get("category", "")
        planning_machines = request.GET.getlist("machine")

        plannings = Planning.objects.all().select_related("machine", "thickness", "height", "width", "size", "length", "color", "density").order_by("-date", "-id")

        if planning_date:
            plannings = plannings.filter(date=planning_date)

        if planning_category:
            plannings = plannings.filter(category=planning_category)

        if planning_machines:
            plannings = plannings.filter(machine_id__in=planning_machines)

        # Group planning data
        grouped_plannings = {}
        for plan in plannings:
            if plan.category in ["Door", "Sheet"]:
                key = (
                    plan.date.strftime("%Y-%m-%d") if plan.date else "-",
                    plan.machine.name if plan.machine else "-",
                    f"{plan.thickness.thickness} mm" if plan.thickness else "-",
                    str(plan.density) if plan.density else "-",
                    str(plan.color) if plan.color else "-",
                    plan.remark or "-"
                )
            else:  # Frame
                key = (
                    plan.date.strftime("%Y-%m-%d") if plan.date else "-",
                    plan.machine.name if plan.machine else "-",
                    plan.size.size if plan.size else "-",
                    str(plan.color) if plan.color else "-",
                    "-",
                    plan.remark or "-"
                )
            
            if key not in grouped_plannings:
                grouped_plannings[key] = {
                    "items": [],
                    "total_quantity": 0,
                    "total_weight": 0,
                }
            grouped_plannings[key]["items"].append(plan)
            grouped_plannings[key]["total_quantity"] += plan.quantity or 0
            grouped_plannings[key]["total_weight"] += float(plan.weight or 0) * float(plan.quantity or 0)

        context["grouped_plannings"] = grouped_plannings
        context["planning_date"] = planning_date or ""
        context["planning_category"] = planning_category
        context["planning_machines"] = planning_machines

    elif report_type == "formulation":
        formulation_date = request.GET.get("date")
        formulation_machines = request.GET.getlist("machine")

        formulations = Formulation.objects.all().select_related("machine", "planning").prefetch_related("items__raw_material").order_by("-date", "-id")

        if formulation_date:
            formulations = formulations.filter(date=formulation_date)

        if formulation_machines:
            formulations = formulations.filter(machine_id__in=formulation_machines)

        # Build formulation data with materials
        formulation_data = []
        for f in formulations:
            materials = []
            for item in f.items.all():
                materials.append({
                    "name": item.raw_material.name,
                    "weight": item.weight,
                    "total_weight": float(item.weight) * float(f.batches or 1)
                })
            
            formulation_data.append({
                "formulation": f,
                "materials": materials,
                "total_material_weight": sum(m["total_weight"] for m in materials)
            })

        context["formulation_data"] = formulation_data
        context["formulation_date"] = formulation_date or ""
        context["formulation_machines"] = formulation_machines

    elif report_type == "production_new":
        production_report_date = request.GET.get("date")
        production_report_machines = request.GET.getlist("machine")

        productions_qs = Production.objects.all().select_related("product", "machine", "operator", "thickness", "height", "width", "density", "color", "size", "length")

        if production_report_date:
            productions_qs = productions_qs.filter(created_at__date=production_report_date)

        if production_report_machines:
            productions_qs = productions_qs.filter(machine_id__in=production_report_machines)

        productions_report = productions_qs.order_by("-created_at", "-id")

        # Group production data
        grouped_productions = {}
        for p in productions_report:
            if p.category in ["Door", "Sheet"]:
                key = (
                    p.created_at.strftime("%Y-%m-%d") if p.created_at else "-",
                    p.machine.name if p.machine else "-",
                    f"{p.thickness.thickness} mm" if p.thickness else "-",
                    str(p.density) if p.density else "-",
                    str(p.color) if p.color else "-",
                    p.remark or "-"
                )
            else:  # Frame
                key = (
                    p.created_at.strftime("%Y-%m-%d") if p.created_at else "-",
                    p.machine.name if p.machine else "-",
                    p.size.size if p.size else "-",
                    str(p.color) if p.color else "-",
                    "-",
                    p.remark or "-"
                )
            
            if key not in grouped_productions:
                grouped_productions[key] = {
                    "items": [],
                    "total_qty": 0,
                    "total_weight": 0,
                    "total_sidepatti": 0,
                    "total_linesetting": 0,
                }
            grouped_productions[key]["items"].append(p)
            grouped_productions[key]["total_qty"] += p.quantity or 0
            grouped_productions[key]["total_weight"] += float(p.quantity or 0) * float(p.weight_per_piece or 0)
            grouped_productions[key]["total_sidepatti"] += float(p.sidepatti or 0)
            grouped_productions[key]["total_linesetting"] += float(p.linesetting or 0)

        context["grouped_productions_report"] = grouped_productions
        context["production_report_date"] = production_report_date or ""
        context["production_report_machines"] = production_report_machines

    elif report_type == "one_day_stock_alert":
        stock_alert_category = request.GET.get("stock_alert_category", "")
        
        stock_alert_data = []
        raw_materials = RawMaterial.objects.filter(is_active=True)
        
        if stock_alert_category:
            raw_materials = raw_materials.filter(category=stock_alert_category)
        
        for rm in raw_materials:
            if rm.current_stock < rm.one_day_requirement:
                last_received = RawMaterialStock.objects.filter(
                    raw_material=rm,
                    movement_type="IN"
                ).order_by("-created_at").first()
                
                stock_alert_data.append({
                    "id": rm.id,
                    "name": rm.name,
                    "category": rm.category,
                    "current_stock": rm.current_stock or 0,
                    "one_day_requirement": rm.one_day_requirement or 0,
                    "shortage": (rm.one_day_requirement or 0) - (rm.current_stock or 0),
                    "last_received": last_received.created_at if last_received else None,
                })
        
        stock_alert_data.sort(key=lambda x: x["last_received"] if x["last_received"] else timezone.datetime.min.replace(tzinfo=timezone.utc))
        
        context["stock_alert_data"] = stock_alert_data
        context["stock_alert_category"] = stock_alert_category
        context["stock_alert_categories"] = RawMaterial.CATEGORY_CHOICES

    elif report_type == "costing":
        costing_date = request.GET.get("costing_date")
        costing_machine = request.GET.get("costing_machine")
        
        costing_data = []
        machines = Machine.objects.filter(is_active=True).order_by("name")
        
        if costing_date:
            # Fetch all plannings for the selected date
            plannings = Planning.objects.filter(date=costing_date)
            if costing_machine:
                plannings = plannings.filter(machine_id=costing_machine)
            plannings = plannings.select_related("machine", "thickness", "density", "color", "size", "length")
            
            # Build group key → list of plans mapping
            groups = defaultdict(list)
            for p in plannings:
                if p.category in ["Door", "Sheet"]:
                    thickness_str = f"{p.thickness.thickness} mm" if p.thickness else "-"
                    density_str = str(p.density) if p.density else "-"
                    color_str = str(p.color) if p.color else "-"
                    key = (p.date.strftime("%Y-%m-%d"), p.machine.name if p.machine else "-", thickness_str, density_str, color_str)
                else:  # Frame
                    size_str = str(p.size) if p.size else "-"
                    length_str = f"{p.length.length} ft" if p.length else "-"
                    color_str = str(p.color) if p.color else "-"
                    key = (p.date.strftime("%Y-%m-%d"), p.machine.name if p.machine else "-", size_str, length_str, color_str)
                groups[key].append(p)
            
            # Build costing rows
            for key, plans in groups.items():
                date_str, machine_name = key[0], key[1]
                category = plans[0].category if plans else "-"
                if category in ["Door", "Sheet"]:
                    group_name = f"Date: {date_str} | Machine: {machine_name} | Thickness: {key[2]} | Density: {key[3]} | Color: {key[4]}"
                else:
                    group_name = f"Date: {date_str} | Machine: {machine_name} | Size: {key[2]} | Length: {key[3]} | Color: {key[4]}"
                
                # Collect planning IDs in this group
                plan_ids = [p.id for p in plans]
                
                # Find related formulations
                formulations = Formulation.objects.filter(planning__in=plan_ids).select_related("machine")
                formulation_statuses = list(formulations.values_list("status", flat=True))
                
                if not formulation_statuses:
                    status = "Not Created"
                elif all(s == "Approved" for s in formulation_statuses):
                    status = "Approved"
                elif all(s == "Pending" for s in formulation_statuses):
                    status = "Pending"
                else:
                    status = "/".join(formulation_statuses)
                
                # Get the first plan's category for reference
                category = plans[0].category if plans else "-"
                
                costing_data.append({
                    "date": date_str,
                    "machine": machine_name,
                    "group_name": group_name,
                    "status": status,
                    "plan_ids": plan_ids,
                    "category": category,
                })
        
        # Sort by machine then group name
        costing_data.sort(key=lambda x: (x["machine"], x["group_name"]))
        
        context["costing_data"] = costing_data
        context["costing_date"] = costing_date or ""
        context["costing_machine"] = costing_machine or ""
        context["costing_machines"] = machines

    elif report_type == "raw_material_stock":
        export_type = request.GET.get("export")
        
        # Fetch all active raw materials ordered by category, then name
        raw_materials = RawMaterial.objects.filter(is_active=True).order_by('category', 'name')
        
        # Group materials by category
        categories_data = defaultdict(lambda: {
            'materials': [],
            'total_stock': 0,
            'total_one_day_req': 0
        })
        
        for rm in raw_materials:
            category = rm.category
            stock = float(rm.current_stock or 0)
            one_day_req = float(rm.one_day_requirement or 0)
            
            categories_data[category]['materials'].append({
                'name': rm.name,
                'stock': stock,
                'one_day_req': one_day_req,
            })
            categories_data[category]['total_stock'] += stock
            categories_data[category]['total_one_day_req'] += one_day_req
        
        # Build final report structure with category-level calculations
        raw_material_stock_report = []
        for category, data in categories_data.items():
            total_stock = data['total_stock']
            total_one_day_req = data['total_one_day_req']
            
            # Calculate category days remaining
            if total_one_day_req > 0:
                category_days = round(total_stock / total_one_day_req, 1)
            else:
                category_days = 0
            
            # Calculate minimum stock level (10 days requirement)
            min_stock_level = round(total_one_day_req * 10, 2)
            
            # Calculate days for each material
            materials_with_days = []
            for mat in data['materials']:
                if mat['one_day_req'] > 0:
                    days = round(mat['stock'] / mat['one_day_req'], 1)
                else:
                    days = 0
                materials_with_days.append({
                    'name': mat['name'],
                    'stock': mat['stock'],
                    'days': days
                })
            
            raw_material_stock_report.append({
                'category': category,
                'category_stock': total_stock,
                'category_one_day_req': total_one_day_req,
                'category_days': category_days,
                'min_stock_level': min_stock_level,
                'materials': materials_with_days
            })
        
        context['raw_material_stock_report'] = raw_material_stock_report
        
        # Handle Excel export
        if export_type == "excel":
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime as dt
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Raw Material Stock"
            
            # Headers
            headers = ["S.No", "Category", "Raw Material", "Stock", "Category Stock", "Min Stock Level", "Days Remaining"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="3F2E97", end_color="3F2E97", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            row_num = 1
            for cat_group in raw_material_stock_report:
                if cat_group['materials']:
                    for idx, material in enumerate(cat_group['materials']):
                        if idx == 0:
                            ws.append([
                                row_num,
                                cat_group['category'],
                                material['name'],
                                material['stock'],
                                cat_group['category_stock'],
                                cat_group['min_stock_level'],
                                cat_group['category_days']
                            ])
                        else:
                            ws.append([
                                "",
                                "",
                                material['name'],
                                material['stock'],
                                "",
                                "",
                                ""
                            ])
                        row_num += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"raw_material_stock_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            wb.save(response)
            return response
        
        # Handle CSV export
        elif export_type == "csv":
            import csv
            from datetime import datetime as dt
            
            response = HttpResponse(content_type='text/csv')
            filename = f"raw_material_stock_{dt.now().strftime('%Y%m%d_%H%M%S')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            writer = csv.writer(response)
            writer.writerow(["S.No", "Category", "Raw Material", "Stock", "Category Stock", "Min Stock Level", "Days Remaining"])
            
            row_num = 1
            for cat_group in raw_material_stock_report:
                if cat_group['materials']:
                    for idx, material in enumerate(cat_group['materials']):
                        if idx == 0:
                            writer.writerow([
                                row_num,
                                cat_group['category'],
                                material['name'],
                                material['stock'],
                                cat_group['category_stock'],
                                cat_group['min_stock_level'],
                                cat_group['category_days']
                            ])
                        else:
                            writer.writerow([
                                "",
                                "",
                                material['name'],
                                material['stock'],
                                "",
                                "",
                                ""
                            ])
                        row_num += 1
            
            return response

    return render(request, "reports.html", context)

# done
def export_report(request):
    report_type = request.GET.get("type", "production")
    mode = request.GET.get("mode")

    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    machines_filter = request.GET.getlist("machine")
    shift_filter = request.GET.get("shift")

    wb = openpyxl.Workbook()
    
    # ================= PRODUCTION =================
    if report_type == "production":
        data = Production.objects.all()

        # ---- SAME FILTER LOGIC AS YOUR VIEW ----
        if mode == "by_date":
            date_only = request.GET.get("date")
            if date_only:
                data = data.filter(created_at__date=date_only)
            else:
                data = data.none()
        else:
            if from_date and to_date:
                data = data.filter(created_at__date__range=[from_date, to_date])
            if machines_filter:
                data = data.filter(machine_id__in=machines_filter)
            if shift_filter:
                data = data.filter(shift=shift_filter)

        # ---- SAME ANNOTATIONS ----
        data = data.annotate(calc_total_weight=F("quantity") * F("weight_per_piece"))
        data = data.annotate(calc_actual=F("quantity") * F("weight_per_piece"))
        data = data.annotate(
            calc_overall=F("quantity") * F("weight_per_piece") + F("sidepatti") + F("linesetting")
        )

        # ================= SHEET 1: SUMMARY =================
        ws1 = wb.active
        ws1.title = "Summary"

        if mode == "by_machine":
            ws1.append(["Machine", "Shift", "Category", "Total Qty", "Total Weight"])

            machines = data.values("machine__id", "machine__name").distinct()
            for m in machines:
                shifts = data.filter(machine_id=m["machine__id"]).values("shift", "category").distinct()
                for s in shifts:
                    agg = data.filter(
                        machine_id=m["machine__id"],
                        shift=s["shift"],
                        category=s["category"]
                    ).aggregate(
                        total_qty=Sum("quantity"),
                        total_weight=Sum(F("quantity") * F("weight_per_piece"))
                    )

                    ws1.append([
                        m["machine__name"],
                        s["shift"],
                        s["category"],
                        agg["total_qty"] or 0,
                        round(float(agg["total_weight"] or 0), 2),
                    ])

        elif mode == "by_operator":
            ws1.append(["Operator", "Shift", "Total Qty", "Total Weight"])

            operators = data.values("operator__id", "operator__full_name").distinct()
            for op in operators:
                shifts = data.filter(operator_id=op["operator__id"]).values("shift").distinct()
                for s in shifts:
                    agg = data.filter(
                        operator_id=op["operator__id"],
                        shift=s["shift"]
                    ).aggregate(
                        total_qty=Sum("quantity"),
                        total_weight=Sum(F("quantity") * F("weight_per_piece"))
                    )

                    ws1.append([
                        op["operator__full_name"],
                        s["shift"],
                        agg["total_qty"] or 0,
                        round(float(agg["total_weight"] or 0), 2),
                    ])

        elif mode == "by_date":
            ws1.append(["Date", "Shift", "Total Qty", "Total Weight"])

            date_only = request.GET.get("date")
            shifts = data.values("shift").distinct()

            for s in shifts:
                agg = data.filter(shift=s["shift"]).aggregate(
                    total_qty=Sum("quantity"),
                    total_weight=Sum(F("quantity") * F("weight_per_piece"))
                )

                ws1.append([
                    date_only,
                    s["shift"],
                    agg["total_qty"] or 0,
                    round(float(agg["total_weight"] or 0), 2),
                ])

        # ================= SHEET 2: DETAILS =================
        ws2 = wb.create_sheet(title="Details")

        ws2.append([
            "Date", "Machine", "Operator", "Shift", "Category",
            "Product", "Side Patti", "Line Setting",
            "Qty", "Actual", "Overall"
        ])

        for p in data.select_related("machine", "operator", "product"):
            ws2.append([
                p.created_at.strftime("%Y-%m-%d") if p.created_at else "",
                p.machine.name if p.machine else "",
                p.operator.full_name if p.operator else "",
                p.shift,
                p.category,
                p.product.product_name if p.product else "",
                p.sidepatti,
                p.linesetting,
                p.quantity,
                round(float(p.calc_actual or 0), 2),
                round(float(p.calc_overall or 0), 2),
            ])

    # ================= DISPATCH =================
    elif report_type == "dispatch":
        ws = wb.active
        ws.title = "Dispatch"

        dispatches = DeliveryChallan.objects.all()
        if from_date and to_date:
            dispatches = dispatches.filter(challan_date__range=[from_date, to_date])

        dispatches = dispatches.prefetch_related("items__product").order_by("-challan_date", "-id")

        ws.append(["Date", "Challan No", "Party", "Vehicle", "LR/RR No", "Products", "Total Qty", "Operator"])

        for challan in dispatches:
            items = list(challan.items.all())
            if not items:
                continue

            products_list = ", ".join([
                f"{item.product.product_name} ({item.quantity})"
                for item in items
            ])
            total_qty = sum(item.quantity for item in items)

            ws.append([
                challan.challan_date.strftime("%Y-%m-%d") if challan.challan_date else "",
                challan.challan_no,
                challan.party_name,
                challan.vehicle_no,
                challan.lr_rr_no or "-",
                products_list,
                total_qty,
                challan.operator.full_name if challan.operator else "",
            ])

    # ================= CONSUMPTION =================
    elif report_type == "consumption":
        ws = wb.active
        ws.title = "Consumption"

        # Get date range for consumption
        from_date = request.GET.get("from_date")
        to_date = request.GET.get("to_date")

        if from_date and to_date:
            start_date = datetime.strptime(from_date, "%Y-%m-%d").date()
            end_date = datetime.strptime(to_date, "%Y-%m-%d").date()

            # Get machines and columns like in the view
            all_machines = list(Machine.objects.filter(is_active=True).order_by("name"))
            machine_cols = []
            for m in all_machines:
                machine_cols.extend([
                    f"{m.name} Day",
                    f"{m.name} Night"
                ])

            # Get materials used in range (from BatchAddHistory via +Add button)
            materials = RawMaterial.objects.filter(
                formulationbatch__batch_history__date__range=[start_date, end_date]
            ).distinct()

            # Get received stock
            received_rows = RawMaterialStock.objects.filter(
                movement_type="IN",
                created_at__date__range=[start_date, end_date],
                raw_material__in=materials
            ).values(
                "raw_material_id",
                "created_at__date"
            ).annotate(
                qty=Sum("quantity")
            )

            received_map = {}
            for r in received_rows:
                key = (r["raw_material_id"], str(r["created_at__date"]))
                received_map[key] = r["qty"] or 0

            # Get consumption data (from BatchAddHistory via +Add button)
            usage_rows = FormulationBatch.objects.filter(
                raw_material__in=materials,
                batch_history__date__range=[start_date, end_date]
            ).values(
                "raw_material_id",
                "batch_history__date",
                "batch_history__machine__name",
                "batch_history__shift",
                "batch_history__formulation__is_loss",
            ).annotate(
                per_batch_weight=Sum(F("weight")),
                total_qty=Sum(F("weight") * F("batch_history__batch_count")),
                batches=Sum("batch_history__batch_count")
            )

            usage_map = {}

            for row in usage_rows:
                raw_material_id = row["raw_material_id"]
                date_key = str(row["batch_history__date"])
                is_loss = row["batch_history__formulation__is_loss"]

                if is_loss:
                    key = (raw_material_id, date_key, "Production Loss")
                else:
                    machine_name = row["batch_history__machine__name"]
                    shift = row["batch_history__shift"]

                    col = f"{machine_name} {shift}"

                    key = (raw_material_id, date_key, col)

                existing = usage_map.get(key, {"per_batch_weight": 0, "total_qty": 0, "batches": 0})
                usage_map[key] = {
                    "per_batch_weight": existing["per_batch_weight"] + (row["per_batch_weight"] or 0),
                    "total_qty": existing["total_qty"] + (row["total_qty"] or 0),
                    "batches": existing["batches"] + (row["batches"] or 0),
                }

            # Build header
            header = ["S.No", "DATE", "CHEMICAL NAME", "OPENING STOCK", "ON DAY RECEIVED"]
            for col in machine_cols:
                header.extend([f"{col} Qty", f"{col} Batch", f"{col} Total"])
            header.extend([
                "Production Loss Qty", "Production Loss Batch", "Production Loss Total",
                "Total Chemical Used", "AVL. BALANCE"
            ])
            ws.append(header)

            # Build data rows
            row_counter = 1
            current_date = start_date

            while current_date <= end_date:
                date_str = current_date.strftime("%Y-%m-%d")

                for rm in materials:
                    opening = rm.current_stock or 0
                    received = received_map.get((rm.id, date_str), 0)

                    loss_by_combo = {}
                    total_loss = 0

                    for col in machine_cols:
                        entry = usage_map.get((rm.id, date_str, col), {"per_batch_weight": 0, "total_qty": 0, "batches": 0})
                        loss_by_combo[col] = entry
                        total_loss += entry["total_qty"]

                    production_loss_entry = usage_map.get(
                        (rm.id, date_str, "Production Loss"),
                        {"per_batch_weight": 0, "total_qty": 0, "batches": 0}
                    )
                    total_loss += production_loss_entry["total_qty"]

                    # Skip zero rows like in view
                    if received == 0 and total_loss == 0:
                        continue

                    total_consumed = total_loss
                    balance = opening + received - total_loss

                    row = [
                        row_counter,
                        date_str,
                        rm.name,
                        round(float(opening), 2),
                        round(float(received), 2)
                    ]

                    for col in machine_cols:
                        entry = loss_by_combo[col]
                        row.extend([
                            round(float(entry["per_batch_weight"]), 2),
                            entry["batches"],
                            round(float(entry["total_qty"]), 2)
                        ])

                    row.extend([
                        round(float(production_loss_entry["per_batch_weight"]), 2),
                        production_loss_entry["batches"],
                        round(float(production_loss_entry["total_qty"]), 2),
                        round(float(total_consumed), 2),
                        round(float(balance), 2)
                    ])

                    ws.append(row)
                    row_counter += 1

                current_date += timedelta(days=1)

            # Add grand total row
            # First, calculate totals
            total_opening = 0
            total_received = 0
            total_machine_qty = {col: 0 for col in machine_cols}
            total_machine_batches = {col: 0 for col in machine_cols}
            total_machine_total = {col: 0 for col in machine_cols}
            total_prod_loss_qty = 0
            total_prod_loss_batches = 0
            total_prod_loss_total = 0
            total_consumed = 0
            total_balance = 0

            # Reset and recalculate
            current_date = start_date
            while current_date <= end_date:
                date_str = current_date.strftime("%Y-%m-%d")

                for rm in materials:
                    opening = rm.current_stock or 0
                    received = received_map.get((rm.id, date_str), 0)

                    loss_by_combo = {}
                    total_loss = 0

                    for col in machine_cols:
                        entry = usage_map.get((rm.id, date_str, col), {"per_batch_weight": 0, "total_qty": 0, "batches": 0})
                        loss_by_combo[col] = entry
                        total_loss += entry["total_qty"]

                    production_loss_entry = usage_map.get(
                        (rm.id, date_str, "Production Loss"),
                        {"per_batch_weight": 0, "total_qty": 0, "batches": 0}
                    )
                    total_loss += production_loss_entry["total_qty"]

                    if received == 0 and total_loss == 0:
                        continue

                    total_opening += opening
                    total_received += received
                    total_prod_loss_qty += production_loss_entry["per_batch_weight"]
                    total_prod_loss_batches += production_loss_entry["batches"]
                    total_prod_loss_total += production_loss_entry["total_qty"]
                    total_consumed += total_loss
                    total_balance += (opening + received - total_loss)

                    for col in machine_cols:
                        entry = loss_by_combo[col]
                        total_machine_qty[col] += entry["per_batch_weight"]
                        total_machine_batches[col] += entry["batches"]
                        total_machine_total[col] += entry["total_qty"]

                current_date += timedelta(days=1)

            # Append total row
            total_row = ["", "TOTAL", "", round(float(total_opening), 2), round(float(total_received), 2)]
            for col in machine_cols:
                total_row.extend([
                    round(float(total_machine_qty[col]), 2),
                    "-",  # batches not summed in display
                    round(float(total_machine_total[col]), 2)
                ])
            total_row.extend([
                round(float(total_prod_loss_qty), 2),
                "-",  # batches not summed
                round(float(total_prod_loss_total), 2),
                round(float(total_consumed), 2),
                round(float(total_balance), 2)
            ])
            ws.append(total_row)

    elif report_type == "planning":
        ws = wb.active
        ws.title = "Planning"

        planning_date = request.GET.get("date")
        planning_category = request.GET.get("category", "")
        planning_machines = request.GET.getlist("machine")

        data = Planning.objects.all().select_related("machine", "thickness", "height", "width", "size", "length", "color", "density").order_by("-date", "-id")

        if planning_date:
            data = data.filter(date=planning_date)

        if planning_category:
            data = data.filter(category=planning_category)

        if planning_machines:
            data = data.filter(machine_id__in=planning_machines)

        ws.append(["S.No", "Date", "Category", "Machine", "Quantity", "Weight", "Stamp", "Masking", "Remark"])

        for idx, plan in enumerate(data, 1):
            ws.append([
                idx,
                plan.date.strftime("%Y-%m-%d") if plan.date else "",
                plan.category,
                plan.machine.name if plan.machine else "",
                plan.quantity,
                round(float(plan.weight), 2),
                plan.stamp or "-",
                plan.masking or "-",
                plan.remark or "-",
            ])

    elif report_type == "formulation":
        ws = wb.active
        ws.title = "Formulation"

        formulation_date = request.GET.get("date")
        formulation_machines = request.GET.getlist("machine")

        data = Formulation.objects.all().select_related("machine", "planning").order_by("-date", "-id")

        if formulation_date:
            data = data.filter(date=formulation_date)

        if formulation_machines:
            data = data.filter(machine_id__in=formulation_machines)

        ws.append(["S.No", "Date", "Name", "Machine", "Shift", "Batches", "Status", "Planning"])

        for idx, f in enumerate(data, 1):
            ws.append([
                idx,
                f.date.strftime("%Y-%m-%d") if f.date else "",
                f.name,
                f.machine.name if f.machine else "",
                f.shift,
                f.batches or "-",
                f.status,
                f"{f.planning.date.strftime('%Y-%m-%d')} - {f.planning.category} - {f.planning.machine.name}" if f.planning else "-",
            ])

    elif report_type == "production_new":
        ws = wb.active
        ws.title = "Production Report"

        production_report_date = request.GET.get("date")
        production_report_machines = request.GET.getlist("machine")

        data = Production.objects.all().select_related("product", "machine", "operator")

        if production_report_date:
            data = data.filter(created_at__date=production_report_date)

        if production_report_machines:
            data = data.filter(machine_id__in=production_report_machines)

        data = data.order_by("-created_at", "-id")

        ws.append(["S.No", "Date", "Machine", "Shift", "Category", "Product", "Qty", "Weight", "Side Patti", "Line Setting"])

        for idx, p in enumerate(data, 1):
            ws.append([
                idx,
                p.created_at.strftime("%Y-%m-%d") if p.created_at else "",
                p.machine.name if p.machine else "",
                p.shift,
                p.category,
                p.product.product_name if p.product else "",
                p.quantity,
                round(float(p.quantity * p.weight_per_piece), 2),
                round(float(p.sidepatti), 2),
                round(float(p.linesetting), 2),
            ])

    elif report_type == "raw_material":
        supplier_filter = request.GET.get("supplier", "")
        receiving_from = request.GET.get("receiving_from", "")
        receiving_to = request.GET.get("receiving_to", "")

        ws = wb.active
        ws.title = "Raw Material Purchases"

        data = RawMaterialStock.objects.filter(
            movement_type='IN',
            purchase_invoice__isnull=False
        ).select_related(
            'raw_material',
            'purchase_invoice',
            'purchase_invoice__supplier'
        ).order_by('-purchase_invoice__receiving_date')

        if supplier_filter:
            data = data.filter(purchase_invoice__supplier__name__icontains=supplier_filter)
        
        if receiving_from:
            data = data.filter(purchase_invoice__receiving_date__gte=receiving_from)
        
        if receiving_to:
            data = data.filter(purchase_invoice__receiving_date__lte=receiving_to)

        ws.append([
            "S.No",
            "Receiving Date",
            "Supplier Name",
            "Invoice No",
            "Invoice Date",
            "Raw Material Name",
            "Quantity"
        ])

        for idx, stock in enumerate(data, 1):
            ws.append([
                idx,
                stock.purchase_invoice.receiving_date if stock.purchase_invoice else "-",
                stock.purchase_invoice.supplier.name if stock.purchase_invoice else "-",
                stock.purchase_invoice.invoice_no if stock.purchase_invoice else "-",
                stock.purchase_invoice.invoice_date if stock.purchase_invoice else "-",
                stock.raw_material.name,
                round(float(stock.quantity), 2)
            ])

    elif report_type == "one_day_stock_alert":
        stock_alert_category = request.GET.get("stock_alert_category", "")

        ws = wb.active
        ws.title = "One Day Stock Alert"

        ws.append([
            "S.No",
            "Raw Material Name",
            "Category",
            "In Stock",
            "One Day Req",
            "Shortage",
            "Last Received"
        ])

        raw_materials = RawMaterial.objects.filter(is_active=True)
        if stock_alert_category:
            raw_materials = raw_materials.filter(category=stock_alert_category)

        alert_data = []
        for rm in raw_materials:
            if rm.current_stock < rm.one_day_requirement:
                last_received = RawMaterialStock.objects.filter(
                    raw_material=rm,
                    movement_type="IN"
                ).order_by("-created_at").first()

                alert_data.append({
                    "id": rm.id,
                    "name": rm.name,
                    "category": rm.category,
                    "current_stock": rm.current_stock or 0,
                    "one_day_requirement": rm.one_day_requirement or 0,
                    "shortage": (rm.one_day_requirement or 0) - (rm.current_stock or 0),
                    "last_received": last_received.created_at if last_received else None,
                })

        alert_data.sort(key=lambda x: x["last_received"] if x["last_received"] else timezone.datetime.min.replace(tzinfo=timezone.utc))

        for idx, item in enumerate(alert_data, 1):
            ws.append([
                idx,
                item["name"],
                item["category"],
                round(float(item["current_stock"]), 2),
                round(float(item["one_day_requirement"]), 2),
                round(float(item["shortage"]), 2),
                item["last_received"].strftime("%d %b %Y") if item["last_received"] else "-"
            ])

    # ================= RESPONSE =================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"{report_type}_report.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

# calculator
def density_rate_calculator(request):
    return render(request, 'calculator.html')


# ================= API ENDPOINTS FOR ADDING NEW DROPDOWN VALUES =================
@login_required
def api_create_size(request):
    """API endpoint to create a new Size"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        size_value = request.POST.get('size', '').strip()
        standard_weight = request.POST.get('standard_weight', 0)
        
        if not size_value:
            return JsonResponse({'error': 'Size value is required'}, status=400)
        
        # Check if size already exists
        existing = Size.objects.filter(size__iexact=size_value).first()
        if existing:
            return JsonResponse({
                'id': existing.id,
                'size': existing.size,
                'standard_weight': existing.standard_weight,
                'already_exists': True
            })
        
        # Create new size
        size_obj = Size.objects.create(size=size_value, standard_weight=standard_weight)
        return JsonResponse({
            'id': size_obj.id,
            'size': size_obj.size,
            'standard_weight': size_obj.standard_weight,
            'success': True
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_create_color(request):
    """API endpoint to create a new Color"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        color_value = request.POST.get('color', '').strip()
        
        if not color_value:
            return JsonResponse({'error': 'Color value is required'}, status=400)
        
        # Check if color already exists
        existing = Color.objects.filter(color__iexact=color_value).first()
        if existing:
            return JsonResponse({
                'id': existing.id,
                'color': existing.color,
                'already_exists': True
            })
        
        # Create new color
        color_obj = Color.objects.create(color=color_value)
        return JsonResponse({
            'id': color_obj.id,
            'color': color_obj.color,
            'success': True
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



@login_required
def api_create_length(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        length_value = request.POST.get('length')
        unit = request.POST.get('unit', 'ft')
        display = request.POST.get('display')

        if not length_value:
            return JsonResponse({'error': 'Length required'}, status=400)

        length_float = float(length_value)

        # Normalize to feet for storage; keep original value and unit
        if unit == 'mm':
            length_ft = length_float / 304.8
        else:
            unit = 'ft'
            length_ft = length_float

        existing = Length.objects.filter(length=length_ft).first()
        if existing:
            return JsonResponse({
                'id': existing.id,
                'length': format_float(existing.length),
                'unit': existing.unit,
                'display': display or str(existing),
                'already_exists': True
            })

        length_obj = Length.objects.create(length=length_ft, unit=unit, original_value=length_float)

        return JsonResponse({
            'id': length_obj.id,
            'length': format_float(length_obj.length),
            'unit': length_obj.unit,
            'display': display or str(length_obj),
            'success': True
        })

    except ValueError:
        return JsonResponse({'error': 'Invalid number'}, status=400)

@login_required
def api_create_thickness(request):
    """API endpoint to create a new Thickness"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        thickness_value = request.POST.get('thickness', '')
        
        if not thickness_value:
            return JsonResponse({'error': 'Thickness value is required'}, status=400)
        
        thickness_float = float(thickness_value)
        
        if thickness_float > 36:
            return JsonResponse({'error': 'Invalid number. Thickness cannot be greater than 36.'}, status=400)

        # Check if thickness already exists
        existing = Thickness.objects.filter(thickness=thickness_float).first()
        if existing:
            return JsonResponse({
                'id': existing.id,
                'thickness': existing.thickness,
                'already_exists': True
            })
        
        # Create new thickness
        thickness_obj = Thickness.objects.create(thickness=thickness_float)
        return JsonResponse({
            'id': thickness_obj.id,
            'thickness': thickness_obj.thickness,
            'success': True
        })
    except ValueError:
        return JsonResponse({'error': 'Thickness must be a number'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_create_height(request):
    """API endpoint to create a new Height"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        height_value = request.POST.get('height', '')
        unit = request.POST.get('unit', 'inch')
        original_value = request.POST.get('original_value', '')
        
        if not height_value or not original_value:
            return JsonResponse({'error': 'Height value is required'}, status=400)
        
        height_float = float(height_value)
        original_float = float(original_value)
        
        if height_float < 48:
            return JsonResponse({'error': 'Invalid number. Height cannot be less than 48 inches.'}, status=400)

        # Check if height already exists
        existing = Height.objects.filter(height=height_float, unit=unit, original_value=original_float).first()
        if existing:
            return JsonResponse({
                'id': existing.id,
                'height': existing.height,
                'unit': existing.unit,
                'original_value': existing.original_value,
                'already_exists': True
            })
        
        # Create new height
        height_obj = Height.objects.create(height=height_float, unit=unit, original_value=original_float)
        return JsonResponse({
            'id': height_obj.id,
            'height': height_obj.height,
            'unit': height_obj.unit,
            'original_value': height_obj.original_value,
            'success': True
        })
    except ValueError:
        return JsonResponse({'error': 'Height must be a number'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_create_width(request):
    """API endpoint to create a new Width"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        width_value = request.POST.get('width', '')
        unit = request.POST.get('unit', 'inch')
        original_value = request.POST.get('original_value', '')
        
        if not width_value or not original_value:
            return JsonResponse({'error': 'Width value is required'}, status=400)
        
        width_float = float(width_value)
        original_float = float(original_value)
        
        if width_float > 49:
            return JsonResponse({'error': 'Invalid number. Width cannot be greater than 49 inches.'}, status=400)

        # Check if width already exists
        existing = Width.objects.filter(width=width_float, unit=unit, original_value=original_float).first()
        if existing:
            return JsonResponse({
                'id': existing.id,
                'width': existing.width,
                'unit': existing.unit,
                'original_value': existing.original_value,
                'already_exists': True
            })
        
        # Create new width
        width_obj = Width.objects.create(width=width_float, unit=unit, original_value=original_float)
        return JsonResponse({
            'id': width_obj.id,
            'width': width_obj.width,
            'unit': width_obj.unit,
            'original_value': width_obj.original_value,
            'success': True
        })
    except ValueError:
        return JsonResponse({'error': 'Width must be a number'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_create_density(request):
    """API endpoint to create a new Density"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        density_value = request.POST.get('density', '')
        density_name = request.POST.get('density_name', '').strip()
        
        if not density_value or not density_name:
            return JsonResponse({'error': 'Density value and name are required'}, status=400)
        
        density_int = int(density_value)
        
        # Check if density already exists
        existing = Density.objects.filter(density=density_int).first()
        if existing:
            return JsonResponse({
                'id': existing.id,
                'density': existing.density,
                'density_name': existing.density_name,
                'already_exists': True
            })
        
        # Create new density
        density_obj = Density.objects.create(density=density_int, density_name=density_name)
        return JsonResponse({
            'id': density_obj.id,
            'density': density_obj.density,
            'density_name': density_obj.density_name,
            'success': True
        })
    except ValueError:
        return JsonResponse({'error': 'Density must be a number'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)



# upload excel for production data

def upload_excel(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        df = pd.read_excel(file)

        for _, row in df.iterrows():

            # ===== COMMON =====
            date = pd.to_datetime(row.get("Date"))
            category = str(row.get("Category")).strip().title()
            shift = row.get("Shift")
            operator_name = str(row.get("Operator")).strip()
            machine_name = str(row.get("Line")).strip()

            qty = float(row.get("Quantity") or 0)
            weight = float(row.get("1 Pcs Kg") or 0)

            sidepatti = float(row.get("SidePatti") or 0)
            linesetting = float(row.get("LineSetting") or 0)

            color_val = str(row.get("Color")).strip()

            # ===== GET MASTER =====
            machine = Machine.objects.filter(name__iexact=machine_name).first()

            operator = CustomUser.objects.filter(full_name__iexact=operator_name).first()
            if not operator:
                operator = request.user   # fallback

            color, _ = Color.objects.get_or_create(color=color_val)

            # ===============================
            # 🟦 FRAME LOGIC
            # ===============================
            if category == "Frame":

                size_val = str(row.get("Size")).strip()
                length_val = float(row.get("Length Feet") or 0)

                size, _ = Size.objects.get_or_create(size=size_val)

                length, _ = Length.objects.get_or_create(
                    length=length_val,
                    defaults={
                        "original_value": length_val,
                        "unit": "ft"
                    }
                )

                temp_product = Product(
                    category="Frame",
                    size=size,
                    length=length,
                    color=color,
                )

                product_name = temp_product.generate_name()

                product = Product.objects.filter(product_name=product_name).first()

                if not product:
                    temp_product.product_name = product_name
                    temp_product.save()
                    product = temp_product

                Production.objects.create(
                    product=product,
                    category="Frame",
                    machine=machine,
                    operator=operator,
                    shift=shift,
                    size=size,
                    color=color,
                    length=length,
                    quantity=qty,
                    weight_per_piece=weight,
                    sidepatti=sidepatti,
                    linesetting=linesetting,
                    created_at=date
                )

            # ===============================
            # 🟥 DOOR LOGIC
            # ===============================
            elif category == "Door":

                thickness_val = float(row.get("Thickness") or 0)
                size_val = str(row.get("Size")).strip()
                density_val = row.get("DENSITY (AS PER ORDER)")
                density_name_val = str(row.get("PRODUCT TYPE") or "").strip()

                density = None
                density_name_obj = None

                if density_val and str(density_val).strip() != "":
                    density_val = int(float(density_val))
                    density_name_text = density_name_val or f"{density_val}D"

                    density_name_obj, _ = DensityName.objects.get_or_create(
                        name=density_name_text
                    )

                    density, _ = Density.objects.get_or_create(density=density_val)

                try:
                    width_val, height_val = size_val.lower().replace(" ", "").split("x")
                    width_val = float(width_val)
                    height_val = float(height_val)
                except:
                    print("❌ Invalid size:", size_val)
                    continue

                thickness, _ = Thickness.objects.get_or_create(thickness=thickness_val)

                width, _ = Width.objects.get_or_create(
                    width=width_val,
                    defaults={"original_value": width_val}
                )

                height, _ = Height.objects.get_or_create(
                    height=height_val,
                    defaults={"original_value": height_val}
                )

                temp_product = Product(
                    category="Door",
                    thickness=thickness,
                    width=width,
                    height=height,
                    color=color,
                    density=density,
                    density_name=density_name_obj
                )

                product_name = temp_product.generate_name()

                product = Product.objects.filter(product_name=product_name).first()

                if not product:
                    temp_product.product_name = product_name
                    temp_product.save()
                    product = temp_product

                Production.objects.create(
                    product=product,
                    category="Door",
                    machine=machine,
                    operator=operator,
                    shift=shift,
                    thickness=thickness,
                    width=width,
                    height=height,
                    color=color,
                    density=density,
                    density_name=density_name_obj,
                    quantity=qty,
                    weight_per_piece=weight,
                    sidepatti=sidepatti,
                    linesetting=linesetting,
                    created_at=date
                )

            else:
                continue

        return redirect("production")

    return render(request, "upload_excel.html")


# consumption view
#raw material page 
@login_required
def raw_material_list(request):

    if request.method == "POST":

        material_id = request.POST.get("material_id")
        name = request.POST.get("name")
        category = request.POST.get("category")
        opening_stock = request.POST.get("opening_stock")
        rate_per_kg = request.POST.get("rate_per_kg")

        # EDIT
        if material_id:
            rm = get_object_or_404(RawMaterial, id=material_id)
            rm.name = name
            rm.category = category
            rm.save()

            # Update latest stock rate if provided
            if rate_per_kg:
                latest_stock = RawMaterialStock.objects.filter(
                    raw_material=rm
                ).order_by('-created_at').first()
                if latest_stock:
                    latest_stock.rate_per_kg = float(rate_per_kg)
                    latest_stock.save()

        # ADD
        else:
            rm = RawMaterial.objects.create(
                name=name,
                category=category,
                current_stock=float(opening_stock or 0)
            )

            if opening_stock:
                RawMaterialStock.objects.create(
                    raw_material=rm,
                    quantity=float(opening_stock),
                    movement_type='IN'
                )

        return redirect("raw_material_list")

    today = now().date()
    materials = RawMaterial.objects.all().order_by("category", "name")

    # Apply category filter
    selected_category = request.GET.get('category')
    if selected_category:
        materials = materials.filter(category=selected_category)

    # Apply stock remaining filter
    stock_remaining = request.GET.get('stock_remaining')
    if stock_remaining:
        try:
            days = int(stock_remaining)
            materials = materials.filter(
                current_stock__lt=F('one_day_requirement') * days
            )
        except (ValueError, TypeError):
            pass

    for rm in materials:
        rm.today_added = RawMaterialStock.objects.filter(
            raw_material=rm,
            movement_type='IN',
            created_at__date=today
        ).aggregate(total=Sum('quantity'))['total'] or 0

        rm.today_consumed = RawMaterialStock.objects.filter(
            raw_material=rm,
            movement_type='OUT',
            created_at__date=today
        ).aggregate(total=Sum('quantity'))['total'] or 0

        last = RawMaterialStock.objects.filter(
            raw_material=rm
        ).order_by('-created_at').first()

        rm.last_updated = last.created_at if last else None

    suppliers = Supplier.objects.all().order_by('name')

    latest_stock_map = {}
    for rm in materials:
        last_stock = RawMaterialStock.objects.filter(
            raw_material=rm,
            movement_type='IN'
        ).order_by('-created_at').first()
        if last_stock:
            latest_stock_map[rm.id] = {
                'rate_per_kg': last_stock.rate_per_kg,
                'transportation_rate': last_stock.transportation_rate,
                'actual_rate': last_stock.actual_rate,
                'total_rate': last_stock.total_rate,
            }

    return render(request, "raw_material.html", {
        "materials": materials,
        "categories": RawMaterial.CATEGORY_CHOICES,
        "suppliers": suppliers,
        "latest_stock_map": latest_stock_map,
        "selected_category": selected_category,
        "selected_stock_remaining": stock_remaining,
    })

@login_required
def raw_material_history(request, rm_id):
    rm = get_object_or_404(RawMaterial, id=rm_id)

    last_10_days = now() - timedelta(days=10)

    logs_qs = RawMaterialStock.objects.filter(
        raw_material=rm
    ).select_related('purchase_invoice__supplier').order_by('-created_at')

    show_all = request.GET.get("all") == "1"

    logs = logs_qs if show_all else logs_qs.filter(created_at__gte=last_10_days)

    data = []
    for log in logs:
        supplier_name = "-"
        invoice_no = "-"
        invoice_date = "-"
        receiving_date = "-"
        
        if log.purchase_invoice:
            supplier_name = log.purchase_invoice.supplier.name or "-"
            invoice_no = log.purchase_invoice.invoice_no or "-"
            invoice_date = log.purchase_invoice.invoice_date.strftime("%d %b %Y") if log.purchase_invoice.invoice_date else "-"
            receiving_date = log.purchase_invoice.receiving_date.strftime("%d %b %Y") if log.purchase_invoice.receiving_date else "-"
        
        data.append({
            "date": log.created_at.strftime("%d %b %Y"),
            "type": log.movement_type,
            "qty": log.quantity,
            "supplier": supplier_name,
            "invoice_no": invoice_no,
            "invoice_date": invoice_date,
            "receiving_date": receiving_date,
        })

    return JsonResponse({
        "material": rm.name,
        "logs": data,
        "show_all": show_all
    })

@login_required
def raw_material_history_download(request, rm_id):
    rm = get_object_or_404(RawMaterial, id=rm_id)

    # full history
    logs = RawMaterialStock.objects.filter(
        raw_material=rm
    ).order_by('-created_at')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{rm.name}_history.csv"'

    writer = csv.writer(response)
    writer.writerow(["Date", "Type", "Quantity"])

    for log in logs:
        writer.writerow([
            log.created_at.strftime("%d-%m-%Y "),
            log.movement_type,
            log.quantity
        ])

    return response

@login_required
def add_stock(request):
    if request.method == "POST":
        supplier_name = (request.POST.get("supplier") or "").strip()
        invoice_no = (request.POST.get("invoice_no") or "").strip()
        invoice_date = parse_date(request.POST.get("invoice_date"))
        receiving_date = parse_date(request.POST.get("receiving_date"))
        transportation_rate = (request.POST.get("transportation_rate") or "0").strip()

        if supplier_name and invoice_no and invoice_date and receiving_date:
            supplier = Supplier.objects.filter(name__iexact=supplier_name).first()
            if not supplier:
                supplier = Supplier.objects.create(name=supplier_name)

            purchase_invoice, _ = PurchaseInvoice.objects.get_or_create(
                supplier=supplier,
                invoice_no=invoice_no,
                defaults={
                    "invoice_date": invoice_date,
                    "receiving_date": receiving_date,
                }
            )

            material_ids = request.POST.getlist("material[]")
            quantities = request.POST.getlist("qty[]")
            rate_per_kg_list = request.POST.getlist("rate_per_kg[]")

            # Sum total qty to compute actual_rate per row
            raw_qty_total = 0
            parsed_quantities = []
            for qty_value in quantities:
                try:
                    q = float(qty_value)
                except (TypeError, ValueError):
                    q = 0
                parsed_quantities.append(q)
                raw_qty_total += q

            try:
                trans_rate = Decimal(transportation_rate)
            except Exception:
                trans_rate = Decimal("0")

            for material_id, qty_value, rate_val in zip(material_ids, parsed_quantities, rate_per_kg_list):
                if not material_id or qty_value <= 0:
                    continue

                try:
                    rate_per_kg = Decimal(rate_val)
                except Exception:
                    rate_per_kg = Decimal("0")

                if raw_qty_total > 0:
                    actual_rate = rate_per_kg + (trans_rate / Decimal(str(raw_qty_total)))
                else:
                    actual_rate = rate_per_kg

                total_rate = actual_rate * Decimal(str(qty_value))

                rm = RawMaterial.objects.filter(id=material_id).first()
                if not rm:
                    continue

                rm.current_stock += qty_value
                rm.save()

                RawMaterialStock.objects.create(
                    raw_material=rm,
                    quantity=qty_value,
                    movement_type='IN',
                    purchase_invoice=purchase_invoice,
                    rate_per_kg=rate_per_kg,
                    transportation_rate=trans_rate,
                    actual_rate=actual_rate,
                    total_rate=total_rate
                )
    return redirect("raw_material_list")


@login_required
def costing_details(request):
    """API endpoint to get formulation material details with actual rates for costing report"""
    plan_ids = request.GET.get("plan_ids", "")
    if not plan_ids:
        return JsonResponse({"error": "No plan IDs provided"}, status=400)
    
    ids = [pid.strip() for pid in plan_ids.split(",") if pid.strip()]
    if not ids:
        return JsonResponse({"error": "Invalid plan IDs"}, status=400)
    
    try:
        # Find formulations linked to these planning IDs
        formulations = Formulation.objects.filter(
            planning_id__in=ids,
            status__in=["Pending", "Approved"]
        ).prefetch_related("items__raw_material")
        
        if not formulations.exists():
            return JsonResponse({"error": "No formulations found for these planning IDs"}, status=404)
        
        # Aggregate material weights across all formulations in the group
        material_totals = {}
        for f in formulations:
            for item in f.items.all():
                rm_id = item.raw_material.id
                if rm_id not in material_totals:
                    material_totals[rm_id] = {
                        "name": item.raw_material.name,
                        "weight": 0,
                    }
                material_totals[rm_id]["weight"] += float(item.weight) * float(f.batches or 1)
        
        # Get actual_rate from latest RawMaterialStock entry for each material
        materials = []
        total_weight = 0
        for rm_id, mat in material_totals.items():
            total_weight += mat["weight"]
            
            # Fetch the latest actual_rate from RawMaterialStock
            latest_stock = RawMaterialStock.objects.filter(
                raw_material_id=rm_id,
                movement_type='IN'
            ).order_by('-created_at').first()
            
            actual_rate = float(latest_stock.actual_rate) if latest_stock and latest_stock.actual_rate else 0
            
            materials.append({
                "id": rm_id,
                "name": mat["name"],
                "weight": round(mat["weight"], 2),
                "actual_rate": round(actual_rate, 2),
            })
        
        return JsonResponse({
            "materials": materials,
            "total_weight": round(total_weight, 2),
            "formulation_count": formulations.count(),
        })
        
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# @login_required
# @module_required('formulation')
# def formulation_page(request):

#     loss_obj = Formulation.objects.filter(
#         is_loss=True
#     ).order_by("-id").first()

#     if not loss_obj:
#         loss_obj = Formulation.objects.create(
#             name="Production Loss",
#             date=now().date(),
#             shift="",
#             machine=None,
#             created_by=request.user,
#             batches=0,
#             status="Pending",
#             is_loss=True
#         )

#     def ensure_loss_items(formulation):
#         default_materials = [
#             ("CALCIUM CARBONATE / MARBLE POWDER", 15),
#             ("BAEROPAN SMS 305NI", 1.5),
#             ("PV WAX", 0.4),
#             ("STEARIC ACID", 0.4),
#             ("PVC Resin", 10),
#         ]
#         # fixed_ids = []

#         for query, weight in default_materials:
#             rm = RawMaterial.objects.filter(name__icontains=query).first()
#             if not rm:
#                 rm = RawMaterial.objects.filter(name__iexact=query).first()
#             if not rm:
#                 continue

#             # fixed_ids.append(rm.id)
#             FormulationItem.objects.get_or_create(
#                 formulation=formulation,
#                 raw_material=rm,
#                 defaults={"weight": weight}
#             )

#         # formulation.items.exclude(raw_material__id__in=fixed_ids).delete()

#     ensure_loss_items(loss_obj)

#     # =========================
#     # LOSS DATE UPDATE (ADD HERE)
#     # =========================
#     if request.method == "POST" and "loss_date" in request.POST:
#         loss_obj.date = request.POST.get("loss_date")
#         loss_obj.save()
#         return redirect("formulation_page")
    

#     # =========================
#     # SAVE FORMULATION
#     # =========================
#     if request.method == "POST":

#         try:
#             name = request.POST.get("name")
#             date = request.POST.get("date")
#             machine_id = request.POST.get("machine")
#             is_loss = request.POST.get("is_loss") == "1"
#             materials_json = request.POST.get("materials")
#             plan_ids_raw = request.POST.get("plan_id")
#             plan_ids = [pid.strip() for pid in plan_ids_raw.split(",") if pid.strip()] if plan_ids_raw else []

#             if not name or not date or not machine_id:
#                 messages.error(request, "Fill all required fields.")
#                 return redirect("formulation_page")

#             machine = Machine.objects.get(id=machine_id)
#             materials = json.loads(materials_json)

#             created_count = 0
#             if plan_ids:
#                 for single_plan_id in plan_ids:
#                     plan = Planning.objects.filter(id=single_plan_id).first() if single_plan_id else None
#                     form = Formulation.objects.create(
#                         name=name,
#                         date=date,
#                         shift="Day",
#                         machine=machine,
#                         created_by=request.user,
#                         status="Pending",
#                         is_loss=is_loss,
#                         batches=1,
#                         planning=plan
#                     )

#                     for row in materials:
#                         raw = RawMaterial.objects.get(id=row["id"])
#                         FormulationItem.objects.create(
#                             formulation=form,
#                             raw_material=raw,
#                             weight=float(row["weight"])
#                         )
#                     created_count += 1
#             else:
#                 form = Formulation.objects.create(
#                     name=name,
#                     date=date,
#                     shift="Day",
#                     machine=machine,
#                     created_by=request.user,
#                     status="Pending",
#                     is_loss=is_loss,
#                     batches=1,
#                     planning=None
#                 )

#                 for row in materials:
#                     raw = RawMaterial.objects.get(id=row["id"])
#                     FormulationItem.objects.create(
#                         formulation=form,
#                         raw_material=raw,
#                         weight=float(row["weight"])
#                     )
#                 created_count = 1

#             audit_logger.info(f"Formulation(s) created: {name} ({created_count} formulation) by user {request.user.username}, Machine: {machine.name}, Loss: {is_loss}, Plan IDs: {plan_ids}")
#             business_logger.info(f"{created_count} formulation(s) added - Materials count: {len(materials)}, Date: {date}, Plan IDs: {plan_ids}")
#             messages.success(request, f"{created_count} Formulation(s) Added Successfully.")

#         except Exception as e:
#             messages.error(request, f"Error : {str(e)}")

#         return redirect("formulation_page")

#     # =========================
#     # GET DATA
#     # =========================

#     formulations = Formulation.objects.filter(
#         status="Pending"
#     ).prefetch_related("items__raw_material").annotate(
#     total_weight=Sum("items__weight")
#     ).order_by("-id")

#     machines = Machine.objects.filter(is_active=True)
#     # materials = RawMaterial.objects.filter(is_active=True)

#     materials_qs = RawMaterial.objects.filter(is_active=True).order_by("name")

#     category_sequence = [
#         'Raw Material',
#         'Processing Head (Processing Aids)',
#         'Stabilizer',
#         'CP',
#         'Lubrication',
#         'Wax',
#         'Foaming White',
#         'Foaming Yellow',
#         'Blister',
#         'Scrap',
#         'Pigment',
#     ]

#     materials_grouped = OrderedDict()

#     for cat in category_sequence:
#         materials_grouped[cat] = materials_qs.filter(category=cat)

#     context = {
#         "formulations": all_formulations,
#         "loss_form": loss_obj,   
#         "machines": machines,
#         # "materials": materials,
#         "materials_grouped": materials_grouped,
#     }

#     return render(request, "formulation.html", context)
@login_required
@module_required('formulation')
def formulation_page(request):

    loss_obj = Formulation.objects.filter(
        is_loss=True
    ).order_by("-id").first()

    if not loss_obj:
        loss_obj = Formulation.objects.create(
            name="Production Loss",
            date=now().date(),
            shift="",
            machine=None,
            created_by=request.user,
            batches=0,
            status="Pending",
            is_loss=True
        )

    def ensure_loss_items(formulation):
        default_materials = [
            ("CALCIUM CARBONATE / MARBLE POWDER", 15),
            ("BAEROPAN SMS 305NI", 1.5),
            ("PV WAX", 0.4),
            ("STEARIC ACID", 0.4),
            ("PVC Resin", 10),
        ]

        for query, weight in default_materials:
            rm = RawMaterial.objects.filter(name__icontains=query).first()
            if not rm:
                rm = RawMaterial.objects.filter(name__iexact=query).first()
            if not rm:
                continue

            FormulationItem.objects.get_or_create(
                formulation=formulation,
                raw_material=rm,
                defaults={"weight": weight}
            )

    ensure_loss_items(loss_obj)

    # =========================
    # LOSS DATE UPDATE
    # =========================
    if request.method == "POST" and "loss_date" in request.POST:
        loss_obj.date = request.POST.get("loss_date")
        loss_obj.save()
        return redirect("formulation_page")
    

    # =========================
    # SAVE FORMULATION
    # =========================
    if request.method == "POST":
        try:
            name = request.POST.get("name")
            date = request.POST.get("date")
            machine_id = request.POST.get("machine")
            is_loss = request.POST.get("is_loss") == "1"
            materials_json = request.POST.get("materials")
            plan_ids_raw = request.POST.get("plan_id")
            plan_ids = [pid.strip() for pid in plan_ids_raw.split(",") if pid.strip()] if plan_ids_raw else []

            if not name or not date or not machine_id:
                messages.error(request, "Fill all required fields.")
                return redirect("formulation_page")

            machine = Machine.objects.get(id=machine_id)
            materials = json.loads(materials_json) if materials_json else []

            created_count = 0
            if len(plan_ids) > 1:
                # Group submission: create ONE formulation per plan so all show up in production page
                for pid in plan_ids:
                    plan = Planning.objects.filter(id=pid).first()
                    if not plan:
                        continue
                    form = Formulation.objects.create(
                        name=name,
                        date=date,
                        shift="Day",
                        machine=machine,
                        created_by=request.user,
                        status="Pending",
                        is_loss=is_loss,
                        batches=1,
                        planning=plan
                    )
                    for row in materials:
                        raw = RawMaterial.objects.get(id=row["id"])
                        FormulationItem.objects.create(
                            formulation=form,
                            raw_material=raw,
                            weight=float(row["weight"])
                        )
                    created_count += 1
                messages.success(request, f"{created_count} Formulations Added Successfully for the group ({len(plan_ids)} plans).")
            elif len(plan_ids) == 1:
                # Single plan submission  
                single_plan = Planning.objects.filter(id=plan_ids[0]).first()
                form = Formulation.objects.create(
                    name=name,
                    date=date,
                    shift="Day",
                    machine=machine,
                    created_by=request.user,
                    status="Pending",
                    is_loss=is_loss,
                    batches=1,
                    planning=single_plan
                )
                for row in materials:
                    raw = RawMaterial.objects.get(id=row["id"])
                    FormulationItem.objects.create(
                        formulation=form,
                        raw_material=raw,
                        weight=float(row["weight"])
                    )
                created_count = 1
                messages.success(request, f"Formulation Added Successfully.")
            else:
                # No plan_ids: standalone formulation
                form = Formulation.objects.create(
                    name=name,
                    date=date,
                    shift="Day",
                    machine=machine,
                    created_by=request.user,
                    status="Pending",
                    is_loss=is_loss,
                    batches=1,
                    planning=None
                )
                for row in materials:
                    raw = RawMaterial.objects.get(id=row["id"])
                    FormulationItem.objects.create(
                        formulation=form,
                        raw_material=raw,
                        weight=float(row["weight"])
                    )
                created_count = 1
                messages.success(request, f"Formulation Added Successfully.")

            audit_logger.info(f"Formulation processed: {name} by user {request.user.username}, Machine: {machine.name}, Loss: {is_loss}, Created Count: {created_count}, Plan IDs: {plan_ids}")
            business_logger.info(f"Formulation processed - Materials count: {len(materials)}, Date: {date}, Created Count: {created_count}, Plan IDs: {plan_ids}")

        except Exception as e:
            messages.error(request, f"Error : {str(e)}")

        return redirect("formulation_page")

    # =========================
    # GET DATA
    # =========================
    formulations = Formulation.objects.filter(
        status="Pending"
    ).prefetch_related("items__raw_material").annotate(
        total_weight=Sum("items__weight")
    ).order_by("-id")

    machines = Machine.objects.filter(is_active=True)
    materials_qs = RawMaterial.objects.filter(is_active=True).order_by("name")

    category_sequence = [
        'Raw Material',
        'PVC Resin',
        'Marble Powder',
        'Processing Head (Processing Aids)',
        'Stabilizer',
        'CS',
        'CPE',
        'Lubrication',
        'Internal Lubricant',
        'External Lubricant',
        'Wax',
        'PE Wax',
        'Foaming White',
        'Foaming Yellow',
        'Blister',
        'Scrap',
        'Pigment',
        'Stearic Acid',
        'Other',
    ]

    materials_grouped = OrderedDict()
    for cat in category_sequence:
        materials_grouped[cat] = materials_qs.filter(category=cat)

    # =========================
    # HISTORY DATA (all formulations, not just pending)
    # =========================
    history_formulations = Formulation.objects.all().prefetch_related(
        "items__raw_material"
    ).select_related(
        "machine", "created_by", "planning"
    ).annotate(
        total_weight=Sum("items__weight")
    ).order_by("-id")[:10]  # Last 10 entries (newest first)

    # =========================
    # BATCH ADDITION REPORT (from +Add button) - uses BatchAddHistory for accuracy
    # =========================
    batch_report = []
    batch_history = BatchAddHistory.objects.all().select_related(
        "formulation__machine", "added_by"
    ).prefetch_related(
        "batches__raw_material"
    ).order_by("group_name", "-date")

    report_groups = defaultdict(lambda: {
        "times_added": 0,
        "total_batches": 0,
        "last_date": None,
        "machine": "-",
        "added_by": "-",
        "materials": {},
    })

    for entry in batch_history:
        g = report_groups[entry.group_name]
        g["times_added"] += 1
        g["total_batches"] += (entry.batch_count or 0)
        if g["last_date"] is None or entry.date > g["last_date"]:
            g["last_date"] = entry.date
            g["machine"] = entry.formulation.machine.name if entry.formulation and entry.formulation.machine else "-"
            g["added_by"] = entry.added_by.full_name if entry.added_by else "-"
        
        # Collect raw materials from FormulationBatch records
        for fb in entry.batches.all():
            if fb.raw_material:
                mat_name = fb.raw_material.name
                if mat_name not in g["materials"]:
                    g["materials"][mat_name] = 0
                g["materials"][mat_name] += float(fb.weight) * (entry.batch_count or 1)

    batch_report = [{
        "group_name": name,
        "times_added": data["times_added"],
        "total_batches": data["total_batches"],
        "avg_batches": round(data["total_batches"] / data["times_added"], 1) if data["times_added"] else 0,
        "last_date": data["last_date"],
        "machine": data["machine"],
        "added_by": data["added_by"],
        "materials": [{"name": m, "weight": round(w, 2)} for m, w in data["materials"].items()],
    } for name, data in report_groups.items()]

    # Sort by total batches descending
    batch_report.sort(key=lambda x: x["total_batches"], reverse=True)

    # =========================
    # BATCH DETAILS (individual +Add entries with materials)
    # =========================
    batch_details_qs = BatchAddHistory.objects.all().select_related(
        "formulation__machine", "added_by"
    ).prefetch_related(
        "batches__raw_material"
    ).order_by("-date", "-created_at")

    batch_details = []
    for entry in batch_details_qs:
        materials = []
        for fb in entry.batches.all():
            materials.append({
                'name': fb.raw_material.name if fb.raw_material else '-',
                'weight': fb.weight,
                'shift': fb.shift,
                'date': fb.date,
            })
        
        batch_details.append({
            'group_name': entry.group_name,
            'formulation_name': entry.formulation.name if entry.formulation else '-',
            'machine': entry.formulation.machine.name if entry.formulation and entry.formulation.machine else '-',
            'shift': entry.shift,
            'date': entry.date,
            'batch_count': entry.batch_count,
            'added_by': entry.added_by.full_name if entry.added_by else '-',
            'created_at': entry.created_at,
            'materials': materials,
        })

    context = {
        "formulations": formulations,
        "loss_form": loss_obj,   
        "machines": machines,
        "materials_grouped": materials_grouped,
        "history_formulations": history_formulations,
        "batch_report": batch_report,
        "batch_details": batch_details,
    }

    return render(request, "formulation.html", context)

@login_required
@module_required('formulation')
def frameline_page(request):
    # Handle Formulation POST submission from frameline page
    if request.method == "POST":
        try:
            name = request.POST.get("name")
            date = request.POST.get("date")
            machine_id = request.POST.get("machine")
            is_loss = request.POST.get("is_loss") == "1"
            materials_json = request.POST.get("materials")

            if not name or not date or not machine_id:
                messages.error(request, "Fill all required fields.")
                return redirect("frameline_page")

            machine = Machine.objects.get(id=machine_id)
            materials = json.loads(materials_json)

            plan_ids_raw = request.POST.get("plan_id")
            plan_ids = [pid.strip() for pid in plan_ids_raw.split(",") if pid.strip()] if plan_ids_raw else []

            created_count = 0
            if len(plan_ids) > 1:
                # Group submission: create ONE formulation per plan so all show up in production page
                for pid in plan_ids:
                    plan = Planning.objects.filter(id=pid).first()
                    if not plan:
                        continue
                    form = Formulation.objects.create(
                        name=plan.category,
                        date=date,
                        shift="Day",
                        machine=machine,
                        created_by=request.user,
                        status="Pending",
                        is_loss=is_loss,
                        batches=1,
                        planning=plan
                    )
                    for row in materials:
                        raw = RawMaterial.objects.get(id=row["id"])
                        FormulationItem.objects.create(
                            formulation=form,
                            raw_material=raw,
                            weight=float(row["weight"])
                        )
                    created_count += 1
            elif len(plan_ids) == 1:
                # Single plan submission
                single_plan = Planning.objects.filter(id=plan_ids[0]).first() if plan_ids[0] else None
                form = Formulation.objects.create(
                    name=name,
                    date=date,
                    shift="Day",
                    machine=machine,
                    created_by=request.user,
                    status="Pending",
                    is_loss=is_loss,
                    batches=1,
                    planning=single_plan
                )
                for row in materials:
                    raw = RawMaterial.objects.get(id=row["id"])
                    FormulationItem.objects.create(
                        formulation=form,
                        raw_material=raw,
                        weight=float(row["weight"])
                    )
                created_count = 1
            else:
                form = Formulation.objects.create(
                    name=name,
                    date=date,
                    shift="Day",
                    machine=machine,
                    created_by=request.user,
                    status="Pending",
                    is_loss=is_loss,
                    batches=1,
                    planning=None
                )

                for row in materials:
                    raw = RawMaterial.objects.get(id=row["id"])
                    FormulationItem.objects.create(
                        formulation=form,
                        raw_material=raw,
                        weight=float(row["weight"])
                    )
                created_count = 1

            audit_logger.info(f"Frame formulation(s) created: {name} ({created_count} formulation) by user {request.user.username}, Machine: {machine.name}")
            business_logger.info(f"Frame formulation(s) {name} added - Materials: {len(materials)}")
            messages.success(request, f"{created_count} Formulation(s) Added Successfully.")
        except Exception as e:
            messages.error(request, f"Error : {str(e)}")
        return redirect("frameline_page")

    # GET - render the page with all needed context
    machines = Machine.objects.filter(category="Frame" ,is_active=True)
    materials_qs = RawMaterial.objects.filter(is_active=True).order_by("name")

    category_sequence = [
        'Raw Material',
        'PVC Resin',
        'Marble Powder',
        'Processing Head (Processing Aids)',
        'Stabilizer',
        'CS',
        'CPE',
        'Lubrication',
        'Internal Lubricant',
        'External Lubricant',
        'Wax',
        'PE Wax',
        'Foaming White',
        'Foaming Yellow',
        'Blister',
        'Scrap',
        'Pigment',
        'Stearic Acid',
        'Other',
    ]

    materials_grouped = OrderedDict()
    for cat in category_sequence:
        materials_grouped[cat] = materials_qs.filter(category=cat)

    # # Get remaining planning records (ones not yet formulated)
    # frame_plans = Planning.objects.filter(category="Frame", formulation__isnull=True).select_related("size", "length").order_by(
    #     Case(
    #         When(remark__iexact="Urgent", then=Value(0)),
    #         When(remark__iexact="Order", then=Value(1)),
    #         When(remark__iexact="Stock", then=Value(2)),
    #         default=Value(2),
    #         output_field=IntegerField(),
    #     )
    # )
    
    # Get formulations created from frameline (all pending Frame formulations)
    formulations = Formulation.objects.filter(
        status="Pending",
        machine__category="Frame"
    ).select_related("machine").prefetch_related("items__raw_material").order_by("-id")

    # context = {
    #     "frame_plans": frame_plans,
    #     "formulations": formulations,
    #     "machines": machines,
    #     "materials_grouped": materials_grouped,
    # }
    # return render(request, "frameline.html", context)
    
# GET - render the page with all needed context
    machine_filter = request.GET.get('machine')
    frame_plans = Planning.objects.filter(category="Frame").select_related("size", "length", "machine")    
    # Apply machine filter
    if machine_filter:
        frame_plans = frame_plans.filter(machine__name=machine_filter)
    
    frame_plans = frame_plans.order_by(
        Case(
            When(remark__iexact="Urgent", then=Value(0)),
            When(remark__iexact="Order", then=Value(1)),
            When(remark__iexact="Stock", then=Value(2)),
            default=Value(2),
            output_field=IntegerField(),
        )
    )

    # Group the plans
    groups = defaultdict(list)
    for p in frame_plans:
        # Calculate tweight for each plan item
        p.tweight = max(0.0, float(p.weight) * float(p.quantity))
        
        date_str = p.date.strftime("%Y-%m-%d") if p.date else "-"
        machine_str = p.machine.name if p.machine else "-"
        size_str = p.size.size if p.size else "-"
        color_str = str(p.color) if p.color else "-"
        
        key = (date_str, machine_str, size_str, color_str)
        groups[key].append(p)

    grouped_frame_plans = []
    for key, items in groups.items():
        # Calculate totals for the group
        total_quantity = sum(item.quantity for item in items)
        total_tweight = sum(item.tweight for item in items)
        expected_days = (total_tweight / 2500) if total_tweight else 0
        
        grouped_frame_plans.append({
            'title': f"Date: {key[0]} | Machine: {key[1]} | Size: {key[2]} | Color: {key[3]}",
            'items': items,
            'total_quantity': total_quantity,
            'total_tweight': total_tweight,
            'expected_days': expected_days,
        })

    # Pass 'grouped_frame_plans' to the context instead of 'frame_plans'
    context = {
        "grouped_frame_plans": grouped_frame_plans,
        "formulations": formulations,
        "machines": machines,
        "machine_filter": machine_filter,
        "materials_grouped": materials_grouped,
    }
    return render(request, "frameline.html", context)


@login_required
@module_required('formulation')
def batch_issue_page(request):

    # Handle group formulation POST (set 1 batch for all pending formulations)
    if request.method == "POST" and request.POST.get("action") == "group_formulation":
        group_name = (request.POST.get("group_name") or "").strip()
        if not group_name:
            messages.error(request, "Group name is required.")
            return redirect("batch_issue_page")

        updated = Formulation.objects.filter(status="Pending").update(
            batches=1,
            group_name=group_name,
        )
        messages.success(request, f"Grouped formulation updated: {updated} pending formulation(s) set to 1 batch.")
        return redirect("batch_issue_page")

    # Handle batch approval POST
    if request.method == "POST":
        # Existing per-formulation batch update
        formulation_ids = request.POST.getlist("formulation_ids")

        # New: batch update for complete group (all pending formulations)
        group_id = request.POST.get("group_id")

        approved_count = 0

        if group_id:
            qs = Formulation.objects.filter(group_id=group_id)
            if not qs.exists():
                messages.error(request, "Invalid group for batch update.")
            else:
                # Use single input value (batch_group_qty) applied to ALL in group
                batch_qty_raw = request.POST.get("batch_group_qty")
                try:
                    batch_qty = int(batch_qty_raw or 0)
                except ValueError:
                    batch_qty = 0

                if batch_qty <= 0:
                    messages.error(request, "Batch quantity must be > 0.")
                else:
                    for obj in qs:
                        obj.batches = batch_qty
                        # keep status as-is; only setting batches
                        obj.save()
                        approved_count += 1

        elif formulation_ids:
            for fid in formulation_ids:
                try:
                    obj = Formulation.objects.get(id=fid, status="Approved")
                    batch_key = f"batch_{fid}"
                    batch_val = request.POST.get(batch_key)
                    if batch_val:
                        qty = int(batch_val)
                        if qty > 0:
                            obj.batches = qty
                            obj.save()
                            approved_count += 1
                except Formulation.DoesNotExist:
                    continue

        if approved_count:
            messages.success(request, f"{approved_count} batch(es) updated successfully.")
        else:
            messages.warning(request, "No valid formulations were updated.")

        return redirect("batch_issue_page")

    # GET: display data
    pending_formulations = Formulation.objects.filter(status="Pending").select_related("machine").prefetch_related("items__raw_material").order_by("-date", "-id")
    approved_formulations = Formulation.objects.filter(status="Approved").select_related("machine").prefetch_related("items__raw_material").order_by("-date", "-id")
    all_formulations = list(pending_formulations) + list(approved_formulations)

    # door_plans = Planning.objects.filter(category__in=["Door", "Sheet"], formulation__isnull=True).order_by(
    #     Case(
    #         When(remark__iexact="Urgent", then=Value(0)),
    #         When(remark__iexact="Order", then=Value(1)),
    #         When(remark__iexact="Stock", then=Value(2)),
    #         default=Value(2),
    #         output_field=IntegerField(),
    #     )
    # )

    machines = Machine.objects.filter(category = "Door" ,is_active=True)
    materials_qs = RawMaterial.objects.filter(is_active=True).order_by("name")

    category_sequence = [
        'Raw Material',
        'PVC Resin',
        'Marble Powder',
        'Processing Head (Processing Aids)',
        'Stabilizer',
        'CS',
        'CPE',
        'Lubrication',
        'Internal Lubricant',
        'External Lubricant',
        'Wax',
        'PE Wax',
        'Foaming White',
        'Foaming Yellow',
        'Blister',
        'Scrap',
        'Pigment',
        'Stearic Acid',
        'Other',
    ]

    materials_grouped = OrderedDict()
    for cat in category_sequence:
        materials_grouped[cat] = materials_qs.filter(category=cat)

    # context = {
    #     "pending_formulations": pending_formulations,
    #     "approved_formulations": approved_formulations,
    #     "all_formulations": all_formulations,
    #     "door_plans": door_plans,
    #     "machines": machines,
    #     "materials_grouped": materials_grouped,
    # }
    # return render(request, "doorline.html", context)
    
    machine_filter = request.GET.get('machine')
    door_plans = Planning.objects.filter(category__in=["Door", "Sheet"]).select_related("thickness", "height", "width", "density", "machine")    
    # Apply machine filter
    if machine_filter:
        door_plans = door_plans.filter(machine__name=machine_filter)
    
    door_plans = door_plans.order_by(
        Case(
            When(remark__iexact="Urgent", then=Value(0)),
            When(remark__iexact="Order", then=Value(1)),
            When(remark__iexact="Stock", then=Value(2)),
            default=Value(2),
            output_field=IntegerField(),
        )
    )

    groups = defaultdict(list)
    for p in door_plans:
        # Calculate tweight for each plan item
        p.tweight = max(0.0, float(p.weight) * float(p.quantity))
        
        date_str = p.date.strftime("%Y-%m-%d") if p.date else "-"
        machine_str = p.machine.name if p.machine else "-"
        thickness_str = f"{p.thickness.thickness} mm" if p.thickness else "-"
        density_str = str(p.density) if p.density else "-"
        color_str = str(p.color) if p.color else "-"
        
        key = (date_str, machine_str, thickness_str, density_str, color_str)
        groups[key].append(p)

    grouped_door_plans = []
    for key, items in groups.items():
        # Calculate totals for the group
        total_quantity = sum(item.quantity for item in items)
        total_tweight = sum(item.tweight for item in items)
        expected_days = (total_tweight / 5000) if total_tweight else 0
        
        grouped_door_plans.append({
            'title': f"Date: {key[0]} | Machine: {key[1]} | Thickness: {key[2]} | Density: {key[3]} | Color: {key[4]}",
            'items': items,
            'total_quantity': total_quantity,
            'total_tweight': total_tweight,
            'expected_days': expected_days,
        })

    # Pass 'grouped_door_plans' to the context instead of 'door_plans'
    context = {
        "grouped_door_plans": grouped_door_plans,
        "formulations": all_formulations,
        "machines": machines,
        "machine_filter": machine_filter,
        "materials_grouped": materials_grouped,
    }
    return render(request, "doorline.html", context)


@login_required
@module_required('batch')
def add_batch(request, id):

    obj = get_object_or_404(Formulation, id=id)

    if request.method == "POST":
        qty = int(request.POST.get("batches") or 0)

        obj.batches = qty
        obj.save()

        messages.success(request, "Batch Saved.")

    return redirect("formulation_page")


@admin_only
def approve_formulation(request, id):

    obj = get_object_or_404(Formulation, id=id)

    # Prevent double approval
    if obj.status == "Approved":
        messages.warning(request, "Already approved.")
        return redirect("formulation_page")

    try:

        batch_qty = obj.batches or 0

        if batch_qty <= 0:
            messages.error(request, "Batch is 0, cannot approve.")
            return redirect("formulation_page")

        # ==============================
        # DEDUCT STOCK (ALL FORMULATIONS)
        # ==============================
        for item in obj.items.all():

            total_weight = float(item.weight) * float(batch_qty)

            rm = item.raw_material

            # safety check (optional but recommended)
            if rm.current_stock < total_weight:
                messages.error(request, f"Not enough stock for {rm.name}")
                return redirect("formulation_page")

            # deduct stock
            rm.current_stock -= total_weight
            rm.save()

            # log stock movement
            RawMaterialStock.objects.create(
                raw_material=rm,
                quantity=total_weight,
                movement_type='OUT',
                formulation=obj
            )

        # ==============================
        # MARK APPROVED
        # ==============================
        obj.status = "Approved"
        obj.approved_at = now()
        obj.stock_deducted = True
        obj.save()

        audit_logger.info(f"Formulation {obj.id} approved by user {request.user.username} - Name: {obj.name}, Batches: {obj.batches}")
        business_logger.info(f"Formulation {obj.id} approved - Stock deducted for {len(obj.items.all())} materials")
        messages.success(request, "Approved & Stock Deducted Successfully.")

    except Exception as e:
        messages.error(request, f"Error approving formulation: {str(e)}")

    return redirect("formulation_page")


@module_required('formulation')
def delete_formulation(request, id):

    obj = get_object_or_404(Formulation, id=id)
    formulation_name = obj.name

    if obj.status == "Approved":
        messages.error(request, "Approved formulation cannot be deleted.")
        return redirect("formulation_page")

    obj.delete()

    audit_logger.info(f"Formulation deleted: {formulation_name} (ID: {id}) by user {request.user.username}")
    messages.success(request, "Deleted Successfully.")

    return redirect("formulation_page")


@module_required('formulation')
def edit_loss_formula(request):

    loss_id = request.POST.get("form_id") or request.GET.get("id")

    if loss_id:
            latest_loss = Formulation.objects.filter(id=loss_id, is_loss=True).first()
    else:
            latest_loss = Formulation.objects.filter(is_loss=True).order_by("-id").first()

    if not latest_loss:
        latest_loss = Formulation.objects.create(
            name="Production Loss",
            date=now().date(),
            shift="",
            machine=None,
            created_by=request.user,
            batches=0,
            status="Pending",
            is_loss=True
        )

    if latest_loss.items.count() == 0:
        default_materials = [
            ("CALCIUM CARBONATE / MARBLE POWDER", 15),
            ("BAEROPAN SMS 305NI", 1.5),
            ("PV WAX", 0.4),
            ("STEARIC ACID", 0.4),
            ("PVC Resin", 10),
        ]

        for query, weight in default_materials:
            rm = RawMaterial.objects.filter(name__icontains=query).first()
            if not rm:
                rm = RawMaterial.objects.filter(name__iexact=query).first()
            if not rm:
                # Try partial match
                words = query.split()
                for word in words:
                    if len(word) > 3:
                        rm = RawMaterial.objects.filter(name__icontains=word).first()
                        if rm:
                            break
            if rm:
                FormulationItem.objects.create(
                    formulation=latest_loss,
                    raw_material=rm,
                    weight=weight
                )

    if request.method == "POST":
        action = request.POST.get("action", "save_template")

        if action == "save_template":
            raw_ids = request.POST.getlist("raw_id[]")
            weights = request.POST.getlist("weight[]")

            messages.info(request, f"Received {len(raw_ids)} raw_ids and {len(weights)} weights")

            latest_loss.items.all().delete()
            for raw_id, weight in zip(raw_ids, weights):
                raw_id = (raw_id or "").strip()
                weight = (weight or "").strip()
                if raw_id and weight:
                    try:
                        FormulationItem.objects.create(
                            formulation=latest_loss,
                            raw_material_id=int(raw_id),
                            weight=float(weight)
                        )
                    except (ValueError, TypeError) as e:
                        messages.error(request, f"Error saving item {raw_id}: {str(e)}")
                        continue

            messages.success(request, "Production Loss formula updated successfully.")
            return redirect("formulation_page")

        if action == "create_entry":
            date_str = request.POST.get("date")
            batch = int(request.POST.get("batch", 0))

            if not date_str:
                messages.error(request, "Date is required.")
                return redirect("edit_loss_formula")

            date = timezone.datetime.strptime(date_str, "%Y-%m-%d").date()

            new_loss = Formulation.objects.create(
                name="Production Loss",
                date=date,
                shift="",
                machine=None,
                created_by=request.user,
                batches=batch,
                status="Pending",
                is_loss=True
            )

            for item in latest_loss.items.all():
                FormulationItem.objects.create(
                    formulation=new_loss,
                    raw_material=item.raw_material,
                    weight=item.weight
                )

            messages.success(request, f"New Production Loss entry added for {date}.")
            return redirect("formulation_page")

    materials = RawMaterial.objects.filter(is_active=True)

    return render(request, "edit_loss_formula.html", {
    "loss": latest_loss,
    "materials": materials
})


@module_required('formulation')
def edit_formulation(request, id):
    formulation = get_object_or_404(Formulation, id=id)

    if request.method == "POST":
        try:
            # materials from JS
            items = json.loads(request.POST.get("materials", "[]"))

            # update top fields
            formulation.date = request.POST.get("date")
            formulation.machine_id = request.POST.get("machine")
            formulation.shift = request.POST.get("shift")
            formulation.name = request.POST.get("name")
            formulation.save()

            # remove old materials
            formulation.items.all().delete()

            # add new materials
            for item in items:
                if float(item["weight"]) > 0:
                    FormulationItem.objects.create(
                        formulation=formulation,
                        raw_material_id=item["id"],
                        weight=item["weight"]
                    )

            messages.success(request, "Formulation updated successfully")

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

    return redirect("formulation_page")


@module_required('production')
def weight_sheet(request, pk=None):
    mode = request.GET.get("mode", "list")

    # ==========================================
    # VIEW MODE => no extra blank row
    # ==========================================
    extra_rows = 0 if mode == "view" else 1

    # ==========================================
    # FORMSET (Door + Frame Both Fields)
    # ==========================================
    RowFormSet = modelformset_factory(
        WeightSheetRow,
        fields=(
            "color",
            "weight_per_piece",

            # door
            "thickness",
            "density",
            "height",
            "width",

            # frame
            "size",
            "length",
        ),
        extra=extra_rows,
        can_delete=True
    )

    # ==========================================
    # LIST PAGE
    # ==========================================
    if mode == "list":
        date = request.GET.get("date")

        sheets = WeightSheet.objects.all().order_by("-date", "-id")

        if date:
            sheets = sheets.filter(date=parse_date(date))

        return render(request, "weight_sheet.html", {
            "mode": "list",
            "sheets": sheets,
            "selected_date": date
        })

    # ==========================================
    # CREATE
    # ==========================================
    if mode == "create":

        formset = RowFormSet(
            queryset=WeightSheetRow.objects.none()
        )

        if request.method == "POST":

            formset = RowFormSet(request.POST)

            if formset.is_valid():

                sheet = WeightSheet.objects.create(
                    date=request.POST.get("date"),
                    machine_id=request.POST.get("machine"),
                    operator=request.user,
                    shift=request.POST.get("shift"),
                    category=request.POST.get("category"),
                )

                instances = formset.save(commit=False)

                for obj in instances:
                    obj.sheet = sheet
                    obj.save()

                # deleted rows if any
                for obj in formset.deleted_objects:
                    if obj.pk:
                        obj.delete()

                return redirect("/weightsheets/?mode=list")

        return render(request, "weight_sheet.html", {
            "mode": "create",
            "formset": formset,
            "machines": Machine.objects.filter(is_active=True),
            "sizes": Size.objects.filter(is_active=True),

        })

    # ==========================================
    # EDIT / VIEW
    # ==========================================
    sheet = get_object_or_404(WeightSheet, pk=pk)

    formset = RowFormSet(
        queryset=sheet.rows.all()
    )

    # ==========================================
    # EDIT SAVE
    # ==========================================
    if request.method == "POST" and mode == "edit":

        formset = RowFormSet(
            request.POST,
            queryset=sheet.rows.all()
        )

        if formset.is_valid():

            # -----------------------------
            # TOP SECTION UPDATE
            # -----------------------------
            sheet.date = request.POST.get("date")
            sheet.machine_id = request.POST.get("machine")
            sheet.shift = request.POST.get("shift")
            sheet.category = request.POST.get("category")
            sheet.save()

            # -----------------------------
            # SAVE ROWS
            # -----------------------------
            instances = formset.save(commit=False)

            for obj in instances:
                obj.sheet = sheet
                obj.save()

            # -----------------------------
            # DELETE ROWS
            # -----------------------------
            for obj in formset.deleted_objects:
                obj.delete()

            return redirect("/weightsheets/?mode=list")

    # ==========================================
    # FINAL RENDER
    # ==========================================
    return render(request, "weight_sheet.html", {
        "mode": mode,      # list/create/edit/view
        "sheet": sheet,
        "formset": formset,
        "machines": Machine.objects.filter(is_active=True),
        "sizes": Size.objects.filter(is_active=True),
    })


# -------------------------
# JSON API for weight sheets
# -------------------------
@login_required
def weight_sheet_api(request):
    """POST: create or update a WeightSheet for a Production using JSON payload
       payload: { production_id: int, rows: [number,...] }
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    production_id = data.get('production_id')
    rows = data.get('rows', [])

    if not production_id:
        return JsonResponse({'error': 'production_id required'}, status=400)

    try:
        prod = Production.objects.get(id=production_id)
    except Production.DoesNotExist:
        return JsonResponse({'error': 'Production not found'}, status=404)

    # if production already has a sheet, update it; otherwise create new
    sheet = prod.weight_sheet
    if not sheet:
        sheet = WeightSheet.objects.create(
            date=prod.created_at.date() if prod.created_at else timezone.now().date(),
            machine=prod.machine,
            operator=prod.operator,
            shift=prod.shift,
            category=prod.category,
        )
        prod.weight_sheet = sheet
        prod.save()
    else:
        # clear existing rows
        sheet.rows.all().delete()

    created_rows = []
    for w in rows:
        try:
            val = float(w)
        except Exception:
            continue
        row = WeightSheetRow.objects.create(sheet=sheet, weight_per_piece=val)
        created_rows.append({'id': row.id, 'weight_per_piece': row.weight_per_piece})

    return JsonResponse({'success': True, 'sheet_id': sheet.id, 'rows': created_rows})


@login_required
def weight_sheet_detail_api(request, pk):
    """GET: return weight sheet rows as JSON"""
    try:
        sheet = WeightSheet.objects.get(pk=pk)
    except WeightSheet.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)

    rows = list(sheet.rows.all().values('id', 'weight_per_piece'))
    return JsonResponse({'id': sheet.id, 'date': sheet.date.isoformat(), 'rows': rows})
#         if valid_weights:
#             avg_weight = sum(valid_weights) / len(valid_weights)

#             # SAVE INTO PRODUCTION
#             prod.weight_per_piece = avg_weight
#             prod.save()

#             messages.success(request, "Weight updated successfully")
#             return redirect(f"/production/?category={prod.category}")

#         messages.error(request, "No valid weights entered")

#     return render(request, "weight_sheet.html", {
#         "prod": prod
#     })

@module_required('dashboard')
def dashboard_v2(request):

    # ---------------- DATE RANGE FILTER ----------------
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    machine_name = request.GET.get("machine")

    if from_date and to_date:
        from_date = datetime.strptime(from_date, "%Y-%m-%d").date()
        to_date = datetime.strptime(to_date, "%Y-%m-%d").date()
    else:
        to_date = timezone.now().date() - timedelta(days=1)
        from_date = to_date

    productions = Production.objects.filter(
        created_at__date__range=[from_date, to_date]
    )

    if machine_name:
        productions = productions.filter(machine__name=machine_name)

    machines = Machine.objects.all()[:6]


    # ---------------- TOTALS ----------------
    total_overallproduction = sum(p.overall_production for p in productions)
    total_sidepatti = sum(p.sidepatti for p in productions)
    total_linesetting = sum(p.linesetting for p in productions)
    total_scrap = total_sidepatti + total_linesetting
    total_rejection_weight = sum(p.rejection_weight for p in productions)
    total_production = sum(p.actual_production for p in productions) + total_sidepatti + total_rejection_weight

    # ---------------- TARGET CALCULATION ----------------
    MACHINE_TARGETS = {
        "L-1": 2000,
        "L-2": 2000,
        "L-3": 2000,
        "L-4": 5000,
        "L-5": 6000,
        "L-6": 6000,
    }

    total_days = (to_date - from_date).days + 1

    total_target = 0
    for m in machines:
        daily_target = MACHINE_TARGETS.get(m.name, 0)
        total_target += daily_target * total_days

    # ---------------- CONSUMPTION ----------------
    # raw_in = RawMaterialStock.objects.filter(
    #     created_at__date__range=[from_date, to_date],
    #     movement_type="IN"
    # ).aggregate(total=Sum("quantity"))["total"] or 0

    raw_out = FormulationItem.objects.filter(
                formulation__status="Approved",
                formulation__date__range=[from_date, to_date]
            ).aggregate(
                total=Sum(F("weight") * F("formulation__batches"))
            )["total"] or 0

    consumption = raw_out  # actual used material

    total_production_loss = FormulationItem.objects.filter(
                formulation__is_loss=True,
                formulation__date__range=[from_date, to_date]
            ).aggregate(
                total=Sum(F("weight") * F("formulation__batches"))
            )["total"] or 0

    # ---------------- KPIs ----------------
    kpis = {
        "actual_production": round(total_production, 2),
        "overall_production": round(total_overallproduction, 2),

        "side_patti": round(total_sidepatti, 2),
        "scrap": round(total_linesetting, 2),

        "target_achieved": round((total_production / total_target) * 100, 2) if total_target else 0,

        "scrap_percent": round((total_scrap / total_production) * 100, 2) if total_production else 0,

        "consumption": consumption,

        "consumption_percent": round((consumption / total_production) * 100, 2) if total_production else 0,

        "consumption_vs_production": round(consumption - total_production - total_production_loss, 2),

        "rejection_percent": round((total_rejection_weight / total_production) * 100, 2) if total_production else 0,
        "rejection": total_rejection_weight ,
    }

    # ---------------- CHART ----------------
    chart_data = []
    for m in machines:
        day = productions.filter(machine=m, shift="Day").aggregate(total=Sum("quantity"))["total"] or 0
        night = productions.filter(machine=m, shift="Night").aggregate(total=Sum("quantity"))["total"] or 0

        chart_data.append({
            "machine": m.name,
            "day": day,
            "night": night
        })

    # ---------------- PRODUCTION vs SCRAP CHART ----------------
    scrap_chart_data = []

    for m in machines:
        day_prod = productions.filter(machine=m, shift="Day").aggregate(
            total=Sum(ExpressionWrapper(
            F("quantity") * F("weight_per_piece"),
            output_field=FloatField() ))
        )["total"] or 0

        night_prod = productions.filter(machine=m, shift="Night").aggregate(
            total=Sum(ExpressionWrapper(
            F("quantity") * F("weight_per_piece"),
            output_field=FloatField() ))
        )["total"] or 0

        day_scrap = productions.filter(machine=m, shift="Day").aggregate(
            total=Sum("linesetting")
        )["total"] or 0

        night_scrap = productions.filter(machine=m, shift="Night").aggregate(
            total=Sum("linesetting")
        )["total"] or 0

        scrap_chart_data.append({
            "machine": m.name,
            "day_prod": day_prod,
            "night_prod": night_prod,
            "day_scrap": day_scrap,
            "night_scrap": night_scrap,
        })

    # ---------------- MACHINE STATUS ----------------
    
    machine_status = []
    for m in machines:
        m_prods = productions.filter(machine=m)
        
        # Calculate production (quantity * weight_per_piece)
        production = m_prods.aggregate(
            total=Sum(ExpressionWrapper(
                F("quantity") * F("weight_per_piece") + F("sidepatti") ,
                output_field=FloatField()
            ))
        )["total"] or 0
        
        # Calculate consumption from formulations for this machine
        consumption = FormulationItem.objects.filter(
            formulation__machine=m,
            formulation__status="Approved",
            formulation__date__range=[from_date, to_date]
        ).aggregate(
            total=Sum(ExpressionWrapper(
                F("weight") * F("formulation__batches"),
                output_field=FloatField()
            ))
        )["total"] or 0

        total_production_loss = FormulationItem.objects.filter(
                formulation__is_loss=True,
                formulation__date__range=[from_date, to_date]
            ).aggregate(
                total=Sum(F("weight") * F("formulation__batches"))
            )["total"] or 0
        
        difference = round(consumption - production - total_production_loss, 2)

        machine_status.append({
            "machine": m.name,
            "status": "RUNNING" if m_prods.exists() else "In-active",
            "operator": m_prods.first().operator.full_name if m_prods.exists() else "-",
            "hours": round(m_prods.count() * 1.5, 2),
            "target_achieved": round((production) / (MACHINE_TARGETS.get(m.name, 0) * total_days) * 100, 2) if MACHINE_TARGETS.get(m.name, 0) else 0,
            "production": round(production, 2),
            "consumption": round(consumption, 2),
            "difference": difference
        })

    # ---------------- INVENTORY ----------------
    raw_materials = RawMaterial.objects.all()
    inventory = []

    for r in raw_materials:
        if r.current_stock < 500:
            inventory.append({
                "name": r.name,
                "stock": r.current_stock
            })

    # ---------------- STOCK REMAINING ALERTS ----------------
    stock_1_day = []
    stock_7_days = []
    stock_10_days = []

    for rm in RawMaterial.objects.filter(is_active=True):
        if rm.one_day_requirement and rm.one_day_requirement > 0:
            if rm.current_stock < rm.one_day_requirement * 1:
                stock_1_day.append({
                    "name": rm.name,
                    "category": rm.category,
                    "current_stock": rm.current_stock or 0,
                    "one_day_req": rm.one_day_requirement or 0,
                })
            if rm.current_stock < rm.one_day_requirement * 7:
                stock_7_days.append({
                    "name": rm.name,
                    "category": rm.category,
                    "current_stock": rm.current_stock or 0,
                    "one_day_req": rm.one_day_requirement or 0,
                })
            if rm.current_stock < rm.one_day_requirement * 10:
                stock_10_days.append({
                    "name": rm.name,
                    "category": rm.category,
                    "current_stock": rm.current_stock or 0,
                    "one_day_req": rm.one_day_requirement or 0,
                })

    return render(request, "dashboard_v2.html", {
        "kpis": kpis,
        "chart_data": chart_data,
        "scrap_chart_data": scrap_chart_data,
        "machine_status": machine_status,
        "inventory": inventory,
        "from_date": from_date,
        "to_date": to_date,
        "selected_machine": machine_name,
        "machines": machines,
        "stock_1_day": stock_1_day,
        "stock_7_days": stock_7_days,
        "stock_10_days": stock_10_days,
    })


def mobile_dashboard(request):
    """Mobile wrapper for dashboard_v2"""
    return dashboard_v2(request)


def mobile_stock(request):
    """Mobile wrapper for stock_view"""
    return stock_view(request)


def mobile_raw_materials(request):
    """Mobile wrapper for raw_material_list"""
    return raw_material_list(request)


def mobile_dispatch(request):
    """Mobile wrapper for dispatch"""
    return dispatch(request)


@admin_only
def roles_responsibilities(request):
    """Admin-only view to manage role permissions"""
    if request.method == 'POST':
        # Save permissions from checkbox data
        try:
            for role, role_label in RolePermission.ROLES:
                for module, module_label in RolePermission.MODULES:
                    # Get checkbox value (will be 'on' if checked, or not in POST if unchecked)
                    checkbox_name = f'{role}_{module}'
                    has_permission = checkbox_name in request.POST
                    
                    # Update or create permission
                    RolePermission.objects.update_or_create(
                        role=role,
                        module=module,
                        defaults={'has_permission': has_permission}
                    )
            
            messages.success(request, "Permissions updated successfully!")
            return redirect('roles_responsibilities')
        except Exception as e:
            messages.error(request, f"Error updating permissions: {str(e)}")
    
    # GET: Fetch all permissions and organize by role
    permissions_by_role = {}
    for role, role_label in RolePermission.ROLES:
        permissions_by_role[role] = RolePermission.get_role_permissions(role)
    
    modules = [m for m in RolePermission.MODULES]
    
    context = {
        'roles': RolePermission.ROLES,
        'modules': modules,
        'permissions_by_role': permissions_by_role,
    }
    
    return render(request, 'roles_responsibilities.html', context)


# rejection handling
def update_production(request, pk):
    if request.method == "POST":

        productions = Production.objects.filter(batch_id=pk)

        quantities = request.POST.getlist("quantity")
        rejections = request.POST.getlist("rejection")
        unfinished_qtys = request.POST.getlist("unfinished_qty")
        scrap_qtys = request.POST.getlist("scrap_qty")

        for i, prod in enumerate(productions):

            qty = float(quantities[i] or 0)
            rejection = rejections[i]
            unfinished_qty = float(unfinished_qtys[i] or 0)
            scrap_qty = float(scrap_qtys[i] or 0)

            prod.quantity = qty
            prod.save()

            # 🔥 UNFINISHED LOG
            if rejection == "unfinished" and unfinished_qty > 0:
                UnfinishedProduction.objects.create(
                    product=prod.product,
                    production=prod,
                    quantity=unfinished_qty,
                    date=date.today()
                )

            # 🔥 SCRAP LOG
            if rejection == "scrap" and scrap_qty > 0:
                prod.linesetting += scrap_qty
                prod.save()

                ScrapLog.objects.create(
                    production=prod,
                    quantity=scrap_qty
                )

        return redirect("production_list")
    

@login_required
@login_required
@module_required('stock')
def unfinished_list(request):
    data = UnfinishedProduction.objects.select_related(
        "product",
        "product__width",
        "product__height",
        "product__thickness",
        "product__density",
        "product__density_name",
        "product__color",
        "production",
    ).order_by("-date")

    return render(request, "unfinished_list.html", {
        "data": data,
        "widths": Width.objects.filter(is_active=True).order_by("width"),
        "heights": Height.objects.filter(is_active=True).order_by("height"),
    })


@login_required
@module_required('stock')
def add_to_scrap(request, item_id):
    if request.method != "POST":
        return redirect("unfinished_list")

    item = get_object_or_404(
        UnfinishedProduction.objects.select_related(
            "product",
            "product__width",
            "product__height",
            "product__thickness",
            "product__density",
            "product__density_name",
            "product__color",
            "production",
        ),
        id=item_id,
    )

    product = item.product
    scrap_qty = float(request.POST.get("scrap_quantity") or 0)

    if scrap_qty <= 0:
        messages.error(request, "Scrap quantity must be greater than zero.")
        return redirect("unfinished_list")

    if scrap_qty > item.quantity:
        messages.error(request, "Scrap quantity cannot be greater than unfinished quantity.")
        return redirect("unfinished_list")

    # Calculate weight for the scrap quantity
    weight = 0
    if item.production and item.production.weight_per_piece:
        weight = float(item.production.weight_per_piece) * scrap_qty
    else:
        weight = _door_piece_weight(
            product.thickness, product.width, product.height, product.density
        ) * scrap_qty

    remaining_qty = max(float(item.quantity) - scrap_qty, 0)

    with transaction.atomic():
        # Create ScrapLog if there's a related production
        if item.production:
            if weight > 0 or scrap_qty > 0:
                ScrapLog.objects.create(
                    production=item.production,
                    quantity=scrap_qty,
                    weight=weight,
                    date=timezone.now().date(),
                )

            # Update production rejection status
            item.production.rejected_quantity = remaining_qty
            item.production.rejection_status = "Scrap" if remaining_qty > 0 else "OK"
            item.production.save(update_fields=["rejected_quantity", "rejection_status"])

        # Update or delete the unfinished item
        if remaining_qty > 0:
            item.quantity = remaining_qty
            item.save(update_fields=["quantity"])
        else:
            item.delete()

    messages.success(
        request,
        f"Added {format_float(scrap_qty)} item(s) to scrap. Weight: {weight:.2f} kg."
    )
    return redirect("unfinished_list")


def _door_piece_weight(thickness_obj, width_obj, height_obj, density_obj):
    if not (thickness_obj and width_obj and height_obj and density_obj):
        return 0

    h = float(height_obj.height) * 0.0254
    w = float(width_obj.width) * 0.0254
    t = float(thickness_obj.thickness) / 1000
    density = float(density_obj.density)
    return h * w * t * density


@login_required
@module_required('stock')
def convert_unfinished(request, item_id):
    if request.method != "POST":
        return redirect("unfinished_list")

    item = get_object_or_404(
        UnfinishedProduction.objects.select_related(
            "product",
            "product__width",
            "product__height",
            "product__thickness",
            "product__density",
            "product__density_name",
            "product__color",
            "production",
        ),
        id=item_id,
    )

    product = item.product
    if product.category not in ["Door", "Sheet"]:
        messages.error(request, "Only door/sheet unfinished products can be converted from this page.")
        return redirect("unfinished_list")

    width_id = request.POST.get("width")
    height_id = request.POST.get("height")
    stamp = request.POST.get("stamp", "")
    convert_qty = float(request.POST.get("convert_quantity") or 0)

    if not width_id or not height_id:
        messages.error(request, "Please select both width and height.")
        return redirect("unfinished_list")

    if convert_qty <= 0:
        messages.error(request, "Converted quantity must be greater than zero.")
        return redirect("unfinished_list")

    if convert_qty > item.quantity:
        messages.error(request, "Converted quantity cannot be greater than unfinished quantity.")
        return redirect("unfinished_list")

    new_width = get_object_or_404(Width, id=width_id)
    new_height = get_object_or_404(Height, id=height_id)

    old_piece_weight = (
        float(item.production.weight_per_piece)
        if item.production and item.production.weight_per_piece
        else _door_piece_weight(product.thickness, product.width, product.height, product.density)
    )
    new_piece_weight = _door_piece_weight(product.thickness, new_width, new_height, product.density)

    if new_piece_weight <= 0:
        messages.error(request, "Unable to calculate new product weight for the selected size.")
        return redirect("unfinished_list")

    with transaction.atomic():
        converted_product, _ = Product.objects.get_or_create(
            category=product.category,
            color=product.color,
            thickness=product.thickness,
            density=product.density,
            density_name=product.density_name,
            width=new_width,
            height=new_height,
            stamp=stamp or product.stamp or "",
            defaults={
                "size": None,
                "length": None,
            },
        )

        Stock.objects.create(
            product=converted_product,
            operator=request.user,
            quantity=convert_qty,
            movement_type="IN",
        )

        # Reduce unfinished quantity or delete if fully converted
        remaining_qty = max(float(item.quantity) - convert_qty, 0)
        if remaining_qty > 0:
            item.quantity = remaining_qty
            item.save(update_fields=["quantity"])
        else:
            item.delete()

    messages.success(
        request,
        f"Converted {format_float(convert_qty)} unfinished item(s) to stock. Remaining unfinished: {format_float(remaining_qty)}."
    )
    return redirect("unfinished_list")

def product_stock_history_json(request, pk):
    history = Stock.objects.filter(product_id=pk).order_by("-created_at")

    data = []
    for h in history:
        data.append({
            "date": h.created_at.strftime("%Y-%m-%d %H:%M"),
            "type": h.movement_type,
            "quantity": h.quantity,
            "operator": h.operator.username if h.operator else ""
        })

    return JsonResponse({"data": data})


#upload formulation from excel
@login_required
@module_required('formulation')
def upload_formulation(request):
    if request.method != "POST":
        return redirect("formulation_page")

    is_loss = request.POST.get("is_loss") == "1"

    file = request.FILES.get("file")
    if not file:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect("formulation_page")

    if not file.name.lower().endswith((".xlsx", ".xls", ".xlsm", ".xltx", ".xltm", ".xlsr")):
        messages.error(request, "Only Excel files are supported (.xlsx, .xls, .xlsr).")
        return redirect("formulation_page")

    try:
        df = pd.read_excel(file)
    except Exception as e:
        messages.error(request, f"Unable to read Excel file: {str(e)}")
        return redirect("formulation_page")

    df.columns = [str(col).strip().upper() for col in df.columns]
    if "RAW MATERIAL IN" not in df.columns:
         df["RAW MATERIAL IN"] = 0
    required_columns = ["DATE", "CHEMICAL NAME"]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        messages.error(request, f"Missing required columns: {', '.join(missing_columns)}.")
        return redirect("formulation_page")

    machine_mapping = {
        "L-1  (DAY)": ("L-1", "Day", ["L-1 BATCH  (DAY)"]),
        "L-2  (DAY)": ("L-2", "Day", ["L-2 BATCH  (DAY)"]),
        "L-3  (DAY)": ("L-3", "Day", ["L-3 BATCH  (DAY)"]),
        "L-4  (DAY)": ("L-4", "Day", ["L-4 BATCH  (DAY)"]),
        "L-5  (DAY)": ("L-5", "Day", ["L-5 BATCH  (DAY)"]),
        "L-6  (DAY)": ("L-6", "Day", ["L-6 BATCH  (DAY)"]),
        "L-1 (NIGHT)": ("L-1", "Night", ["L-1 BATCH (NIGHT)"]),
        "L-2 (NIGHT)": ("L-2", "Night", ["L-2 BATCH (NIGHT)"]),
        "L-3 (NIGHT)": ("L-3", "Night", ["L-3 BATCH (NIGHT)"]),
        "L-4 (NIGHT)": ("L-4", "Night", ["L-4 BATCH (NIGHT)"]),
        "L-5 (NIGHT)": ("L-5", "Night", ["L-5 BATCH (NIGHT)"]),
        "L-6 (NIGHT)": ("L-6", "Night", ["L-6 BATCH (NIGHT)"]),
    }

    def safe_float(value, default=0.0):
        try:
            if pd.isna(value):
                return default
            return float(value)
        except Exception:
            return default

    def safe_int(value, default=0):
        try:
            if pd.isna(value):
                return default
            return int(float(value))
        except Exception:
            return default

    def resolve_machine(machine_label):
        machine = Machine.objects.filter(name__iexact=machine_label).first()
        if machine:
            return machine
        if machine_label == "L-1 and L-2":
            return Machine.objects.filter(name__iexact="L-1").first() or Machine.objects.filter(name__iexact="L-2").first()
        return None

    pending_formulations = {}
    pending_loss_forms = {}
    created_formulations = 0
    created_loss_forms = 0
    warnings = []

    for index, row in df.iterrows():
        raw_date = row.get("DATE")
        raw_name = row.get("CHEMICAL NAME")

        if pd.isna(raw_name) or str(raw_name).strip() == "":
            warnings.append(f"Row {index + 2}: missing CHEMICAL NAME.")
            continue
        raw_name = str(raw_name).strip()

        if pd.isna(raw_date) or str(raw_date).strip() == "":
            warnings.append(f"Row {index + 2}: missing DATE.")
            continue

        if isinstance(raw_date, datetime):
            date_value = raw_date.date()
        elif isinstance(raw_date, date):
            date_value = raw_date
        else:
            date_value = parse_date(str(raw_date).strip())

        if not date_value:
            warnings.append(f"Row {index + 2}: invalid DATE value '{raw_date}'.")
            continue

        raw_material = RawMaterial.objects.filter(name__iexact=raw_name).first()
        if not raw_material:
            warnings.append(f"Row {index + 2}: raw material '{raw_name}' not found.")
            continue

        # ✅ STOCK IN LOGIC (ADD HERE)
        stock_in = safe_float(row.get("RAW MATERIAL IN", 0.0), 0.0)

        if stock_in > 0:
            # Create stock entry (NO extra fields)
            RawMaterialStock.objects.create(
                raw_material=raw_material,
                quantity=stock_in,
                movement_type="IN"
            )

            # Increase current stock
            raw_material.current_stock += stock_in
            raw_material.save()


# till here
        row_has_weight = False

        for machine_col, (machine_label, shift, batch_cols) in machine_mapping.items():
            if machine_col not in df.columns:
                continue

            weight = safe_float(row.get(machine_col), 0.0)
            if weight <= 0:
                continue

            row_has_weight = True
            batch_count = 0
            for batch_col in batch_cols:
                batch_count += safe_int(row.get(batch_col), 0)

            machine = resolve_machine(machine_label)
            if not machine:
                warnings.append(f"Row {index + 2}: machine '{machine_label}' not found.")
                continue

            key = (date_value, machine_label, shift)
            formulation = pending_formulations.get(key)
            if formulation is None:
                formulation = Formulation.objects.filter(
                    date=date_value,
                    machine=machine,
                    shift=shift,
                    status="Pending",
                    is_loss=is_loss
                ).first()

                if not formulation:
                    formulation = Formulation.objects.create(
                        name=f"{machine_label} {shift} {date_value}",
                        date=date_value,
                        shift=shift,
                        machine=machine,
                        created_by=request.user,
                        batches=batch_count,
                        status="Pending",
                        is_loss=is_loss,
                    )
                    created_formulations += 1
                else:
                    formulation.batches = batch_count or formulation.batches or 0
                    formulation.save()

                pending_formulations[key] = formulation

            FormulationItem.objects.create(
                formulation=formulation,
                raw_material=raw_material,
                weight=weight,
            )

        loss_weight = safe_float(row.get("PRODUCTION LOSS FORMULATION", 0.0), 0.0)
        if loss_weight > 0:
            row_has_weight = True
            loss_batches = safe_int(row.get("PRODUCTION LOSS BATCHES", 0), 0)
            loss_key = date_value
            loss_form = pending_loss_forms.get(loss_key)
            if loss_form is None:
                loss_form = Formulation.objects.filter(
                    date=date_value,
                    is_loss=True,
                    status="Pending"
                ).first()

                if not loss_form:
                    loss_form = Formulation.objects.create(
                        name=f"Production Loss {date_value}",
                        date=date_value,
                        shift="",
                        machine=None,
                        created_by=request.user,
                        batches=loss_batches,
                        status="Pending",
                        is_loss=True,
                    )
                    created_loss_forms += 1
                else:
                    loss_form.batches = loss_batches or loss_form.batches or 0
                    loss_form.save()

                pending_loss_forms[loss_key] = loss_form

            FormulationItem.objects.create(
                formulation=loss_form,
                raw_material=raw_material,
                weight=loss_weight,
            )

        if not row_has_weight:
            warnings.append(f"Row {index + 2}: no machine or loss weights found.")

    if created_formulations:
        messages.success(request, f"Uploaded {created_formulations} formulations successfully.")
    if created_loss_forms:
        messages.success(request, f"Uploaded {created_loss_forms} loss formulations successfully.")
    if warnings:
        messages.warning(request, "\n".join(warnings))

    return redirect("formulation_page")

# for both door and frame production
def _save_planning_entry(request, allowed_categories, rows_data=None):
    category = allowed_categories[0]
    rows = rows_data or []

    # Support editing an existing planning record if plan_id provided
    plan_id = request.POST.get("plan_id")
    if plan_id:
        plan = get_object_or_404(Planning, pk=plan_id, category__in=allowed_categories)
    else:
        plan = None

    # Get constant fields from POST
    plan_date = request.POST.get("plan_date")
    machine_id = request.POST.get("machine")
    thickness_id = request.POST.get("thickness")
    size_id = request.POST.get("size")
    color_id = request.POST.get("color")
    density_id = request.POST.get("density")
    remark = request.POST.get("remark", "").strip()
    remark = remark if remark in dict(Planning.REMARK_CHOICES) else ""
    stamp = request.POST.get("stamp", "").strip()
    masking = request.POST.get("masking", "").strip()

    machine = Machine.objects.get(id=machine_id) if machine_id else None

    created_plans = []

    if rows:
        # Multi-row mode: create one Planning per row
        for index, row in enumerate(rows):
            if plan_id and index == 0:
                # Update the existing plan with the first row
                current_plan = plan
            else:
                current_plan = Planning(category=category)
            current_plan.date = plan_date
            current_plan.machine = machine
            current_plan.thickness = Thickness.objects.get(id=thickness_id) if thickness_id else None
            current_plan.size = Size.objects.get(id=size_id) if size_id else None
            current_plan.color = Color.objects.get(id=color_id) if color_id else None
            current_plan.density = Density.objects.get(id=density_id) if density_id else None
            current_plan.remark = remark
            current_plan.stamp = stamp
            current_plan.masking = masking
            if category == "Door":
                # row contains height_id, width_id, quantity
                h_id = row.get("height_id")
                w_id = row.get("width_id")
                h_unit = row.get("height_unit", "inch")
                w_unit = row.get("width_unit", "inch")
                h_mm = row.get("height_mm", "")
                w_mm = row.get("width_mm", "")

                if h_unit == "mm" and h_mm:
                    mm_value = float(h_mm)
                    inch_value = mm_value / 25.4
                    current_plan.height, _ = Height.objects.get_or_create(
                        height=inch_value,
                        defaults={"unit": "mm", "original_value": mm_value},
                    )
                else:
                    current_plan.height = Height.objects.get(id=h_id) if h_id else None

                if w_unit == "mm" and w_mm:
                    mm_value = float(w_mm)
                    inch_value = mm_value / 25.4
                    current_plan.width, _ = Width.objects.get_or_create(
                        width=inch_value,
                        defaults={"unit": "mm", "original_value": mm_value},
                    )
                else:
                    current_plan.width = Width.objects.get(id=w_id) if w_id else None

                quantity = int(row.get("quantity", 0) or 0)

                thickness = float(current_plan.thickness.thickness if current_plan.thickness else 0)
                height = float(current_plan.height.height if current_plan.height else 0)
                width = float(current_plan.width.width if current_plan.width else 0)
                density = float(current_plan.density.density if current_plan.density else 0)
                h_m = height * 0.0254
                w_m = width * 0.0254
                t_m = thickness / 1000
                current_plan.weight = h_m * w_m * t_m * density
            else:
                # Frame row: length_id, quantity
                l_id = row.get("length_id")
                l_unit = row.get("length_unit", "ft")
                l_mm = row.get("length_mm", "")
                quantity = int(row.get("quantity", 0) or 0)

                if l_unit == "mm" and l_mm:
                    mm_value = float(l_mm)
                    ft_value = mm_value / 304.8
                    current_plan.length, _ = Length.objects.get_or_create(
                        length=ft_value,
                        defaults={"unit": "mm", "original_value": mm_value},
                    )
                else:
                    current_plan.length = Length.objects.get(id=l_id) if l_id else None

                current_plan.weight = float(request.POST.get("weight") or 0)

            current_plan.quantity = quantity
            current_plan.save()
            created_plans.append(current_plan)
    else:
        # Single row / edit mode
        if plan:
            current_plan = plan
        else:
            current_plan = Planning(category=category)

        current_plan.date = plan_date
        current_plan.machine = machine
        current_plan.thickness = Thickness.objects.get(id=thickness_id) if thickness_id else None
        current_plan.size = Size.objects.get(id=size_id) if size_id else None
        current_plan.color = Color.objects.get(id=color_id) if color_id else None
        current_plan.density = Density.objects.get(id=density_id) if density_id else None
        current_plan.remark = remark
        current_plan.stamp = stamp
        current_plan.masking = masking

        height_id = request.POST.get("height")
        width_id = request.POST.get("width")
        length_id = request.POST.get("length")
        height_unit = request.POST.get("height_unit", "inch")
        width_unit = request.POST.get("width_unit", "inch")
        length_unit = request.POST.get("length_unit", "ft")
        height_mm = request.POST.get("height_mm", "")
        width_mm = request.POST.get("width_mm", "")
        length_mm = request.POST.get("length_mm", "")

        if height_unit == "mm" and height_mm:
            mm_value = float(height_mm)
            inch_value = mm_value / 25.4
            current_plan.height, _ = Height.objects.get_or_create(
                height=inch_value,
                defaults={"unit": "mm", "original_value": mm_value},
            )
        else:
            current_plan.height = Height.objects.get(id=height_id) if height_id else None

        if width_unit == "mm" and width_mm:
            mm_value = float(width_mm)
            inch_value = mm_value / 25.4
            current_plan.width, _ = Width.objects.get_or_create(
                width=inch_value,
                defaults={"unit": "mm", "original_value": mm_value},
            )
        else:
            current_plan.width = Width.objects.get(id=width_id) if width_id else None

        if length_unit == "mm" and length_mm:
            mm_value = float(length_mm)
            ft_value = mm_value / 304.8
            current_plan.length, _ = Length.objects.get_or_create(
                length=ft_value,
                defaults={"unit": "mm", "original_value": mm_value},
            )
        else:
            current_plan.length = Length.objects.get(id=length_id) if length_id else None

        quantity = int(float(request.POST.get("quantity") or 0))
        current_plan.quantity = quantity

        if category == "Door":
            thickness = float(current_plan.thickness.thickness if current_plan.thickness else 0)
            height = float(current_plan.height.height if current_plan.height else 0)
            width = float(current_plan.width.width if current_plan.width else 0)
            density = float(current_plan.density.density if current_plan.density else 0)
            h_m = height * 0.0254
            w_m = width * 0.0254
            t_m = thickness / 1000
            current_plan.weight = h_m * w_m * t_m * density
        else:
            current_plan.weight = float(request.POST.get("weight") or 0)

        current_plan.save()
        created_plans.append(current_plan)

    return created_plans[0] if created_plans else None

#

def _planning_context(allowed_categories, page_title, machine_filter=None):
    plans = Planning.objects.filter(category__in=allowed_categories)
    
    # Apply machine filter if provided
    if machine_filter:
        plans = plans.filter(machine__name=machine_filter)
    
    plans = plans.order_by(
        Case(
            When(remark__iexact="Urgent", then=Value(0)),
            When(remark__iexact="Order", then=Value(1)),
            When(remark__iexact="Stock", then=Value(2)),
            default=Value(2),
            output_field=IntegerField(),
        )
    )

    # Calculate balance (remaining) quantity for each plan
    plan_ids = [p.id for p in plans]
    if plan_ids:
        production_totals = Production.objects.filter(
            planning_id__in=plan_ids,
            planning__isnull=False
        ).values('planning').annotate(total_qty=Sum('quantity'))
        produced_map = {item['planning']: item['total_qty'] for item in production_totals}
    else:
        produced_map = {}

    # Gather all items unconditionally so they can be grouped together first
    visible_plans = []
    for p in plans:
        produced = produced_map.get(p.id, 0)
        p.produced = produced
        p.balance = p.quantity - produced
        
        # Calculate raw weight, forcing it to 0 if balance is negative (overproduced)
        if p.category == "Frame":
            # For Frame: weight * balance * length
            length_val = float(p.length.length) if p.length else 1
            raw_tweight = p.weight * p.balance * length_val
        else:
            # For Door/Sheet: weight * balance (existing formula)
            raw_tweight = p.weight * p.balance
        p.tweight = max(0.0, raw_tweight)
        
        visible_plans.append(p) 

    # -----------------------------------------------------------------
    # ADVANCED MULTI-TABLE GROUPING LOGIC
    # -----------------------------------------------------------------
    grouped_plans = []
    if allowed_categories == ["Door", "Sheet"]:
        # Group by: Thickness, Density, Color
        groups = defaultdict(list)
        for p in visible_plans:
            date_str = p.date.strftime("%Y-%m-%d") if p.date else "-"
            machine_str = p.machine.name if p.machine else "-"
            thickness_str = f"{p.thickness.thickness} mm" if p.thickness else "-"
            density_str = str(p.density) if p.density else "-"
            color_str = str(p.color) if p.color else "-"
            remarks_str = p.remark if p.remark else "-"
            
            key = (date_str, machine_str, thickness_str, density_str, color_str, remarks_str)
            groups[key].append(p)
            
        for key, items in groups.items():
            # Sort items by width (small to large), then by height (small to large) within same width
            items.sort(key=lambda p: (
                float(p.width.width) if p.width else float('inf'),
                float(p.height.height) if p.height else float('inf')
            ))
            # Calculate metrics across the grouped items
            total_balance = sum(item.balance for item in items)
            total_tweight = sum(item.tweight for item in items)  # Sum of individual safe weights
            total_quantity = sum(item.quantity for item in items)
            total_produced = sum(item.produced for item in items)
            expected_days = (total_tweight / 5000) if total_tweight else 0
            
            # FILTER: Only include group if total weight is strictly greater than 0
            if total_tweight > 0: 
                grouped_plans.append({
                    'title': f"Date: {key[0]} | Machine: {key[1]} | Thickness: {key[2]} | Density: {key[3]} | Color: {key[4] }| Remarks: {key[5]}",
                    'items': items,
                    'total_quantity': total_quantity,
                    'total_produced': total_produced,
                    'total_balance': total_balance,  # Balance can be negative if overproduced
                    'total_tweight': total_tweight,  # Guaranteed non-negative
                    'expected_days': expected_days
                })
            
        machines = Machine.objects.filter(name__in=["L-4", "L-5", "L-6"], is_active=True).order_by("name")
    else:
        # Group by: Size, Color for Frames
        groups = defaultdict(list)
        for p in visible_plans:
            date_str = p.date.strftime("%Y-%m-%d") if p.date else "-"
            machine_str = p.machine.name if p.machine else "-"
            size_str = p.size.size if p.size else "-"
            color_str = str(p.color) if p.color else "-"
            remarks_str = p.remark if p.remark else "-"
            
            key = (date_str, machine_str, size_str, color_str, remarks_str)
            groups[key].append(p)
            
        for key, items in groups.items():
            # Sort items by length (small to large)
            items.sort(key=lambda p: (
                float(p.length.length) if p.length else float('inf')
            ))
            # Calculate metrics across the grouped items
            total_balance = sum(item.balance for item in items)
            total_tweight = sum(item.tweight for item in items)  # Sum of individual safe weights
            total_quantity = sum(item.quantity for item in items)
            total_produced = sum(item.produced for item in items)
            expected_days = (total_tweight / 2500) if total_tweight else 0
            
            # FILTER: Only include group if total weight is strictly greater than 0
            if total_tweight > 0:
                grouped_plans.append({
                    'title': f"Date: {key[0]} | Machine: {key[1]} | Size: {key[2]} | Color: {key[3] }| Remarks: {key[4]}",
                    'items': items,
                    'total_quantity': total_quantity,
                    'total_produced': total_produced,
                    'total_balance': total_balance,  # Balance can be negative if overproduced
                    'total_tweight': total_tweight,  # Guaranteed non-negative
                    'expected_days': expected_days
                })
            
        machines = Machine.objects.filter(name__in=["L-1", "L-2", "L-3"], is_active=True).order_by("name")
    
    return {
        "grouped_plans": grouped_plans,
        "has_records": len(grouped_plans) > 0,  # Dynamically driven by remaining active blocks
        "page_title": page_title,
        "page_heading": page_title,
        "sizes": Size.objects.filter(is_active=True).order_by("size"),
        "lengths": Length.objects.filter(is_active=True).order_by("length"),
        "thicknesses": Thickness.objects.filter(is_active=True, thickness__lte=36).order_by("thickness"),
        "heights": Height.objects.filter(is_active=True, height__gte=48).order_by('height'),
        "widths": Width.objects.filter(is_active=True, width__lte=49).order_by('width'),
        "densities": Density.objects.filter(is_active=True).order_by("density"),
        "colors": Color.objects.filter(is_active=True).order_by("color"),
        "machine_filter": machine_filter,
        "machines": machines,
    }

@login_required
@module_required('planning')
def planning_door_sheet(request):
    allowed_categories = ["Door", "Sheet"]

    if request.method == "POST":
        rows_data = []
        raw_rows = request.POST.get("door_rows")
        if raw_rows:
            try:
                rows_data = json.loads(raw_rows)
            except json.JSONDecodeError:
                rows_data = []

        if rows_data:
            plan = _save_planning_entry(request, allowed_categories, rows_data=rows_data)
            if plan:
                audit_logger.info(f"Planning created: {plan.category} (ID: {plan.id}) by user {request.user.username}, Qty: {plan.quantity}, Rows: {len(rows_data)}")
                business_logger.info(f"Planning multi-row saved: {plan.category} - Rows: {len(rows_data)}")
                messages.success(request, f"Planning saved successfully ({len(rows_data)} entries).")
        else:
            plan = _save_planning_entry(request, allowed_categories)
            if plan is not None:
                audit_logger.info(f"Planning created/updated: {plan.category} (ID: {plan.id}) by user {request.user.username}, Qty: {plan.quantity}")
                business_logger.info(f"Planning saved: {plan.category} - {plan.thickness} - {plan.height} - {plan.width}, Qty: {plan.quantity}")
                messages.success(request, "Planning saved successfully.")
        return redirect("planning_door_sheet")

    # Get machine filter from request
    machine_filter = request.GET.get('machine')
    
    context = _planning_context(
        allowed_categories,
        "Planning - Door / Sheet",
        machine_filter=machine_filter
    )
    return render(request, "planning_door_sheet.html", context)

@login_required
@module_required('planning')
def planning_frame_sheet(request):
    allowed_categories = ["Frame"]

    if request.method == "POST":
        rows_data = []
        raw_rows = request.POST.get("frame_rows")
        if raw_rows:
            try:
                rows_data = json.loads(raw_rows)
            except json.JSONDecodeError:
                rows_data = []

        if rows_data:
            plan = _save_planning_entry(request, allowed_categories, rows_data=rows_data)
            if plan:
                audit_logger.info(f"Planning created: Frame (ID: {plan.id}) by user {request.user.username}, Qty: {plan.quantity}, Rows: {len(rows_data)}")
                business_logger.info(f"Planning multi-row saved: Frame - Rows: {len(rows_data)}")
                messages.success(request, f"Planning saved successfully ({len(rows_data)} entries).")
        else:
            plan = _save_planning_entry(request, allowed_categories)
            if plan is not None:
                audit_logger.info(f"Frame planning created/updated: (ID: {plan.id}) by user {request.user.username}, Qty: {plan.quantity}")
                business_logger.info(f"Frame planning saved: {plan.size} - {plan.length}, Qty: {plan.quantity}")
                messages.success(request, "Planning saved successfully.")
        return redirect("planning_frame_sheet")

    # Get machine filter from request
    machine_filter = request.GET.get('machine')
    
    context = _planning_context(
        allowed_categories,
        "Planning - Frame",
        machine_filter=machine_filter
    )
    return render(request, "planning_frame_sheet.html", context)


@login_required
def delete_planning(request, pk):
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("planning_door_sheet")

    plan = get_object_or_404(Planning, pk=pk)
    plan_category = plan.category
    redirect_name = "planning_frame_sheet" if plan.category == "Frame" else "planning_door_sheet"
    plan.delete()
    audit_logger.info(f"Planning deleted: {plan_category} (ID: {pk}) by user {request.user.username}")
    messages.success(request, "Planning deleted successfully.")
    return redirect(redirect_name)


@login_required
def planning_detail_json(request, pk):
    plan = get_object_or_404(Planning, pk=pk)
    return JsonResponse({
        "id": plan.id,
        "category": plan.category,
        "date": plan.date.isoformat() if plan.date else None,
        "machine_id": plan.machine_id,
        "machine": str(plan.machine) if plan.machine else None,
        "thickness": str(plan.thickness) if plan.thickness else None,
        "height": str(plan.height) if plan.height else None,
        "width": str(plan.width) if plan.width else None,
        "size_id": plan.size_id,
        "size": plan.size.size if plan.size else None,
        "length_id": plan.length_id,
        "length": str(plan.length) if plan.length else None,
        "color": str(plan.color) if plan.color else None,
        "density": str(plan.density) if plan.density else None,
        "quantity": plan.quantity,
        "weight": plan.weight,
        "masking": plan.masking,
        "stamp": plan.stamp,
        "remark": plan.remark,
    })


def scrap_list(request):
    productions = Production.objects.select_related(
        "product",
        "machine",
    ).annotate(
        scrap_qty_total=Coalesce(Sum("scrap_logs__quantity"), Value(0.0), output_field=FloatField()),
        scrap_weight_total=Coalesce(Sum("scrap_logs__weight"), Value(0.0), output_field=FloatField()),
    ).order_by("-created_at")

    selected_date = request.GET.get("date")
    if selected_date:
        productions = productions.filter(created_at__date=selected_date)

    return render(request, "scrap_list.html", {
        "productions": productions,
        "selected_date": selected_date,
    })


def export_scrap_report(request):
    productions = Production.objects.select_related(
        "product",
        "machine",
    ).order_by("-created_at")

    selected_date = request.GET.get("date")
    if selected_date:
        productions = productions.filter(created_at__date=selected_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scrap Report"

    ws.append([
        "ProductID",
        "Date",
        "Machine",
        "Category",
        "Line Setting",
        "Side Patti",
        "Excess/Less",
        "Rejected Qty",
    ])

    for production in productions:
        ws.append([
            production.product.id if production.product else "",
            production.created_at.strftime("%Y-%m-%d") if production.created_at else "",
            production.machine.name if production.machine else "",
            production.category,
            round(float(production.linesetting or 0), 2),
            round(float(production.sidepatti or 0), 2),
            round(float(production.excess or 0), 2),
            round(float(production.rejected_quantity or 0), 2),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="scrap_report.xlsx"'
    wb.save(response)
    return response




# ================= PRODUCTION 2.0 VIEWS =================

# @module_required('production')
# @module_required('production')
# def production_2_0_door(request):
#     """Production 2.0 - Door View"""
#     User = get_user_model()
#     category = "Door"

#     # ---- Add Production ----
#     if request.method == "POST":
#         # Collect form fields
#         date_str = request.POST.get("date")
#         date = timezone.datetime.strptime(date_str, "%Y-%m-%d") if date_str else timezone.now()

#         machine_id = request.POST.get("machine")
#         shift = request.POST.get("shift", "Day")
#         operator_id = request.POST.get("operator")
#         remark = request.POST.get("remark", "")

#         thickness_ids = request.POST.getlist("thickness")
#         height_vals = request.POST.getlist("height")
#         width_vals = request.POST.getlist("width")
#         color_ids = request.POST.getlist("color")
#         density_vals = request.POST.getlist("density")
#         quantities = request.POST.getlist("quantity")
#         weights = request.POST.getlist("weight")
#         sidepattis = request.POST.getlist("sidepatti")
#         line_settings = request.POST.getlist("line_setting")

#         # Get planning ID from the hidden field
#         plan_id = request.POST.get("plan_id")
#         plan = None
#         if plan_id:
#             plan = Planning.objects.filter(id=plan_id).first()

#         machine = get_object_or_404(Machine, id=machine_id, category=category)
#         if request.user.role == 'Operator':
#             operator = request.user
#         else:
#             operator = User.objects.get(id=operator_id) if operator_id else request.user

#         num_rows = len(thickness_ids)

#         for idx in range(num_rows):
#             thickness = Thickness.objects.get(id=thickness_ids[idx]) if idx < len(thickness_ids) and thickness_ids[idx] else None
#             color = Color.objects.get(id=color_ids[idx]) if idx < len(color_ids) and color_ids[idx] else None
#             density = Density.objects.get(id=density_vals[idx]) if idx < len(density_vals) and density_vals[idx] else None

#             # Height
#             height_obj = None
#             height_unit = request.POST.getlist('height_unit')[idx] if idx < len(request.POST.getlist('height_unit')) else 'inch'
#             height_mm = request.POST.getlist('height_mm')[idx] if idx < len(request.POST.getlist('height_mm')) else ''

#             if height_unit == 'mm' and height_mm:
#                 try:
#                     mm_value = float(height_mm)
#                     inch_value = mm_value / 25.4
#                     height_obj, _ = Height.objects.get_or_create(height=inch_value, defaults={'unit': 'mm', 'original_value': mm_value})
#                 except ValueError:
#                     height_obj = None
#             else:
#                 height_obj = Height.objects.get(id=height_vals[idx]) if idx < len(height_vals) and height_vals[idx] else None

#             # Width
#             width_obj = None
#             width_unit = request.POST.getlist('width_unit')[idx] if idx < len(request.POST.getlist('width_unit')) else 'inch'
#             width_mm = request.POST.getlist('width_mm')[idx] if idx < len(request.POST.getlist('width_mm')) else ''

#             if width_unit == 'mm' and width_mm:
#                 try:
#                     mm_value = float(width_mm)
#                     inch_value = mm_value / 25.4
#                     width_obj, _ = Width.objects.get_or_create(width=inch_value, defaults={'unit': 'mm', 'original_value': mm_value})
#                 except ValueError:
#                     width_obj = None
#             else:
#                 width_obj = Width.objects.get(id=width_vals[idx]) if idx < len(width_vals) and width_vals[idx] else None

#             quantity = int(quantities[idx]) if idx < len(quantities) and quantities[idx] else 0
#             weight_per_piece = float(weights[idx]) if idx < len(weights) and weights[idx] else 0
#             line_setting = float(line_settings[idx]) if idx < len(line_settings) and line_settings[idx] else 0
#             sidepatti = float(sidepattis[idx]) if idx < len(sidepattis) and sidepattis[idx] else 0

#             if quantity <= 0:
#                 continue
#             if not (thickness and width_obj and height_obj and density and color):
#                 continue

#             product, _ = Product.objects.get_or_create(
#                 category=category,
#                 color=color,
#                 thickness=thickness,
#                 density=density,
#                 width=width_obj,
#                 height=height_obj,
#                 defaults={
#                     "size": None,
#                     "length": None,
#                 }
#             )

#             Production.objects.create(
#                 product=product,
#                 category=category,
#                 machine=machine,
#                 operator=operator,
#                 thickness=thickness,
#                 height=height_obj,
#                 width=width_obj,
#                 density=density,
#                 color=color,
#                 linesetting=line_setting if idx == 0 else 0,
#                 sidepatti=sidepatti if idx == 0 else 0,
#                 quantity=quantity,
#                 weight_per_piece=weight_per_piece,
#                 shift=shift,
#                 status="Pending",
#                 created_at=date,
#                 remark=remark,
#                 planning=plan
#             )

#             thickness = None
#             height = None
#             width = None
#             density = None
#             density_name = None

#         audit_logger.info(f"Production 2.0 {category} entry added by user {request.user.username} - Machine: {machine.name}, Shift: {shift}, Qty: {quantities}")
#         business_logger.info(f"Production 2.0 {category} created - Machine: {machine.name}, Operator: {operator.username}, Items: {len(thickness_ids)}")
#         messages.success(request, f"{category} production added successfully.")
#         machine_id = request.POST.get("machine")
#         if machine_id:
#             return redirect(f"/production-2-0/door/?machine_id={machine_id}")
#         return redirect("production_2_0_door")

#     # ---- GET request ----
#     machines = Machine.objects.filter(category=category, is_active=True).order_by("name")
#     machine_id = request.GET.get("machine_id")
#     selected_machine = None
#     if machine_id:
#         selected_machine = machines.filter(id=machine_id).first()

#     door_formulations = Formulation.objects.none()
#     if selected_machine:
#         door_formulations = Formulation.objects.filter(
#             machine=selected_machine,
#             planning__category__in=["Door", "Sheet"]
#         ).select_related("machine", "planning").order_by("-date", "-id")

#     # Calculate balance quantity for each formulation
#     # Balance = planning.quantity - total_produced_quantity_for_that_planning
#     plan_ids = [f.planning_id for f in door_formulations if f.planning_id]

#     if plan_ids:
#         production_totals = Production.objects.filter(
#             planning_id__in=plan_ids,
#             planning__isnull=False
#         ).values('planning').annotate(total_qty=Sum('quantity'))
#         produced_map = {item['planning']: item['total_qty'] for item in production_totals}
#     else:
#         produced_map = {}

#     for f in door_formulations:
#         if f.planning_id:
#             produced = produced_map.get(f.planning_id, 0)
#             f.balance = f.planning.quantity - produced
#         else:
#             f.balance = 0

#     context = {
#         'formulations': door_formulations,
#         'machines': machines,
#         'selected_machine': selected_machine,
#         'colors': Color.objects.filter(is_active=True),
#         'thicknesses': Thickness.objects.filter(is_active=True, thickness__lte=36).order_by('thickness'),
#         'heights': Height.objects.filter(is_active=True, height__gte=48).order_by('height'),
#         'widths': Width.objects.filter(is_active=True, width__lte=49).order_by('width'),
#         'densities': Density.objects.filter(is_active=True).order_by('density'),
#         'density_names': DensityName.objects.all(),
#         'shifts': [choice[0] for choice in Production.SHIFT_CHOICES],
#         'categories': [choice[0] for choice in Product.CATEGORY_CHOICES],
#         'operators': User.objects.filter(role="Operator").order_by("username"),
#         'current_category': category,
#     }

#     return render(request, 'production_2_0_door.html', context)

@module_required('production')
def production_2_0_door(request):
    """Production 2.0 - Door View with Advanced Multi-Table Grouping Logic"""
    User = get_user_model()
    category = "Door"

    # ---- Add Production (POST) ----
    if request.method == "POST":
        date_str = request.POST.get("date")
        date = timezone.datetime.strptime(date_str, "%Y-%m-%d") if date_str else timezone.now()

        machine_id = request.POST.get("machine")
        shift = request.POST.get("shift", "Day")
        operator_id = request.POST.get("operator")
        remark = request.POST.get("remark", "")

        thickness_ids = request.POST.getlist("thickness")
        height_vals = request.POST.getlist("height")
        width_vals = request.POST.getlist("width")
        color_ids = request.POST.getlist("color")
        density_vals = request.POST.getlist("density")
        quantities = request.POST.getlist("quantity")
        weights = request.POST.getlist("weight")
        stamp = request.POST.get("stamp", "")

        # Read summary values (Side Patti / Line Setting are stored in the summary section, not per-row)
        sidepatti = float(request.POST.get("sidepatti_summary", 0) or 0)
        line_setting = float(request.POST.get("linesetting_summary", 0) or 0)

        # plan_id is only present on per-row hidden fields now (header uses header_plan_id)
        plan_ids = request.POST.getlist("plan_id")

        plan = None
        if plan_ids and plan_ids[0]:
            plan = Planning.objects.filter(id=plan_ids[0]).first()

        machine = get_object_or_404(Machine, id=machine_id, category=category)
        if request.user.role == 'Operator':
            operator = request.user
        else:
            operator = User.objects.get(id=operator_id) if operator_id else request.user

        num_rows = len(thickness_ids)
        created_count = 0

        try:
            with transaction.atomic():
                for idx in range(num_rows):
                    thickness = Thickness.objects.get(id=thickness_ids[idx]) if idx < len(thickness_ids) and thickness_ids[idx] else None
                    color = Color.objects.get(id=color_ids[idx]) if idx < len(color_ids) and color_ids[idx] else None
                    density = Density.objects.get(id=density_vals[idx]) if idx < len(density_vals) and density_vals[idx] else None

                    # Get plan for this specific row
                    row_plan = None
                    if idx < len(plan_ids) and plan_ids[idx]:
                        row_plan = Planning.objects.filter(id=plan_ids[idx]).first()
                    if row_plan:
                        plan = row_plan

                    # Height Unit Parsing
                    height_obj = None
                    height_unit = request.POST.getlist('height_unit')[idx] if idx < len(request.POST.getlist('height_unit')) else 'inch'
                    height_mm = request.POST.getlist('height_mm')[idx] if idx < len(request.POST.getlist('height_mm')) else ''

                    if height_unit == 'mm' and height_mm:
                        try:
                            mm_value = float(height_mm)
                            inch_value = mm_value / 25.4
                            height_obj, _ = Height.objects.get_or_create(height=inch_value, defaults={'unit': 'mm', 'original_value': mm_value})
                        except ValueError:
                            height_obj = None
                    else:
                        height_obj = Height.objects.get(id=height_vals[idx]) if idx < len(height_vals) and height_vals[idx] else None

                    # Width Unit Parsing
                    width_obj = None
                    width_unit = request.POST.getlist('width_unit')[idx] if idx < len(request.POST.getlist('width_unit')) else 'inch'
                    width_mm = request.POST.getlist('width_mm')[idx] if idx < len(request.POST.getlist('width_mm')) else ''

                    if width_unit == 'mm' and width_mm:
                        try:
                            mm_value = float(width_mm)
                            inch_value = mm_value / 25.4
                            width_obj, _ = Width.objects.get_or_create(width=inch_value, defaults={'unit': 'mm', 'original_value': mm_value})
                        except ValueError:
                            width_obj = None
                    else:
                        width_obj = Width.objects.get(id=width_vals[idx]) if idx < len(width_vals) and width_vals[idx] else None

                    quantity = int(quantities[idx]) if idx < len(quantities) and quantities[idx] else 0
                    weight_per_piece = float(weights[idx]) if idx < len(weights) and weights[idx] else 0

                    # Use summary-level values (only applied to first row, matching production_view pattern)
                    row_line_setting = line_setting if idx == 0 else 0
                    row_sidepatti = sidepatti if idx == 0 else 0

                    if quantity <= 0 or not (thickness and width_obj and height_obj and density and color):
                        continue

                    product, _ = Product.objects.get_or_create(
                        category=category, color=color, thickness=thickness, density=density,
                        width=width_obj, height=height_obj, stamp=stamp,
                        defaults={"size": None, "length": None}
                    )

                    Production.objects.create(
                        product=product, category=category, machine=machine, operator=operator,
                        thickness=thickness, height=height_obj, width=width_obj, density=density, color=color,
                        linesetting=row_line_setting, sidepatti=row_sidepatti,
                        quantity=quantity, weight_per_piece=weight_per_piece, shift=shift, status="Pending",
                        created_at=date, remark=remark, planning=plan
                    )
                    created_count += 1
        except Exception as e:
            errors_logger.error(f"Production 2.0 {category} save failed: {str(e)}", exc_info=True)
            messages.error(request, f"Failed to save production: {str(e)}")
            if request.POST.get("machine"):
                return redirect(f"/production-2-0/door/?machine_id={request.POST.get('machine')}")
            return redirect("production_2_0_door")

        if created_count == 0:
            messages.warning(request, "No valid production rows were saved. Please check your entries.")
        else:
            messages.success(request, f"{created_count} production record(s) saved successfully.")

        audit_logger.info(f"Production 2.0 {category} entry added by user {request.user.username}")
        if request.POST.get("machine"):
            return redirect(f"/production-2-0/door/?machine_id={request.POST.get('machine')}")
        return redirect("production_2_0_door")

    # ---- Fetch Data & Group (GET) ----
    machines = Machine.objects.filter(category=category, is_active=True).order_by("name")
    machine_id = request.GET.get("machine_id")
    selected_machine = None
    if machine_id:
        selected_machine = machines.filter(id=machine_id).first()

    door_formulations = Formulation.objects.none()
    if selected_machine:
        door_formulations = Formulation.objects.filter(
            machine=selected_machine,
            planning__category__in=["Door", "Sheet"]
        ).select_related("machine", "planning", "planning__thickness", "planning__height", "planning__width", "planning__density", "planning__color").order_by(
            Case(
                When(planning__remark__iexact="Urgent", then=Value(0)),
                When(planning__remark__iexact="Order", then=Value(1)),
                When(planning__remark__iexact="Stock", then=Value(2)),
                default=Value(2),
                output_field=IntegerField(),
            ),
            "-date", "-id"
        )

    # Process remaining balances
    plan_ids = [f.planning_id for f in door_formulations if f.planning_id]
    if plan_ids:
        production_totals = Production.objects.filter(
            planning_id__in=plan_ids, planning__isnull=False
        ).values('planning').annotate(total_qty=Sum('quantity'))
        produced_map = {item['planning']: item['total_qty'] for item in production_totals}
    else:
        produced_map = {}

    # Apply saved group order if it exists
    saved_orders = {}
    if selected_machine:
        saved_orders_qs = GroupOrder.objects.filter(machine=selected_machine, category=category)
        for go in saved_orders_qs:
            saved_orders[go.group_key] = go.position

    # Multi-Table structural dictionary grouping logic matching planning_context
    groups = defaultdict(list)
    for f in door_formulations:
        if f.planning_id:
            produced = produced_map.get(f.planning_id, 0)
            f.balance = f.planning.quantity - produced
            f.produced = produced
        else:
            f.balance = 0
            f.produced = 0

        # Calculate total weight for display (prevent negative values)
        if f.planning and f.planning.weight and f.balance:
            f.total_weight = max(0.0, float(f.planning.weight) * float(f.balance))
        else:
            f.total_weight = 0

        # Create distinct structure grouping contexts matching your planning tables
        date_str = f.planning.date.strftime("%Y-%m-%d") if f.planning and f.planning.date else "-"
        machine_str = f.machine.name if f.machine else "-"
        thickness_str = f"{f.planning.thickness.thickness} mm" if (f.planning and f.planning.thickness) else "-"
        density_str = str(f.planning.density) if (f.planning and f.planning.density) else "-"
        color_str = str(f.planning.color) if (f.planning and f.planning.color) else "-"
        remark_str = f.planning.remark if (f.planning and f.planning.remark) else "-"

        key = (date_str, machine_str, thickness_str, density_str, color_str, remark_str)
        if f.balance > 0:
            groups[key].append(f)

    grouped_formulations = []
    for key, items in groups.items():
        # Sort items by width (small to large), then by height (small to large) within same width
        items.sort(key=lambda f: (
            float(f.planning.width.width) if (f.planning and f.planning.width) else float('inf'),
            float(f.planning.height.height) if (f.planning and f.planning.height) else float('inf')
        ))
        total_balance = sum(item.balance for item in items)
        total_tweight = sum(
            (item.planning.weight * item.balance) if item.planning else 0
            for item in items
        )
        
        # Only show block groups that still contain active weight remaining
        if total_tweight > 0:
            group_key = f"{key[0]}|{key[1]}|{key[2]}|{key[3]}|{key[4]}|{key[5]}"
            grouped_formulations.append({
                'title': f"Date: {key[0]} | Machine: {key[1]} | Thickness: {key[2]} | Density: {key[3]} | Color: {key[4]} | Remarks: {key[5]}",
                'formulations': items,
                'total_quantity': sum(item.planning.quantity for item in items if item.planning),
                'total_balance': total_balance,
                'total_tweight': total_tweight,
                'group_key': group_key,
            })
    
    # Apply saved group order
    if saved_orders:
        grouped_formulations.sort(key=lambda g: saved_orders.get(g['group_key'], 9999))

    context = {
        'grouped_formulations': grouped_formulations,
        'has_records': len(grouped_formulations) > 0,
        'machines': machines,
        'selected_machine': selected_machine,
        'colors': Color.objects.filter(is_active=True),
        'thicknesses': Thickness.objects.filter(is_active=True, thickness__lte=36).order_by('thickness'),
        'heights': Height.objects.filter(is_active=True, height__gte=48).order_by('height'),
        'widths': Width.objects.filter(is_active=True, width__lte=49).order_by('width'),
        'densities': Density.objects.filter(is_active=True).order_by('density'),
        'density_names': DensityName.objects.all(),
        'shifts': [choice[0] for choice in Production.SHIFT_CHOICES],
        'categories': [choice[0] for choice in Product.CATEGORY_CHOICES],
        'operators': User.objects.filter(role="Operator").order_by("username"),
        'current_category': category,
        'STAMP_CHOICES': Product.STAMP_CHOICES,
    }
    return render(request, 'production_2_0_door.html', context)




# @module_required('production')
# def production_2_0_frame(request):
#     """Production 2.0 - Frame View"""
#     User = get_user_model()
#     category = "Frame"

#     # ---- Add Production (operator) ----
#     if request.method == "POST":
#         # Collect form fields
#         date_str = request.POST.get("date")
#         date = timezone.datetime.strptime(date_str, "%Y-%m-%d") if date_str else timezone.now()

#         machine_id = request.POST.get("machine")
#         shift = request.POST.get("shift", "Day")
#         scrap = float(request.POST.get("scrap", 0))
#         operator_id = request.POST.get("operator")

#         size_ids = request.POST.getlist("size")
#         color_ids = request.POST.getlist("color")
#         length_ids = request.POST.getlist("length")
#         quantities = request.POST.getlist("quantity")
#         weights = request.POST.getlist("weight")
#         remark = request.POST.get("remark", "")

#         # Get planning ID from the hidden field
#         plan_id = request.POST.get("plan_id")
#         plan = None
#         if plan_id:
#             plan = Planning.objects.filter(id=plan_id).first()

#         machine = get_object_or_404(Machine, id=machine_id, category=category)
#         if request.user.role == 'Operator':
#             operator = request.user
#         else:
#             operator = User.objects.get(id=operator_id) if operator_id else request.user

#         num_rows = len(size_ids)

#         for idx in range(num_rows):
#             size = Size.objects.get(id=size_ids[idx]) if idx < len(size_ids) and size_ids[idx] else None
#             color = Color.objects.get(id=color_ids[idx]) if idx < len(color_ids) and color_ids[idx] else None

#             length_obj = None
#             length_unit = request.POST.getlist('length_unit')[idx] if idx < len(request.POST.getlist('length_unit')) else 'ft'
#             length_mm = request.POST.getlist('length_mm')[idx] if idx < len(request.POST.getlist('length_mm')) else ''

#             if length_unit == 'mm' and length_mm:
#                 try:
#                     mm_value = float(length_mm)
#                     ft_value = mm_value / 304.8
#                     length_obj, _ = Length.objects.get_or_create(length=ft_value, defaults={'unit': 'mm', 'original_value': mm_value})
#                 except ValueError:
#                     length_obj = None
#             else:
#                 length_obj = Length.objects.get(id=length_ids[idx]) if idx < len(length_ids) and length_ids[idx] else None

#             length = length_obj

#             quantity = int(quantities[idx]) if idx < len(quantities) and quantities[idx] else 0
#             weight_per_piece = float(weights[idx]) if idx < len(weights) and weights[idx] else 0

#             #  FRAME specific
#             thickness = None
#             height = None
#             width = None
#             density = None
#             density_name = None
#             line_setting = scrap if idx == 0 else 0
#             sidepatti = None

#             #  VALIDATION
#             if quantity <= 0:
#                 continue
#             if not (size and length and color):
#                 continue

#             #  CREATE PRODUCT
#             product, _ = Product.objects.get_or_create(
#                 category=category,
#                 color=color,
#                 size=size,
#                 length=length,
#                 defaults={
#                     "thickness": None,
#                     "density": None,
#                     "width": None,
#                     "height": None,
#                 }
#             )

#             Production.objects.create(
#                 product=product,
#                 category=category,
#                 machine=machine,
#                 operator=operator,
#                 size=size,
#                 length=length,
#                 color=color,
#                 thickness=thickness,
#                 height=height,
#                 width=width,
#                 density=density,
#                 density_name=density_name,
#                 linesetting=line_setting if idx == 0 else 0,
#                 sidepatti=sidepatti if category == "Door" and idx == 0 else 0,
#                 quantity=quantity,
#                 weight_per_piece=weight_per_piece,
#                 shift=shift,
#                 status="Pending",
#                 created_at=date,
#                 remark=remark,
#                 planning=plan
#             )

#         audit_logger.info(f"Production 2.0 {category} entry added by user {request.user.username} - Machine: {machine.name}, Shift: {shift}, Qty: {quantities}")
#         business_logger.info(f"Production 2.0 {category} created - Machine: {machine.name}, Operator: {operator.username}, Items: {len(size_ids)}")
#         messages.success(request, f"{category} production added successfully.")
#         machine_id = request.POST.get("machine")
#         if machine_id:
#             return redirect(f"/production-2-0/frame/?machine_id={machine_id}")
#         return redirect("production_2_0_frame")

#     # ---- GET request ----
#     machines = Machine.objects.filter(category=category, is_active=True).order_by("name")
#     machine_id = request.GET.get("machine_id")
#     selected_machine = None
#     if machine_id:
#         selected_machine = machines.filter(id=machine_id).first()

#     frame_formulations = Formulation.objects.none()
#     if selected_machine:
#         frame_formulations = Formulation.objects.filter(
#             machine=selected_machine,
#             planning__category="Frame"
#         ).select_related("machine", "planning").order_by("-date", "-id")

#     # Calculate balance quantity for each formulation
#     # Balance = planning.quantity - total_produced_quantity_for_that_planning
#     plan_ids = [f.planning_id for f in frame_formulations if f.planning_id]

#     if plan_ids:
#         production_totals = Production.objects.filter(
#             planning_id__in=plan_ids,
#             planning__isnull=False
#         ).values('planning').annotate(total_qty=Sum('quantity'))
#         produced_map = {item['planning']: item['total_qty'] for item in production_totals}
#     else:
#         produced_map = {}

#     for f in frame_formulations:
#         if f.planning_id:
#             produced = produced_map.get(f.planning_id, 0)
#             f.balance = f.planning.quantity - produced
#         else:
#             f.balance = 0

#     context = {
#         "formulations": frame_formulations,
#         "machines": machines,
#         "selected_machine": selected_machine,
#         "colors": Color.objects.filter(is_active=True),
#         "sizes": Size.objects.filter(is_active=True),
#         "lengths": Length.objects.filter(is_active=True).order_by('length'),
#         "today": timezone.now().date(),
#         "thicknesses": Thickness.objects.filter(is_active=True, thickness__lte=36).order_by('thickness'),
#         "heights": Height.objects.filter(is_active=True, height__gte=48).order_by('height'),
#         "widths": Width.objects.filter(is_active=True, width__lte=49).order_by('width'),
#         "densities": Density.objects.filter(is_active=True).order_by('density'),
#         "density_names": DensityName.objects.all(),
#         "shifts": [choice[0] for choice in Production.SHIFT_CHOICES],
#         "categories": [choice[0] for choice in Product.CATEGORY_CHOICES],
#         "operators": User.objects.filter(role="Operator").order_by("username"),
#         "current_category": category,
#     }

#     return render(request, 'production_2_0_frame.html', context)

@module_required('production')
def production_2_0_frame(request):
    """Production 2.0 - Frame View with Advanced Multi-Table Grouping Logic"""
    User = get_user_model()
    category = "Frame"

    # ---- Add Production (POST) ----
    if request.method == "POST":
        date_str = request.POST.get("date")
        date = timezone.datetime.strptime(date_str, "%Y-%m-%d") if date_str else timezone.now()

        machine_id = request.POST.get("machine")
        shift = request.POST.get("shift", "Day")
        scrap = float(request.POST.get("scrap", 0))
        operator_id = request.POST.get("operator")

        size_ids = request.POST.getlist("size")
        color_ids = request.POST.getlist("color")
        length_ids = request.POST.getlist("length")
        quantities = request.POST.getlist("quantity")
        weights = request.POST.getlist("weight")
        remark = request.POST.get("remark", "")
        stamp = request.POST.get("stamp", "")

        # plan_id is only present on per-row hidden fields now (header uses header_plan_id)
        plan_ids = request.POST.getlist("plan_id")

        plan = None
        if plan_ids and plan_ids[0]:
            plan = Planning.objects.filter(id=plan_ids[0]).first()

        machine = get_object_or_404(Machine, id=machine_id, category=category)
        if request.user.role == 'Operator':
            operator = request.user
        else:
            operator = User.objects.get(id=operator_id) if operator_id else request.user

        num_rows = len(size_ids)
        created_count = 0

        try:
            with transaction.atomic():
                for idx in range(num_rows):
                    size = Size.objects.get(id=size_ids[idx]) if idx < len(size_ids) and size_ids[idx] else None
                    color = Color.objects.get(id=color_ids[idx]) if idx < len(color_ids) and color_ids[idx] else None

                    # Get plan for this specific row
                    row_plan = None
                    if idx < len(plan_ids) and plan_ids[idx]:
                        row_plan = Planning.objects.filter(id=plan_ids[idx]).first()
                    if row_plan:
                        plan = row_plan

                    # Length Unit Parsing
                    length_obj = None
                    length_unit = request.POST.getlist('length_unit')[idx] if idx < len(request.POST.getlist('length_unit')) else 'ft'
                    length_mm = request.POST.getlist('length_mm')[idx] if idx < len(request.POST.getlist('length_mm')) else ''

                    if length_unit == 'mm' and length_mm:
                        try:
                            mm_value = float(length_mm)
                            ft_value = mm_value / 304.8
                            length_obj, _ = Length.objects.get_or_create(length=ft_value, defaults={'unit': 'mm', 'original_value': mm_value})
                        except ValueError:
                            length_obj = None
                    else:
                        length_obj = Length.objects.get(id=length_ids[idx]) if idx < len(length_ids) and length_ids[idx] else None

                    quantity = int(quantities[idx]) if idx < len(quantities) and quantities[idx] else 0
                    weight_per_piece = float(weights[idx]) if idx < len(weights) and weights[idx] else 0

                    if quantity <= 0 or not (size and length_obj and color):
                        continue

                    product, _ = Product.objects.get_or_create(
                        category=category, color=color, size=size, length=length_obj,
                        stamp=stamp,
                        defaults={"thickness": None, "density": None, "width": None, "height": None}
                    )

                    Production.objects.create(
                        product=product, category=category, machine=machine, operator=operator,
                        size=size, length=length_obj, color=color, linesetting=scrap if idx == 0 else 0, sidepatti=0,
                        quantity=quantity, weight_per_piece=weight_per_piece, shift=shift, status="Pending",
                        created_at=date, remark=remark, planning=plan
                    )
                    created_count += 1
        except Exception as e:
            errors_logger.error(f"Production 2.0 {category} save failed: {str(e)}", exc_info=True)
            messages.error(request, f"Failed to save production: {str(e)}")
            if request.POST.get("machine"):
                return redirect(f"/production-2-0/frame/?machine_id={request.POST.get('machine')}")
            return redirect("production_2_0_frame")

        if created_count == 0:
            messages.warning(request, "No valid production rows were saved. Please check your entries.")
        else:
            messages.success(request, f"{created_count} production record(s) saved successfully.")

        audit_logger.info(f"Production 2.0 {category} entry added by user {request.user.username}")
        if request.POST.get("machine"):
            return redirect(f"/production-2-0/frame/?machine_id={request.POST.get('machine')}")
        return redirect("production_2_0_frame")

    # ---- Fetch Data & Group (GET) ----
    machines = Machine.objects.filter(category=category, is_active=True).order_by("name")
    machine_id = request.GET.get("machine_id")
    selected_machine = None
    if machine_id:
        selected_machine = machines.filter(id=machine_id).first()

    frame_formulations = Formulation.objects.none()
    if selected_machine:
        frame_formulations = Formulation.objects.filter(
            machine=selected_machine,
            planning__category="Frame"
        ).select_related("machine", "planning", "planning__size", "planning__length", "planning__color").order_by(
            Case(
                When(planning__remark__iexact="Urgent", then=Value(0)),
                When(planning__remark__iexact="Order", then=Value(1)),
                When(planning__remark__iexact="Stock", then=Value(2)),
                default=Value(2),
                output_field=IntegerField(),
            ),
            "-date", "-id"
        )

    # Process remaining balances
    plan_ids = [f.planning_id for f in frame_formulations if f.planning_id]
    if plan_ids:
        production_totals = Production.objects.filter(
            planning_id__in=plan_ids, planning__isnull=False
        ).values('planning').annotate(total_qty=Sum('quantity'))
        produced_map = {item['planning']: item['total_qty'] for item in production_totals}
    else:
        produced_map = {}

    # Apply saved group order if it exists
    saved_orders = {}
    if selected_machine:
        saved_orders_qs = GroupOrder.objects.filter(machine=selected_machine, category=category)
        for go in saved_orders_qs:
            saved_orders[go.group_key] = go.position

    # Structural multi-table grouping mapping sequence
    groups = defaultdict(list)
    for f in frame_formulations:
        if f.planning_id:
            produced = produced_map.get(f.planning_id, 0)
            f.balance = f.planning.quantity - produced
            f.produced = produced
        else:
            f.balance = 0
            f.produced = 0

        # Calculate total weight for display (weight * balance * length for frames, prevent negative values)
        if f.planning and f.planning.weight and f.balance:
            length_val = float(f.planning.length.length) if f.planning.length else 1
            f.total_weight = max(0.0, float(f.planning.weight) * float(f.balance) * length_val)
        else:
            f.total_weight = 0

        date_str = f.planning.date.strftime("%Y-%m-%d") if f.planning and f.planning.date else "-"
        machine_str = f.machine.name if f.machine else "-"
        size_str = f.planning.size.size if (f.planning and f.planning.size) else "-"
        color_str = str(f.planning.color) if (f.planning and f.planning.color) else "-"
        remark_str = f.planning.remark if (f.planning and f.planning.remark) else "-"

        key = (date_str, machine_str, size_str, color_str, remark_str)
        if f.balance > 0:
            groups[key].append(f)

    grouped_formulations = []
    for key, items in groups.items():
        # Sort items by length (small to large)
        items.sort(key=lambda f: (
            float(f.planning.length.length) if (f.planning and f.planning.length) else float('inf')
        ))
        total_balance = sum(item.balance for item in items)
        total_tweight = sum(
            (item.planning.weight * item.balance * float(item.planning.length.length)) if item.planning and item.planning.length else 0
            for item in items
        )
        
        # Only show block groups that still contain active weight remaining
        if total_tweight > 0:
            group_key = f"{key[0]}|{key[1]}|{key[2]}|{key[3]}|{key[4]}"
            grouped_formulations.append({
                'title': f"Date: {key[0]} | Machine: {key[1]} | Size: {key[2]} | Color: {key[3]} | Remarks: {key[4]}",
                'formulations': items,
                'total_quantity': sum(item.planning.quantity for item in items if item.planning),
                'total_balance': total_balance,
                'total_tweight': total_tweight,
                'group_key': group_key,
            })
    
    # Apply saved group order
    if saved_orders:
        grouped_formulations.sort(key=lambda g: saved_orders.get(g['group_key'], 9999))

    context = {
        "grouped_formulations": grouped_formulations,
        "has_records": len(grouped_formulations) > 0,
        "machines": machines,
        "selected_machine": selected_machine,
        "colors": Color.objects.filter(is_active=True),
        "sizes": Size.objects.filter(is_active=True),
        "lengths": Length.objects.filter(is_active=True).order_by('length'),
        "today": timezone.now().date(),
        "shifts": [choice[0] for choice in Production.SHIFT_CHOICES],
        "categories": [choice[0] for choice in Product.CATEGORY_CHOICES],
        "operators": User.objects.filter(role="Operator").order_by("username"),
        "current_category": category,
        "STAMP_CHOICES": Product.STAMP_CHOICES,
    }
    return render(request, 'production_2_0_frame.html', context)


def download_production_template(request, category):
    """Generate a pre-filled Excel template based on the planning rows shown on the
    Production 2.0 page for the selected machine. The user fills in the production
    values (Shift, Operator, Quantity, 1 Pcs Kg, etc.) and uploads it back."""
    from django.http import HttpResponse

    machine_id = request.GET.get("machine_id")
    machine = Machine.objects.filter(id=machine_id, category=category, is_active=True).first()
    if not machine:
        messages.error(request, "Please select a valid machine first.")
        return redirect(f"production_2_0_{category.lower()}")

    # Reuse the same formulation query + balance logic as the 2.0 GET views
    if category == "Frame":
        formulations = Formulation.objects.filter(
            machine=machine, planning__category="Frame"
        ).select_related(
            "machine", "planning", "planning__size", "planning__length", "planning__color"
        )
    else:
        formulations = Formulation.objects.filter(
            machine=machine, planning__category__in=["Door", "Sheet"]
        ).select_related(
            "machine", "planning", "planning__thickness", "planning__height",
            "planning__width", "planning__density", "planning__color"
        )

    plan_ids = [f.planning_id for f in formulations if f.planning_id]
    produced_map = {}
    if plan_ids:
        totals = Production.objects.filter(
            planning_id__in=plan_ids, planning__isnull=False
        ).values('planning').annotate(total=Sum('quantity'))
        produced_map = {item['planning']: item['total'] for item in totals}

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Production"

    if category == "Frame":
        headers = [
            "Plan ID", "Date", "Category", "Shift", "Operator", "Line", "Quantity", "1 Pcs Kg",
            "SidePatti", "LineSetting", "Color", "Size", "Length Feet", "Remark",
        ]
    else:
        headers = [
            "Plan ID", "Date", "Category", "Shift", "Operator", "Line", "Quantity", "1 Pcs Kg",
            "SidePatti", "LineSetting", "Color", "Thickness", "Size WxH",
            "DENSITY (AS PER ORDER)", "PRODUCT TYPE", "Remark",
        ]

    ws.append(headers)
    # Style header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="3F2E97", end_color="3F2E97", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    for f in formulations:
        if not f.planning_id:
            continue
        produced = produced_map.get(f.planning_id, 0)
        balance = f.planning.quantity - produced
        if balance <= 0:
            continue

        plan = f.planning
        plan_id_val = plan.id
        date_val = plan.date.strftime("%Y-%m-%d") if plan.date else ""
        color_val = str(plan.color) if plan.color else ""
        machine_name = machine.name
        remark_val = plan.remark or ""

        if category == "Frame":
            size_val = plan.size.size if plan.size else ""
            length_val = plan.length.length if plan.length else ""
            row = [
                plan_id_val, date_val, "Frame", "", "", machine_name, "", "", "", "", color_val,
                size_val, length_val, remark_val,
            ]
        else:
            thick_val = plan.thickness.thickness if plan.thickness else ""
            width_val = plan.width.width if plan.width else ""
            height_val = plan.height.height if plan.height else ""
            size_wh = f"{width_val}x{height_val}" if width_val and height_val else ""
            density_val = plan.density.density if plan.density else ""
            product_type = f"{density_val}D" if density_val else ""
            row = [
                plan_id_val, date_val, "Door", "", "", machine_name, "", "", "", "", color_val,
                thick_val, size_wh, density_val, product_type, remark_val,
            ]
        ws.append(row)

    # Column widths
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 12), 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"production_{category.lower()}_template.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _safe_float(value, default=0.0):
    try:
        if pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default


def _safe_int(value, default=0):
    try:
        if pd.isna(value):
            return default
        return int(float(value))
    except Exception:
        return default


def _resolve_planning_for_row(category, machine, date_val, row):
    """Resolve the Planning record for an uploaded row.

    Priority:
    1. Use the explicit PLAN ID if provided and valid.
    2. Match by machine + date + attribute FKs.
    3. Create a new Planning if no match found.
    Returns the Planning instance (created or existing).
    """
    plan = None

    # 1. Explicit Plan ID
    plan_id_raw = row.get("PLAN ID")
    if plan_id_raw is not None and str(plan_id_raw).strip() != "":
        try:
            plan = Planning.objects.filter(id=int(float(plan_id_raw))).first()
        except (ValueError, TypeError):
            plan = None

    if plan:
        return plan

    # 2. Match by attributes
    if category == "Frame":
        size_val = str(row.get("SIZE", "")).strip()
        length_val = _safe_float(row.get("LENGTH FEET", 0), 0.0)
        color_val = str(row.get("COLOR", "")).strip()
        if size_val and length_val > 0 and color_val:
            size = Size.objects.filter(size__iexact=size_val).first()
            length = Length.objects.filter(length=length_val).first()
            color = Color.objects.filter(color__iexact=color_val).first()
            if size and length and color:
                plan = Planning.objects.filter(
                    category="Frame", machine=machine, date=date_val,
                    size=size, length=length, color=color
                ).first()
    else:
        thick_val = _safe_float(row.get("THICKNESS", 0), 0.0)
        size_wh = str(row.get("SIZE WXH", "")).strip()
        density_val = _safe_float(row.get("DENSITY (AS PER ORDER)", 0), 0.0)
        color_val = str(row.get("COLOR", "")).strip()
        if thick_val > 0 and size_wh and density_val > 0 and color_val:
            try:
                w_str, h_str = size_wh.lower().replace(" ", "").replace("x", " ").split()
                width_val = float(w_str)
                height_val = float(h_str)
            except Exception:
                width_val = height_val = None
            if width_val and height_val:
                thickness = Thickness.objects.filter(thickness=thick_val).first()
                width = Width.objects.filter(width=width_val).first()
                height = Height.objects.filter(height=height_val).first()
                density = Density.objects.filter(density=int(density_val)).first()
                color = Color.objects.filter(color__iexact=color_val).first()
                if thickness and width and height and density and color:
                    plan = Planning.objects.filter(
                        category="Door", machine=machine, date=date_val,
                        thickness=thickness, width=width, height=height,
                        density=density, color=color
                    ).first()

    # 3. Create new Planning if none matched
    if not plan:
        plan = Planning(category=category, machine=machine, date=date_val)
        if category == "Frame":
            size_val = str(row.get("SIZE", "")).strip()
            length_val = _safe_float(row.get("LENGTH FEET", 0), 0.0)
            color_val = str(row.get("COLOR", "")).strip()
            plan.size = Size.objects.filter(size__iexact=size_val).first()
            plan.length = Length.objects.filter(length=length_val).first()
            plan.color = Color.objects.filter(color__iexact=color_val).first()
            plan.weight = _safe_float(row.get("1 PCS KG", 0), 0.0)
        else:
            thick_val = _safe_float(row.get("THICKNESS", 0), 0.0)
            size_wh = str(row.get("SIZE WXH", "")).strip()
            density_val = _safe_float(row.get("DENSITY (AS PER ORDER)", 0), 0.0)
            color_val = str(row.get("COLOR", "")).strip()
            try:
                w_str, h_str = size_wh.lower().replace(" ", "").replace("x", " ").split()
                width_val = float(w_str)
                height_val = float(h_str)
            except Exception:
                width_val = height_val = 0
            plan.thickness = Thickness.objects.filter(thickness=thick_val).first()
            plan.width = Width.objects.filter(width=width_val).first()
            plan.height = Height.objects.filter(height=height_val).first()
            plan.density = Density.objects.filter(density=int(density_val)).first()
            plan.color = Color.objects.filter(color__iexact=color_val).first()
            plan.weight = _safe_float(row.get("1 PCS KG", 0), 0.0)
        plan.quantity = _safe_int(row.get("QUANTITY"), 0)
        plan.remark = str(row.get("REMARK", "")).strip() if "REMARK" in row else ""
        plan.save()

    return plan


def _ensure_formulation(plan, machine, request_user):
    """Ensure a Formulation exists for the given plan+machine so the row shows
    on the Production 2.0 page. Returns the Formulation."""
    formulation, _ = Formulation.objects.get_or_create(
        planning=plan,
        machine=machine,
        defaults={
            "name": f"{plan.category} {machine.name} {plan.date}",
            "date": plan.date,
            "shift": "Day",
            "created_by": request_user,
            "status": "Pending",
            "is_loss": False,
            "batches": 1,
        },
    )
    return formulation


def _production_exists(category, plan, machine, operator, date_val, shift, qty, weight):
    """Return True if an identical Production already exists for this upload row.
    This prevents duplicate records when the same Excel file is uploaded again."""
    return Production.objects.filter(
        planning=plan,
        machine=machine,
        operator=operator,
        shift=shift,
        created_at__date=date_val,
        quantity=qty,
        weight_per_piece=weight,
        category=category,
    ).exists()


def upload_production_excel(request, category):
    """Upload a pre-filled Excel file and create Production records linked to
    Planning/Formulation so they appear on the 2.0 pages and update balances."""
    if request.method != "POST":
        return redirect(f"production_2_0_{category.lower()}")

    file = request.FILES.get("file")
    if not file:
        messages.error(request, "Please select an Excel file to upload.")
        return redirect(f"production_2_0_{category.lower()}")

    if not file.name.lower().endswith((".xlsx", ".xls", ".xlsm", ".xltx", ".xltm")):
        messages.error(request, "Only Excel files are supported (.xlsx, .xls).")
        return redirect(f"production_2_0_{category.lower()}")

    try:
        df = pd.read_excel(file)
    except Exception as e:
        messages.error(request, f"Unable to read Excel file: {str(e)}")
        return redirect(f"production_2_0_{category.lower()}")

    df.columns = [str(col).strip().upper() for col in df.columns]

    required = ["DATE", "QUANTITY", "1 PCS KG"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        messages.error(request, f"Missing required columns: {', '.join(missing)}.")
        return redirect(f"production_2_0_{category.lower()}")

    created_count = 0
    warnings = []

    def safe_float(value, default=0.0):
        try:
            if pd.isna(value):
                return default
            return float(value)
        except Exception:
            return default

    def safe_int(value, default=0):
        try:
            if pd.isna(value):
                return default
            return int(float(value))
        except Exception:
            return default

    try:
        with transaction.atomic():
            for index, row in df.iterrows():
                qty = safe_int(row.get("QUANTITY"), 0)
                weight = safe_float(row.get("1 PCS KG"), 0.0)
                if qty <= 0 or weight <= 0:
                    warnings.append(f"Row {index + 2}: skipped (invalid quantity/weight).")
                    continue

                # Date
                raw_date = row.get("DATE")
                if pd.isna(raw_date) or str(raw_date).strip() == "":
                    warnings.append(f"Row {index + 2}: missing DATE.")
                    continue
                if isinstance(raw_date, datetime):
                    date_val = raw_date.date()
                elif isinstance(raw_date, date):
                    date_val = raw_date
                else:
                    date_val = parse_date(str(raw_date).strip())
                if not date_val:
                    warnings.append(f"Row {index + 2}: invalid DATE.")
                    continue

                shift = str(row.get("SHIFT", "")).strip() or "Day"
                shift = "Night" if shift.lower().startswith("n") else "Day"
                remark = str(row.get("REMARK", "")).strip() if "REMARK" in df.columns else ""

                # Machine
                machine_name = str(row.get("LINE", "")).strip()
                machine = None
                if machine_name:
                    machine = Machine.objects.filter(name__iexact=machine_name, category=category).first()
                if not machine:
                    warnings.append(f"Row {index + 2}: machine '{machine_name}' not found for {category}.")
                    continue

                # Operator
                operator = request.user
                if "OPERATOR" in df.columns:
                    op_name = str(row.get("OPERATOR", "")).strip()
                    if op_name:
                        op = CustomUser.objects.filter(full_name__iexact=op_name).first() or \
                             CustomUser.objects.filter(username__iexact=op_name).first()
                        if op:
                            operator = op

                # Color
                color_val = str(row.get("COLOR", "")).strip()
                if not color_val:
                    warnings.append(f"Row {index + 2}: missing COLOR.")
                    continue
                color, _ = Color.objects.get_or_create(color=color_val)

                sidepatti = safe_float(row.get("SIDEPATTI", 0), 0.0)
                line_setting = safe_float(row.get("LINESETTING", 0), 0.0)

                if category == "Frame":
                    size_val = str(row.get("SIZE", "")).strip()
                    length_val = safe_float(row.get("LENGTH FEET", 0), 0.0)
                    if not size_val or length_val <= 0:
                        warnings.append(f"Row {index + 2}: missing/invalid SIZE or LENGTH FEET.")
                        continue

                    size, _ = Size.objects.get_or_create(size=size_val)
                    length, _ = Length.objects.get_or_create(
                        length=length_val,
                        defaults={"unit": "ft", "original_value": length_val},
                    )
                    product, _ = Product.objects.get_or_create(
                        category="Frame", color=color, size=size, length=length,
                        defaults={"thickness": None, "density": None, "width": None, "height": None},
                    )
                    plan = _resolve_planning_for_row(category, machine, date_val, row)
                    _ensure_formulation(plan, machine, request.user)
                    if _production_exists("Frame", plan, machine, operator, date_val, shift, qty, weight):
                        warnings.append(f"Row {index + 2}: skipped (duplicate record already exists).")
                        continue
                    Production.objects.create(
                        product=product, category="Frame", machine=machine, operator=operator,
                        size=size, length=length, color=color,
                        linesetting=line_setting, sidepatti=sidepatti,
                        quantity=qty, weight_per_piece=weight, shift=shift, status="Pending",
                        created_at=date_val, remark=remark, planning=plan,
                    )
                    created_count += 1

                elif category == "Door":
                    thick_val = safe_float(row.get("THICKNESS", 0), 0.0)
                    size_wh = str(row.get("SIZE WXH", "")).strip()
                    density_val = safe_float(row.get("DENSITY (AS PER ORDER)", 0), 0.0)
                    product_type = str(row.get("PRODUCT TYPE", "")).strip()

                    if thick_val <= 0 or not size_wh:
                        warnings.append(f"Row {index + 2}: missing/invalid THICKNESS or SIZE WXH.")
                        continue

                    try:
                        w_str, h_str = size_wh.lower().replace(" ", "").replace("x", " ").split()
                        width_val = float(w_str)
                        height_val = float(h_str)
                    except Exception:
                        warnings.append(f"Row {index + 2}: invalid SIZE WXH '{size_wh}' (expected WxH).")
                        continue

                    thickness, _ = Thickness.objects.get_or_create(thickness=thick_val)
                    width, _ = Width.objects.get_or_create(width=width_val, defaults={"original_value": width_val})
                    height, _ = Height.objects.get_or_create(height=height_val, defaults={"original_value": height_val})

                    density = None
                    density_name_obj = None
                    if density_val > 0:
                        density, _ = Density.objects.get_or_create(density=int(density_val))
                        name_text = product_type or f"{int(density_val)}D"
                        density_name_obj, _ = DensityName.objects.get_or_create(name=name_text)

                    product, _ = Product.objects.get_or_create(
                        category="Door", color=color, thickness=thickness, density=density,
                        width=width, height=height,
                        defaults={"size": None, "length": None},
                    )
                    plan = _resolve_planning_for_row(category, machine, date_val, row)
                    _ensure_formulation(plan, machine, request.user)
                    if _production_exists("Door", plan, machine, operator, date_val, shift, qty, weight):
                        warnings.append(f"Row {index + 2}: skipped (duplicate record already exists).")
                        continue
                    Production.objects.create(
                        product=product, category="Door", machine=machine, operator=operator,
                        thickness=thickness, width=width, height=height, density=density,
                        density_name=density_name_obj, color=color,
                        linesetting=line_setting, sidepatti=sidepatti,
                        quantity=qty, weight_per_piece=weight, shift=shift, status="Pending",
                        created_at=date_val, remark=remark, planning=plan,
                    )
                    created_count += 1
                else:
                    warnings.append(f"Row {index + 2}: unsupported category.")
    except Exception as e:
        errors_logger.error(f"Production excel upload ({category}) failed: {str(e)}", exc_info=True)
        messages.error(request, f"Failed to upload production: {str(e)}")
        return redirect(f"production_2_0_{category.lower()}")

    if created_count:
        messages.success(request, f"{created_count} production record(s) uploaded successfully.")
    if warnings:
        messages.warning(request, "\n".join(warnings[:20]) + ("" if len(warnings) <= 20 else f"\n... and {len(warnings) - 20} more."))

    audit_logger.info(f"Production excel upload ({category}) by user {request.user.username}: {created_count} records")
    machine_id = request.POST.get("machine_id")
    if machine_id:
        return redirect(f"/production-2-0/{category.lower()}/?machine_id={machine_id}")
    return redirect(f"production_2_0_{category.lower()}")


def _product_dispatch_amount(product, quantity, rate):
    qty = Decimal(str(quantity or 0))
    rate_decimal = Decimal(str(rate or 0))
    
    if product.category == "Frame":
        # For Frame: rate per piece = size.rate * length
        if product.size and product.size.rate and product.length:
            size_rate = Decimal(str(product.size.rate))
            length = Decimal(str(product.length.length))
            piece_rate = (size_rate * length).quantize(Decimal("0.01"))
            amount = (piece_rate * qty).quantize(Decimal("0.01"))
            
            return {
                "weight_per_qty": None,
                "weight_per_sqft": None,
                "rate_per_sqft": None,
                "piece_rate": piece_rate,
                "amount": amount,
            }
        else:
            return {
                "weight_per_qty": None,
                "weight_per_sqft": None,
                "rate_per_sqft": None,
                "piece_rate": Decimal("0.00"),
                "amount": Decimal("0.00"),
            }

    if not (product.height and product.width and product.thickness and product.density):
        return {
            "weight_per_qty": None,
            "weight_per_sqft": None,
            "rate_per_sqft": None,
            "piece_rate": None,
            "amount": None,
        }

    height = Decimal(str(product.height.height or 0))
    width = Decimal(str(product.width.width or 0))
    thickness = Decimal(str(product.thickness.thickness or 0))
    density = Decimal(str(product.density.density or 0))
    qty = Decimal(str(quantity or 0))
    rate_decimal = Decimal(str(rate or 0))

    height_m = height * Decimal("0.0254")
    width_m = width * Decimal("0.0254")
    thickness_m = thickness / Decimal("1000")
    weight_per_qty = (height_m * width_m * thickness_m * density).quantize(Decimal("0.01"))

    sqft_exact = (height * width) / Decimal("144")
    sqft_rounded = sqft_exact.quantize(Decimal("0.01"))
    if not sqft_rounded:
        return {
            "weight_per_qty": Decimal("0.00"),
            "weight_per_sqft": Decimal("0.00"),
            "rate_per_sqft": Decimal("0.00"),
            "piece_rate": Decimal("0.00"),
            "amount": Decimal("0.00"),
        }

    weight_per_sqft = (weight_per_qty / sqft_rounded).quantize(Decimal("0.01"))
    rate_per_sqft = (weight_per_sqft * rate_decimal).quantize(Decimal("0.01"))
    amount = (sqft_exact * rate_per_sqft * qty).quantize(Decimal("0.01"))
    piece_rate = (amount / qty).quantize(Decimal("0.01")) if qty else Decimal("0.00")

    return {
        "weight_per_qty": weight_per_qty,
        "weight_per_sqft": weight_per_sqft,
        "rate_per_sqft": rate_per_sqft,
        "piece_rate": piece_rate,
        "amount": amount,
    }


def resolve_dispatch_rate(party_name, submitted_rate):
    party_label = (party_name or "").strip()
    normalized_party = party_label.lower()
    if normalized_party in {"maica plastiwood pvt. ltd.", "maica plastiwood"}:
        return Decimal("82")

    if submitted_rate is None:
        raise ValueError("Please enter a rate for this party.")

    raw_rate = str(submitted_rate).strip()
    if not raw_rate:
        raise ValueError("Please enter a rate for this party.")

    try:
        rate = Decimal(raw_rate)
    except Exception as exc:
        raise ValueError("Rate must be a valid number.") from exc

    if rate < 0:
        raise ValueError("Rate must be non-negative.")

    return rate


def _delivery_challan_print_company_details():
    return {
        "company_name": "Maica Plastiwood",
        "company_address_line1": "125-126,Industrial Area,Rao ,",
        "company_address_line2": "Indore ,M. P.,India - 453331",
        "company_address_line3": "Contact No :- 9893674442 ",
        "company_gstin": "23AENPC5208D2Z1",
        "company_state": "Madhya Pradesh",
        "company_state_name": "Madhya Pradesh",
        "company_state_code": "23",
        "company_email": "maicagroup1@gmail.com",
        "company_signature_label": "for Maica Plastiwood",
    }


def _clean_print_value(value):
    value = str(value or "").strip()
    if value in {"-", "--", "None", "none"}:
        return ""
    return value


def _delivery_challan_party_details(challan):
    party_name = _clean_print_value(getattr(challan, "party_name", ""))
    if not party_name:
        return {"address": "", "state": "", "gst": ""}

    party = Party.objects.filter(c_name__iexact=party_name).first()
    if party:
        party_address = _clean_print_value(getattr(party, "address", ""))
        party_state = _clean_print_value(getattr(party, "state", ""))
        party_gst = _clean_print_value(getattr(party, "gst", ""))
        if party_address or party_state or party_gst:
            return {"address": party_address, "state": party_state, "gst": party_gst}

    normalized_party = party_name.lower()
    if normalized_party in {"maica plastiwood pvt. ltd.", "maica plastiwood"}:
        company_details = _delivery_challan_print_company_details()
        return {
            "address": (
                f"{company_details['company_address_line1']}, "
                f"{company_details['company_address_line2']}, "
                f"{company_details['company_address_line3']}"
            ),
            "state": company_details["company_state"],
            "gst": company_details["company_gstin"],
        }

    party_address = _clean_print_value(getattr(challan, "party_address", ""))
    party_state = _clean_print_value(getattr(challan, "party_state", ""))
    party_gst = _clean_print_value(getattr(challan, "party_gst", ""))

    return {"address": party_address, "state": party_state, "gst": party_gst}


def _number_to_words(number):
    number = int(number)
    if number == 0:
        return "Zero"

    ones = [
        "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
        "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
        "Seventeen", "Eighteen", "Nineteen"
    ]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def under_thousand(n):
        parts = []
        if n >= 100:
            parts.append(f"{ones[n // 100]} Hundred")
            n %= 100
        if n >= 20:
            parts.append(tens[n // 10])
            n %= 10
        if n:
            parts.append(ones[n])
        return " ".join(parts)

    parts = []
    for value, label in ((10000000, "Crore"), (100000, "Lakh"), (1000, "Thousand")):
        if number >= value:
            parts.append(f"{under_thousand(number // value)} {label}")
            number %= value
    if number:
        parts.append(under_thousand(number))
    return " ".join(parts)


def _amount_to_words(amount):
    amount = Decimal(str(amount or 0)).quantize(Decimal("0.01"))
    rupees = int(amount)
    paise = int((amount - Decimal(rupees)) * 100)
    words = f"Rupees {_number_to_words(rupees)}"
    if paise:
        words += f" and Paise {_number_to_words(paise)}"
    return f"{words} Only"


def _delivery_challan_rows(challan):
    rows = []
    total = Decimal("0.00")
    for item in challan.items.select_related(
        "product",
        "product__height",
        "product__width",
        "product__thickness",
        "product__density",
    ).all():
        calc = _product_dispatch_amount(item.product, item.quantity, challan.rate)
        if calc["amount"] is not None:
            total += calc["amount"]
        rows.append({
            "stock": item,
            "calc": calc,
        })
    return rows, total


@module_required('stock')
def delivery_challan_pdf(request, pk):
    challan = get_object_or_404(DeliveryChallan, pk=pk)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        messages.error(request, "PDF library is not installed. Run: pip install reportlab")
        return redirect("dispatch")

    rows, total = _delivery_challan_rows(challan)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="delivery_challan_{challan.challan_no.replace("/", "-")}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    
    # Custom text styles 
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='TitleStyle', fontName='Helvetica-Bold', fontSize=12, alignment=1)
    body_style = ParagraphStyle(name='Body', fontName='Helvetica', fontSize=8, leading=10)
    body_bold = ParagraphStyle(name='BodyBold', fontName='Helvetica-Bold', fontSize=8, leading=10)
    italic_style = ParagraphStyle(name='Italic', fontName='Helvetica-Oblique', fontSize=7, alignment=2)
    small_style = ParagraphStyle(name='Small', fontName='Helvetica', fontSize=7.5, leading=9)

    story = []

    # Title
    story.append(Paragraph(f"<i>oed on {timezone.now().strftime('%d-%b-%y at %H:%M')}</i>", italic_style))
    story.append(Paragraph("DELIVERY CHALLAN", title_style))
    story.append(Spacer(1, 2*mm))

    # Box 1: Company & Metadata
    company_details = _delivery_challan_print_company_details()
    company_details_text = f"""<b>{company_details['company_name']}</b><br/>
    {company_details['company_address_line1']}<br/>
    {company_details['company_address_line2']}<br/>
    {company_details['company_address_line3']}<br/>
    GSTIN/UIN: {company_details['company_gstin']}<br/>
    State Name : {company_details['company_state_name']}, Code : {company_details['company_state_code']}<br/>
    E-Mail : {company_details['company_email']}"""

    meta_left = Table([
        [Paragraph("Challan No.<br/><b>" + challan.challan_no + "</b>", small_style),
         Paragraph(f"LR-RR No.<br/><b>{challan.lr_rr_no or '-'}</b>", small_style),
         Paragraph(f"Dated<br/><b>{challan.challan_date.strftime('%d-%b-%y')}</b>", small_style)],
        [Paragraph(f"Bill of Lading<br/><b>{challan.lr_rr_no or '-'}</b>", small_style), "",
         Paragraph(f"Motor Vehicle No.<br/><b>{challan.vehicle_no}</b>", small_style)],
         [Paragraph(f"remarks<br/><b>{challan.requirement or ''}</b>", small_style), "", ""]
    ], colWidths=[31*mm, 31*mm, 32.5*mm])
    
    meta_left.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('SPAN', (0,1), (1,1)),
        ('SPAN', (0,2), (2,2)),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 3),
    ]))

    t1 = Table([[Paragraph(company_details_text, body_style), meta_left]], colWidths=[95*mm, 95*mm])
    t1.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.black),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (1,0), (1,0), 0),
        ('LEFTPADDING', (1,0), (1,0), 0),
        ('TOPPADDING', (1,0), (1,0), 0),
        ('BOTTOMPADDING', (1,0), (1,0), 0),
    ]))
    story.append(t1)

    # Box 2: Consignee & Buyer
    party_details = _delivery_challan_party_details(challan)
    party_lines = [f"<b>{challan.party_name}</b>"]
    if party_details["address"]:
        party_lines.append(party_details["address"])
    if party_details["state"]:
        party_lines.append(f"State Name: {party_details['state']}")
    if party_details["gst"]:
        party_lines.append(f"GSTIN/UIN: {party_details['gst']}")
    party_str = "<br/>".join(party_lines)

    t2 = Table([
        [Paragraph("Consignee (Ship to)<br/>" + party_str, body_style),
         Paragraph("Buyer (Bill to)<br/>" + party_str, body_style)]
    ], colWidths=[95*mm, 95*mm])
    t2.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.black),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('PADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 0),  # Connects seamlessly to upper table
    ]))
    story.append(t2)

    # Main Data Table
    table_data = [[
        Paragraph("Sl<br/>No.", small_style), 
        Paragraph("Description of Goods", small_style), 
        Paragraph("HSN/SAC", small_style), 
        Paragraph("Quantity", small_style), 
        Paragraph("Rate<br/>(Incl. of Tax)", small_style), 
        Paragraph("Rate", small_style), 
        Paragraph("per", small_style), 
        Paragraph("Amount", small_style)
    ]]
    
    total_qty = 0
    for idx, row in enumerate(rows, 1):
        item = row["stock"]
        calc = row["calc"]
        qty = int(item.quantity)
        total_qty += qty
        
        piece_rate_display = "-" if calc["piece_rate"] is None or calc["piece_rate"] == Decimal("0.00") else f"{calc['piece_rate']}"
        amount_display = "-" if calc["amount"] is None or calc["amount"] == Decimal("0.00") else f"{calc['amount']}"
        rate_with_tax_display = "-" if calc["piece_rate"] is None or calc["piece_rate"] == Decimal("0.00") else f"{round(calc['piece_rate'] * Decimal('1.18'), 2)}"
        
        table_data.append([
            Paragraph(str(idx), small_style),
            Paragraph(item.product.product_name, body_bold),
            Paragraph("39211200", small_style),
            Paragraph(f"<b>{qty} No.</b>", small_style),
            Paragraph(str(rate_with_tax_display), small_style),
            Paragraph(str(piece_rate_display), small_style),
            Paragraph("No.", small_style),
            Paragraph(f"<b>{amount_display}</b>", small_style),
        ])

    # Append Tax calculation row & totals
    gst_rate = 18
    tax_amount = round(total * (Decimal(gst_rate)/100), 2)
    grand_total = total + tax_amount

    table_data.append([
        "", Paragraph("<br/><br/><i>IGST OUTPUT</i><br/><i>Round Off</i><br/><br/>", italic_style), "", "", "",
        Paragraph(f"<br/><br/>{gst_rate} %<br/><br/><br/>", small_style), "",
        Paragraph(f"<br/><br/><b>{tax_amount}</b><br/><b>0.00</b><br/><br/>", small_style)
    ])
    
    table_data.append([
        "", Paragraph("<b>Total</b>", body_style), "", Paragraph(f"<b>{total_qty} No.</b>", body_style), 
        "", "", "", Paragraph(f"<b>Rs. {grand_total}</b>", body_style)
    ])

    item_table = Table(table_data, colWidths=[10*mm, 68*mm, 18*mm, 20*mm, 20*mm, 17*mm, 10*mm, 27*mm])
    item_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('PADDING', (0, 0), (-1, -1), 3),
        # Erase inner horizontal lines to mimic Tally style
        ('LINEBELOW', (0, 1), (-1, -2), 0, colors.white),
        # Re-apply the top border of the total row
        ('LINEABOVE', (0, -1), (-1, -1), 0.5, colors.black),
    ]))
    story.append(item_table)

    # Footer Box (Amount & Signature)
    footer_table = Table([
        [Paragraph(f"Amount Chargeable (in words)<br/><b>Indian Rupees {_amount_to_words(grand_total)}</b>", body_style), 
         Paragraph("<i>E. & O.E</i>", italic_style)],
        [Paragraph("Declaration:<br/>We declare that this invoice shows the actual price of the goods described and that all particulars are true and correct.", body_style),
         Paragraph("<b>for Maica Plastiwood</b><br/><br/><br/><br/>Authorised Signatory", ParagraphStyle(name='right', alignment=2, fontSize=8))]
    ], colWidths=[130*mm, 60*mm])
    
    footer_table.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 0.5, colors.black),
        ('LINEBELOW', (0,0), (-1,0), 0.5, colors.black),
        ('LINEAFTER', (0,1), (0,1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('VALIGN', (1,1), (1,1), 'BOTTOM'),
        ('PADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    story.append(footer_table)

    doc.build(story)
    return response


@module_required('stock')
def dispatch_print(request, pk):
    """Print view for delivery challan"""
    challan = get_object_or_404(DeliveryChallan, pk=pk)
    
    # Get all items for this challan
    stock_items = Stock.objects.filter(delivery_challan=challan).select_related('product')
    
    # Prepare items with calculations
    items = []
    total_qty = 0
    total_amount_without_gst = Decimal('0.00')
    total_gst_amount = Decimal('0.00')
    total_amount_with_gst = Decimal('0.00')
    gst_rate = Decimal(str(challan.gst_rate))  # Use dedicated gst_rate field
    
    for stock in stock_items:
        product = stock.product
        qty = int(stock.quantity)
        
        # Calculate rate per piece (same logic as in dispatch form)
        if product.category == "Frame":
            # For Frame: rate = size.rate * length
            rate_per_piece = float(product.size.rate) if product.size and product.size.rate else 0
        else:
            # For Door/Sheet: Use calculator logic
            if product.height and product.width and product.thickness and product.density:
                h = float(product.height.height)
                w = float(product.width.width)
                t = float(product.thickness.thickness)
                d = float(product.density.density)
                rate_per_kg = float(challan.rate)
                
                hm = h * 0.0254
                wm = w * 0.0254
                tm = t / 1000
                wpq = round(hm * wm * tm * d, 2)
                sqft_exact = (h * w) / 144
                sqft_rounded = round(sqft_exact, 2)
                wpsf = round(wpq / sqft_rounded, 2) if sqft_rounded else 0
                rpsf = round(wpsf * rate_per_kg, 2)
                rate_per_piece = round(sqft_exact * rpsf, 2)
            else:
                rate_per_piece = 0
        
        total_without_gst = round(rate_per_piece * qty, 2)
        gst_amount = round(total_without_gst * gst_rate / 100, 2)
        total_with_gst = round(total_without_gst + gst_amount, 2)
        
        # Show "-" for zero or missing rates
        rate_per_piece_display = "-" if rate_per_piece == 0 else rate_per_piece
        rate_with_gst_display = "-" if rate_per_piece == 0 else round(rate_per_piece * (1 + gst_rate / 100), 2)
        total_without_gst_display = "-" if rate_per_piece == 0 else total_without_gst
        total_with_gst_display = "-" if rate_per_piece == 0 else total_with_gst
        
        items.append({
            'name': product.product_name,
            'qty': qty,
            'ratePerPiece': rate_per_piece_display,
            'rateWithGst': rate_with_gst_display,
            'totalWithoutGst': total_without_gst_display,
            'totalWithGst': total_with_gst_display
        })
        
        total_qty += qty
        total_amount_without_gst += Decimal(str(total_without_gst))
        total_gst_amount += Decimal(str(gst_amount))
        total_amount_with_gst += Decimal(str(total_with_gst))
    
    party_details = _delivery_challan_party_details(challan)

    # Convert to float for template
    context = {
        'challan': challan,
        'items': items,
        'total_qty': total_qty,
        'total_amount_without_gst': float(total_amount_without_gst),
        'total_gst_amount': float(total_gst_amount),
        'total_amount_with_gst': float(total_amount_with_gst),
        'gst_rate': float(gst_rate),
        'party_gst': party_details['gst'],
        'party_address': party_details['address'],
        'party_state': party_details['state'],
        'company_details': _delivery_challan_print_company_details(),
    }
    
    # Add amount in words
    from .views import _amount_to_words
    context['amount_in_words'] = _amount_to_words(total_amount_with_gst)
    
    return render(request, 'dispatch_print.html', context)


@module_required('stock')
def dispatch(request):
    products = Product.objects.annotate(
        total_in=Coalesce(
            Sum(
                Case(
                    When(
                        stock_movements__movement_type='IN',
                        then='stock_movements__quantity'
                    ),
                    output_field=IntegerField()
                )
            ),
            Value(0)
        ),
        total_out=Coalesce(
            Sum(
                Case(
                    When(
                        stock_movements__movement_type='OUT',
                        then='stock_movements__quantity'
                    ),
                    output_field=IntegerField()
                )
            ),
            Value(0)
        )
    ).annotate(
        stock=F('total_in') - F('total_out')
    )

    # Get filter parameters (used only for pre-selecting filter dropdown values, not for server-side filtering)
    category = request.GET.get('category')
    thickness_id = request.GET.get('thickness')
    length_id = request.GET.get('length')
    color_id = request.GET.get('color')

    # NOTE: Products are NOT filtered server-side anymore.
    # All products are passed to the template for client-side filtering.
    # This prevents page reset when filters are applied.

    if request.method == "POST":
        dispatch_data = request.POST.get("dispatch_data")
        party_name = request.POST.get("party_name", "").strip()
        vehicle_no = request.POST.get("vehicle_no", "").strip()
        lr_rr_no = request.POST.get("lr_rr_no", "").strip()
        challan_date = parse_date(request.POST.get("challan_date", "")) or timezone.now().date()
        requirement = request.POST.get("requirement", "").strip()

        if not dispatch_data:
            messages.error(request, "No products selected.")
            return redirect("dispatch")
        if not party_name or not vehicle_no:
            messages.error(request, "Party name and vehicle number are required.")
            return redirect("dispatch")

        try:
            rate = resolve_dispatch_rate(party_name, request.POST.get("rate", ""))
        except Exception as exc:
            messages.error(request, str(exc))
            return redirect("dispatch")

        try:
            gst_rate = Decimal(request.POST.get("gst_rate", "18") or "18")
        except Exception:
            messages.error(request, "GST Rate must be a valid number.")
            return redirect("dispatch")

        try:

            items = json.loads(dispatch_data)
            if not items:
                messages.error(request, "No products selected.")
                return redirect("dispatch")

            with transaction.atomic():
                # Get party details from Party model
                party = Party.objects.filter(c_name__iexact=party_name).first()
                
                challan = DeliveryChallan.objects.create(
                    challan_date=challan_date,
                    vehicle_no=vehicle_no,
                    lr_rr_no=lr_rr_no,
                    party_name=party_name,
                    rate=rate,
                    gst_rate=gst_rate,
                    requirement=requirement,
                    operator=request.user,
                    # Store party details directly in challan for persistence
                    party_address=party.address if party else '',
                    party_gst=party.gst if party else '',
                    party_state=party.state if party else ''
                )

                for item in items:

                    product = Product.objects.get(id=item["id"])
                    qty = int(item["qty"])

                    stock_data = Stock.objects.filter(product=product).aggregate(
                        total_in=Sum(
                            Case(
                                When(movement_type='IN', then=F('quantity')),
                                output_field=IntegerField()
                            )
                        ),
                        total_out=Sum(
                            Case(
                                When(movement_type='OUT', then=F('quantity')),
                                output_field=IntegerField()
                            )
                        )
                    )

                    current_stock = (
                        (stock_data["total_in"] or 0)
                        -
                        (stock_data["total_out"] or 0)
                    )

                    if qty > current_stock:
                        messages.error(
                            request,
                            f"Not enough stock for {product.product_name}"
                        )
                        raise ValueError("Insufficient stock")

                    Stock.objects.create(
                        product=product,
                        quantity=qty,
                        movement_type='OUT',
                        operator=request.user,
                        delivery_challan=challan
                    )

            audit_logger.info(f"Dispatch completed by user {request.user.username} - Items: {len(items)}")
            business_logger.info(f"Stock dispatch - Challan: {challan.challan_no} - Total items: {len(items)}")
            messages.success(request, f"Dispatch completed. Challan {challan.challan_no} created.")
            return redirect("delivery_challan_pdf", pk=challan.pk)

        except Exception as e:
            if str(e) != "Insufficient stock":
                messages.error(request, str(e))
            return redirect("dispatch")

    return render(
        request,
        "dispatch.html",
        {
            "products": products,
            "parties": Party.objects.all().order_by('c_name'),
            "today": timezone.now().date(),
            "default_rate": "",
            # Filter context (for pre-selecting filter dropdown values)
            "selected_category": category,
            "all_colors": Color.objects.filter(is_active=True).order_by('color'),
            "all_thicknesses": Thickness.objects.filter(is_active=True).order_by('thickness'),
            "all_lengths": Length.objects.filter(is_active=True).order_by('length'),
            "selected_thickness": thickness_id,
            "selected_length": length_id,
            "selected_color": color_id,
        }
    )


@module_required('stock')
def dispatch_edit(request, pk):
    """Edit an existing delivery challan"""
    challan = get_object_or_404(DeliveryChallan, pk=pk)
    
    products = Product.objects.annotate(
        total_in=Coalesce(
            Sum(
                Case(
                    When(
                        stock_movements__movement_type='IN',
                        then='stock_movements__quantity'
                    ),
                    output_field=IntegerField()
                )
            ),
            Value(0)
        ),
        total_out=Coalesce(
            Sum(
                Case(
                    When(
                        stock_movements__movement_type='OUT',
                        then='stock_movements__quantity'
                    ),
                    output_field=IntegerField()
                )
            ),
            Value(0)
        )
    ).annotate(
        stock=F('total_in') - F('total_out')
    )

    # Get filter parameters (used only for pre-selecting filter dropdown values, not for server-side filtering)
    category = request.GET.get('category')
    thickness_id = request.GET.get('thickness')
    length_id = request.GET.get('length')
    color_id = request.GET.get('color')

    # NOTE: Products are NOT filtered server-side anymore.
    # All products are passed to the template for client-side filtering.
    # This prevents page reset when filters are applied.

    # Get existing stock items for this challan
    existing_stock_items = Stock.objects.filter(delivery_challan=challan).select_related('product')
    
    if request.method == "POST":
        dispatch_data = request.POST.get("dispatch_data")
        party_name = request.POST.get("party_name", "").strip()
        vehicle_no = request.POST.get("vehicle_no", "").strip()
        lr_rr_no = request.POST.get("lr_rr_no", "").strip()
        challan_date = parse_date(request.POST.get("challan_date", "")) or timezone.now().date()
        requirement = request.POST.get("requirement", "").strip()

        if not dispatch_data:
            messages.error(request, "No products selected.")
            return redirect("dispatch_edit", pk=pk)
        if not party_name or not vehicle_no:
            messages.error(request, "Party name and vehicle number are required.")
            return redirect("dispatch_edit", pk=pk)

        try:
            rate = resolve_dispatch_rate(party_name, request.POST.get("rate", ""))
        except Exception as exc:
            messages.error(request, str(exc))
            return redirect("dispatch_edit", pk=pk)

        try:
            gst_rate = Decimal(request.POST.get("gst_rate", "18") or "18")
        except Exception:
            messages.error(request, "GST Rate must be a valid number.")
            return redirect("dispatch_edit", pk=pk)

        try:
            items = json.loads(dispatch_data)
            if not items:
                messages.error(request, "No products selected.")
                return redirect("dispatch_edit", pk=pk)

            with transaction.atomic():
                # Update challan metadata
                party = Party.objects.filter(c_name__iexact=party_name).first()
                challan.challan_date = challan_date
                challan.vehicle_no = vehicle_no
                challan.lr_rr_no = lr_rr_no
                challan.party_name = party_name
                challan.rate = rate
                challan.gst_rate = gst_rate
                challan.requirement = requirement
                challan.party_address = party.address if party else ''
                challan.party_gst = party.gst if party else ''
                challan.party_state = party.state if party else ''
                challan.save()

                # Collect old quantities BEFORE deleting (to use for stock adjustment check)
                old_stock_items = Stock.objects.filter(
                    delivery_challan=challan, movement_type='OUT'
                )
                old_qty_map = {}
                for old_stock in old_stock_items:
                    pid = old_stock.product_id
                    old_qty_map[pid] = old_qty_map.get(pid, 0) + old_stock.quantity

                # Remove old stock items for this challan
                old_stock_items.delete()

                # Create new stock items
                for item in items:
                    product = Product.objects.get(id=item["id"])
                    qty = int(item["qty"])

                    # Calculate current stock considering other OUT movements not related to this challan
                    stock_data = Stock.objects.filter(product=product).aggregate(
                        total_in=Sum(
                            Case(
                                When(movement_type='IN', then=F('quantity')),
                                output_field=IntegerField()
                            )
                        ),
                        total_out=Sum(
                            Case(
                                When(movement_type='OUT', then=F('quantity')),
                                output_field=IntegerField()
                            )
                        )
                    )

                    current_stock = (
                        (stock_data["total_in"] or 0)
                        -
                        (stock_data["total_out"] or 0)
                    )

                    # For edit mode, add back the old quantities that were removed
                    old_qty_for_this_challan = old_qty_map.get(product.id, 0)

                    available_stock = current_stock + old_qty_for_this_challan

                    if qty > available_stock:
                        messages.error(
                            request,
                            f"Not enough stock for {product.product_name}"
                        )
                        raise ValueError("Insufficient stock")

                    Stock.objects.create(
                        product=product,
                        quantity=qty,
                        movement_type='OUT',
                        operator=request.user,
                        delivery_challan=challan
                    )

            audit_logger.info(f"Dispatch edited by user {request.user.username} - Challan: {challan.challan_no}")
            business_logger.info(f"Stock dispatch edited - Challan: {challan.challan_no} - Total items: {len(items)}")
            messages.success(request, f"Dispatch updated. Challan {challan.challan_no} updated.")
            return redirect("delivery_challan_pdf", pk=challan.pk)

        except Exception as e:
            if str(e) != "Insufficient stock":
                messages.error(request, str(e))
            return redirect("dispatch_edit", pk=pk)

    # Build edit data for the template
    edit_data = []
    for stock in existing_stock_items:
        edit_data.append({
            "id": stock.product.id,
            "name": str(stock.product),
            "qty": int(stock.quantity),
        })

    return render(
        request,
        "dispatch.html",
        {
            "products": products,
            "parties": Party.objects.all().order_by('c_name'),
            "today": timezone.now().date(),
            "default_rate": str(challan.rate) if challan.rate else "",
            "edit_mode": True,
            "edit_challan": challan,
            "edit_data_json": json.dumps(edit_data),
            "edit_gst_rate": float(challan.gst_rate) if challan.gst_rate else 18,
            # Filter context (for filter bar UI consistency)
            "selected_category": request.GET.get('category'),
            "all_colors": Color.objects.filter(is_active=True).order_by('color'),
            "all_thicknesses": Thickness.objects.filter(is_active=True).order_by('thickness'),
            "all_lengths": Length.objects.filter(is_active=True).order_by('length'),
            "selected_thickness": request.GET.get('thickness'),
            "selected_length": request.GET.get('length'),
            "selected_color": request.GET.get('color'),
        }
    )


@module_required('stock')
def dispatch_edit_qty(request, pk):
    """Edit quantities of items in an existing delivery challan (quick edit)"""
    # Allow Admin and Dispatch Manager to access this view
    if request.user.role not in ['Admin', 'Dispatch Manager']:
        messages.error(request, "You don't have permission to edit dispatch quantities.")
        return redirect("reports")
    challan = get_object_or_404(DeliveryChallan, pk=pk)
    stock_items = Stock.objects.filter(delivery_challan=challan).select_related('product')
    
    if request.method == "POST":
        try:
            with transaction.atomic():
                updated_count = 0
                for stock in stock_items:
                    new_qty = request.POST.get(f"qty_{stock.id}")
                    if new_qty is not None:
                        new_qty = int(new_qty)
                        if new_qty >= 0:
                            stock.quantity = new_qty
                            stock.save()
                            updated_count += 1
                
                if updated_count > 0:
                    audit_logger.info(f"Dispatch quantities edited by user {request.user.username} - Challan: {challan.challan_no}, Items updated: {updated_count}")
                    business_logger.info(f"Stock dispatch quantities edited - Challan: {challan.challan_no} - Items updated: {updated_count}")
                    messages.success(request, f"Quantities updated successfully for {updated_count} item(s).")
                else:
                    messages.warning(request, "No items were updated.")
                    
        except Exception as e:
            messages.error(request, f"Error updating quantities: {str(e)}")
        
        return redirect(f"/reports/?type=dispatch")
    
    return render(
        request,
        "dispatch_edit_qty.html",
        {
            "challan": challan,
            "stock_items": stock_items,
        }
    )


@login_required
def get_last_formulation(request, group_key):
    """API endpoint to get the last formulation material weights for a group"""
    try:
        # Decode the group key (it's URL-encoded)
        from urllib.parse import unquote
        decoded_key = unquote(group_key)
        
        # Extract machine name from group key
        # Format: "Date: YYYY-MM-DD | Machine: MACHINE_NAME | ..."
        machine_name = None
        if "Machine:" in decoded_key:
            parts = decoded_key.split("|")
            for part in parts:
                if "Machine:" in part:
                    machine_name = part.split("Machine:")[1].strip()
                    break
        
        if not machine_name:
            return JsonResponse({'materials': []})
        
        # Find the most recent formulation for this machine
        last_formulation = Formulation.objects.filter(
            machine__name=machine_name,
            is_loss=False
        ).prefetch_related('items__raw_material').order_by('-date', '-id').first()
        
        if not last_formulation:
            return JsonResponse({'materials': []})
        
        # Extract material weights
        materials = []
        for item in last_formulation.items.all():
            materials.append({
                'id': item.raw_material.id,
                'name': item.raw_material.name,
                'weight': float(item.weight)
            })
        
        return JsonResponse({'materials': materials})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_formulation_by_group(request):
    """API endpoint to auto-fill formulation materials based on group attributes.
    
    Strategy:
    1. First try to find an exact match (same thickness/density/color for Door,
       or same size/color for Frame) on the same machine.
    2. If no exact match exists, fallback to the most recent formulation
       on the same machine (any attributes).
    3. This ensures auto-fill always works even for new product combinations.
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'GET required'}, status=400)
    
    try:
        machine_id = request.GET.get('machine_id')
        category = request.GET.get('category', 'Door')
        
        if not machine_id:
            return JsonResponse({'error': 'machine_id is required'}, status=400)
        
        machine = get_object_or_404(Machine, id=machine_id)
        
        # ----- STEP 1: Try exact match -----
        query = Formulation.objects.filter(
            machine=machine,
            is_loss=False,
            status='Approved'
        )
        
        if category == 'Door':
            thickness_val = request.GET.get('thickness')
            density_val = request.GET.get('density')
            color_val = request.GET.get('color')
            
            if thickness_val:
                query = query.filter(planning__thickness__thickness=float(thickness_val))
            if density_val:
                query = query.filter(planning__density__density=int(density_val))
            if color_val:
                query = query.filter(planning__color__color__iexact=color_val)
                
        elif category == 'Frame':
            size_val = request.GET.get('size')
            color_val = request.GET.get('color')
            
            if size_val:
                query = query.filter(planning__size__size__iexact=size_val)
            if color_val:
                query = query.filter(planning__color__color__iexact=color_val)
        
        # Try to get exact match first
        last_formulation = query.prefetch_related(
            'items__raw_material'
        ).order_by('-date', '-id').first()
        
        match_type = 'exact'
        
        # ----- STEP 2: If no exact match, fallback to machine's most recent -----
        if not last_formulation:
            last_formulation = Formulation.objects.filter(
                machine=machine,
                is_loss=False,
                status='Approved'
            ).prefetch_related(
                'items__raw_material'
            ).order_by('-date', '-id').first()
            
            match_type = 'fallback_machine'
        
        if not last_formulation:
            return JsonResponse({
                'materials': [],
                'message': 'No formulations found for this machine'
            })
        
        # Extract material weights
        materials = []
        for item in last_formulation.items.all():
            materials.append({
                'id': item.raw_material.id,
                'name': item.raw_material.name,
                'weight': float(item.weight)
            })
        
        response_data = {
            'materials': materials,
            'formulation_name': last_formulation.name,
            'formulation_date': last_formulation.date.strftime('%Y-%m-%d'),
            'match_type': match_type,
        }
        
        if match_type == 'fallback_machine':
            # Include the attributes of the matched formulation for transparency
            p = last_formulation.planning
            if p:
                attrs = []
                if p.thickness:
                    attrs.append(f"Thickness: {p.thickness.thickness} mm")
                if p.density:
                    attrs.append(f"Density: {p.density.density}")
                if p.size:
                    attrs.append(f"Size: {p.size.size}")
                if p.color:
                    attrs.append(f"Color: {p.color.color}")
                response_data['fallback_attributes'] = ', '.join(attrs)
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def api_group_batch_input(request):
    """API endpoint to set batch count for all formulations in a group"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if request.user.role not in ['Admin', 'Manager']:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
        group_name = (data.get('group_name') or '').strip()
        batch_count = int(data.get('batch_count') or 0)
        
        if not group_name:
            return JsonResponse({'error': 'Group name is required'}, status=400)
        
        if batch_count < 0:
            return JsonResponse({'error': 'Batch count must be 0 or greater'}, status=400)
        
        # Get all formulations matching the group name (product name)
        formulations = Formulation.objects.filter(name=group_name)
        
        if not formulations.exists():
            return JsonResponse({'error': f'No formulations found for group: {group_name}'}, status=404)
        
        updated_count = 0
        for formulation in formulations:
            formulation.batches = batch_count
            formulation.save()
            updated_count += 1
        
        audit_logger.info(f"Group batch input: {updated_count} formulations in group '{group_name}' set to {batch_count} batches by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'updated_count': updated_count,
            'group_name': group_name,
            'batch_count': batch_count
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except ValueError:
        return JsonResponse({'error': 'Invalid batch count'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_group_approve(request):
    """API endpoint to approve all formulations in a group (no stock deduction)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if request.user.role not in ['Admin', 'Manager']:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
        group_name = (data.get('group_name') or '').strip()
        
        if not group_name:
            return JsonResponse({'error': 'Group name is required'}, status=400)
        
        # Get all formulations in this group
        formulations = Formulation.objects.filter(name=group_name, status="Pending")
        
        if not formulations.exists():
            return JsonResponse({'error': f'No pending formulations found for group: {group_name}'}, status=404)
        
        # Mark all formulations as approved (NO stock deduction)
        approved_count = 0
        with transaction.atomic():
            for formulation in formulations:
                formulation.status = "Approved"
                formulation.approved_at = now()
                formulation.save()
                approved_count += 1
        
        audit_logger.info(f"Group approve: {approved_count} formulations in group '{group_name}' approved by {request.user.username}. No stock deduction.")
        business_logger.info(f"Group approval completed for '{group_name}' - {approved_count} formulations approved")
        
        return JsonResponse({
            'success': True,
            'approved_count': approved_count,
            'group_name': group_name,
            'message': f'{approved_count} formulation(s) approved. Stock deduction should be done via Add button.'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_group_edit(request):
    """API endpoint to update all formulations in a group at once"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)

    if request.user.role not in ['Admin', 'Manager']:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        # Parse form data from URL-encoded POST
        group_name = (request.POST.get('group_name') or '').strip()
        date = request.POST.get('date')
        machine_id = request.POST.get('machine')
        shift = request.POST.get('shift')
        name = request.POST.get('name')
        materials_json = request.POST.get('materials', '[]')

        if not group_name:
            return JsonResponse({'error': 'Group name is required'}, status=400)

        # Parse materials JSON
        try:
            materials = json.loads(materials_json)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid materials data'}, status=400)

        # Get all formulations in this group
        formulations = Formulation.objects.filter(name=group_name)

        if not formulations.exists():
            return JsonResponse({'error': f'No formulations found for group: {group_name}'}, status=404)

        updated_count = 0

        with transaction.atomic():
            for formulation in formulations:
                # Update top fields
                formulation.date = date
                formulation.machine_id = machine_id
                formulation.shift = shift
                formulation.name = name
                formulation.save()

                # Remove old materials
                formulation.items.all().delete()

                # Add new materials
                for item in materials:
                    if float(item['weight']) > 0:
                        FormulationItem.objects.create(
                            formulation=formulation,
                            raw_material_id=item['id'],
                            weight=item['weight']
                        )

                updated_count += 1

        audit_logger.info(f"Group edit: {updated_count} formulations in group '{group_name}' updated by {request.user.username}")

        return JsonResponse({
            'success': True,
            'updated_count': updated_count,
            'group_name': group_name,
            'formulation_ids': [f.id for f in formulations]
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def reorder_groups(request):
    """API endpoint to save the custom display order of groups on Production 2.0 pages"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if request.user.role not in ['Admin', 'Manager']:
        return JsonResponse({'error': 'Permission denied. Only Admin/Manager can reorder.'}, status=403)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
        machine_id = data.get('machine_id')
        category = data.get('category')
        order = data.get('order', [])
        
        if not machine_id or not category or not order:
            return JsonResponse({'error': 'machine_id, category, and order are required'}, status=400)
        
        machine = get_object_or_404(Machine, id=machine_id)
        
        if category not in ['Door', 'Frame']:
            return JsonResponse({'error': 'category must be Door or Frame'}, status=400)
        
        # Delete existing group orders for this machine and category
        GroupOrder.objects.filter(machine=machine, category=category).delete()
        
        # Create new order records
        for idx, group_key in enumerate(order):
            GroupOrder.objects.create(
                machine=machine,
                category=category,
                group_key=group_key,
                position=idx
            )
        
        audit_logger.info(f"Group order updated for {machine.name} - {category} by {request.user.username}")
        return JsonResponse({'success': True, 'saved': len(order)})
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_group_add_stock_deduct(request):
    """API endpoint to add new weights to a group formulation, deduct stock, and reset batches to 0"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    if request.user.role not in ['Admin', 'Manager']:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    try:
        data = json.loads(request.body.decode('utf-8'))
        group_name = (data.get('group_name') or '').strip()
        batch_count = int(data.get('batch_count') or 0)
        date = data.get('date')
        machine_id = data.get('machine')
        shift = data.get('shift')
        name = data.get('name')
        materials_json = data.get('materials', '[]')
        
        if not group_name:
            return JsonResponse({'error': 'Group name is required'}, status=400)
        
        if batch_count <= 0:
            return JsonResponse({'error': 'Batch count must be greater than 0'}, status=400)
        
        # Parse materials
        try:
            materials = json.loads(materials_json) if isinstance(materials_json, str) else materials_json
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid materials data'}, status=400)
        
        if not materials:
            return JsonResponse({'error': 'At least one material weight is required'}, status=400)
        
        # Find the existing formulation for this group
        formulation = Formulation.objects.filter(name=group_name).first()
        if not formulation:
            return JsonResponse({'error': f'No formulation found for group: {group_name}'}, status=404)
        
        deduction_summary = []
        stock_issues = []
        
        with transaction.atomic():
            # STEP 1: Check stock availability and calculate deductions
            for item in materials:
                material_id = item.get('id')
                new_weight = float(item.get('weight', 0))
                
                if new_weight <= 0:
                    continue
                
                raw_material = RawMaterial.objects.get(id=material_id)
                total_deduction = new_weight * batch_count
                
                if raw_material.current_stock < total_deduction:
                    stock_issues.append({
                        'material': raw_material.name,
                        'required': total_deduction,
                        'available': raw_material.current_stock
                    })
            
            if stock_issues:
                return JsonResponse({
                    'error': 'Insufficient stock',
                    'details': stock_issues
                }, status=400)
            
            # STEP 2: Create BatchAddHistory record first (so stock deductions can link to it)
            batch_history = BatchAddHistory.objects.create(
                formulation=formulation,
                group_name=group_name,
                batch_count=batch_count,
                date=date,
                shift=shift or 'Day',
                machine_id=machine_id,
                added_by=request.user
            )
            
            # STEP 3: Deduct stock and link to batch history
            for item in materials:
                material_id = item.get('id')
                new_weight = float(item.get('weight', 0))
                
                if new_weight <= 0:
                    continue
                
                raw_material = RawMaterial.objects.get(id=material_id)
                total_deduction = new_weight * batch_count
                
                raw_material.current_stock -= total_deduction
                raw_material.save()
                
                RawMaterialStock.objects.create(
                    raw_material=raw_material,
                    quantity=total_deduction,
                    movement_type='OUT',
                    formulation=formulation,
                    batch_history=batch_history
                )
                
                deduction_summary.append({
                    'material': raw_material.name,
                    'deducted': round(total_deduction, 2)
                })
            
            # STEP 4: Update formulation fields
            if date:
                from datetime import datetime as dt
                formulation.date = dt.strptime(date, '%Y-%m-%d').date()
            if machine_id:
                formulation.machine_id = machine_id
            if shift:
                formulation.shift = shift
            if name:
                formulation.name = name
            
            # Mark as added via +Add button
            formulation.added_via_button = True
            
            # STEP 5: Reset batches to 0
            formulation.batches = 0
            formulation.save()
            
            # STEP 6: Update material weights
            formulation.items.all().delete()
            for item in materials:
                if float(item.get('weight', 0)) > 0:
                    FormulationItem.objects.create(
                        formulation=formulation,
                        raw_material_id=item['id'],
                        weight=float(item['weight'])
                    )
            
            # STEP 7: Save each raw material row for this batch addition
            for item in materials:
                if float(item.get('weight', 0)) > 0:
                    FormulationBatch.objects.create(
                        batch_history=batch_history,
                        raw_material_id=item['id'],
                        weight=float(item['weight']),
                        date=date,
                        shift=shift or 'Day'
                    )
        
        audit_logger.info(f"Group add with stock deduct: '{group_name}' updated by {request.user.username}. Batches: {batch_count}, Materials deducted: {len(deduction_summary)}")
        business_logger.info(f"Group add stock deduct completed for '{group_name}' - {batch_count} batches, {len(deduction_summary)} materials deducted")
        
        return JsonResponse({
            'success': True,
            'formulation_id': formulation.id,
            'group_name': group_name,
            'batch_count': batch_count,
            'deduction_summary': deduction_summary,
            'message': f'Stock deducted for {len(deduction_summary)} materials. Batch reset to 0.'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except RawMaterial.DoesNotExist as e:
        return JsonResponse({'error': f'Raw material not found: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ================= POWER BI DASHBOARD =================

@module_required('reports')
@admin_only
def power_bi_report(request):
    """Power BI-style dashboard with charts, pivot tables, and multiple views"""
    from datetime import datetime as dt_mod
    from django.db.models import Count, Avg
    from django.core.serializers.json import DjangoJSONEncoder
    from .powerbi_reports import get_report_data
    
    view_format = request.GET.get("view", "chart")
    report_type = request.GET.get("report_type", "production")
    
    # Get filter parameters
    from_date = request.GET.get("from_date", "")
    to_date = request.GET.get("to_date", "")
    machine_ids = request.GET.getlist("machine")
    shift = request.GET.get("shift", "")
    category = request.GET.get("category", "")
    operator_id = request.GET.get("operator", "")
    
    # Get report data from the new module
    report_data = get_report_data(request)
    
    # Get user's saved dashboards
    user_dashboards = Dashboard.objects.filter(user=request.user).order_by('-is_default', '-updated_at')
    
    # Get all machines for filters
    all_machines = Machine.objects.filter(is_active=True).order_by('name')
    operators = CustomUser.objects.filter(role='Operator').order_by('full_name')
    
    # Get today's date for defaults
    today = timezone.now().date()
    today_str = today.isoformat()
    week_ago = (today - timedelta(days=7)).isoformat()
    
    # Build context from report data
    # Use DjangoJSONEncoder to handle Decimal and other types, then mark as safe
    import json as json_mod
    context = {
        'report_type': report_type,
        'view_format': view_format,
        'from_date': from_date or week_ago,
        'to_date': to_date or today_str,
        'selected_machines': machine_ids,
        'selected_shift': shift,
        'selected_category': category,
        'selected_operator': operator_id,
        'all_machines': all_machines,
        'operators': operators,
        'user_dashboards': user_dashboards,
        
        # Serialize with DjangoJSONEncoder for safe rendering
        'json_kpi_data': json_mod.dumps(report_data['kpis'], cls=DjangoJSONEncoder),
        'json_chart_data': json_mod.dumps(report_data['charts'], cls=DjangoJSONEncoder),
        'json_table_data': json_mod.dumps(report_data['table_data'], cls=DjangoJSONEncoder),
        'json_table_config': json_mod.dumps(report_data['table_config'], cls=DjangoJSONEncoder),
        'json_pivot_data': json_mod.dumps(report_data['pivot_data'], cls=DjangoJSONEncoder),
        'json_pivot_config': json_mod.dumps(report_data['pivot_config'], cls=DjangoJSONEncoder),
    }
    
    return render(request, 'power_bi_dashboard.html', context)


@admin_only
def power_bi_data(request):
    """AJAX endpoint to get filtered chart/pivot data"""
    from .powerbi_reports import get_report_data
    
    report_data = get_report_data(request)
    
    return JsonResponse(report_data)


@login_required
def save_dashboard(request):
    """Save or update a dashboard configuration"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    import json as json_mod
    
    try:
        data = json_mod.loads(request.body)
        dashboard_id = data.get('id')
        name = data.get('name', 'New Dashboard')
        description = data.get('description', '')
        default_format = data.get('default_format', 'chart')
        layout_config = data.get('layout_config', {})
        filters_config = data.get('filters_config', {})
        pivot_config = data.get('pivot_config', {})
        set_as_default = data.get('is_default', False)
        auto_refresh = data.get('auto_refresh_interval', 0)
        
        if dashboard_id:
            dash = get_object_or_404(Dashboard, id=dashboard_id, user=request.user)
        else:
            dash = Dashboard(user=request.user)
        
        dash.name = name
        dash.description = description
        dash.default_format = default_format
        dash.layout_config = layout_config
        dash.filters_config = filters_config
        dash.pivot_config = pivot_config
        dash.auto_refresh_interval = int(auto_refresh)
        
        if set_as_default:
            # Unset other defaults
            Dashboard.objects.filter(user=request.user, is_default=True).update(is_default=False)
            dash.is_default = True
        else:
            dash.is_default = False
        
        dash.save()
        
        return JsonResponse({
            'success': True,
            'dashboard': {
                'id': dash.id,
                'name': dash.name,
                'default_format': dash.default_format,
                'is_default': dash.is_default,
                'auto_refresh_interval': dash.auto_refresh_interval,
                'updated_at': dash.updated_at.isoformat(),
            }
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def load_dashboard(request, dashboard_id):
    """Load a saved dashboard configuration"""
    dash = get_object_or_404(Dashboard, id=dashboard_id, user=request.user)
    return JsonResponse({
        'id': dash.id,
        'name': dash.name,
        'description': dash.description,
        'default_format': dash.default_format,
        'layout_config': dash.layout_config,
        'filters_config': dash.filters_config,
        'pivot_config': dash.pivot_config,
        'is_default': dash.is_default,
        'auto_refresh_interval': dash.auto_refresh_interval,
        'created_at': dash.created_at.isoformat(),
        'updated_at': dash.updated_at.isoformat(),
    })


@login_required
def delete_dashboard(request, dashboard_id):
    """Delete a saved dashboard"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    dash = get_object_or_404(Dashboard, id=dashboard_id, user=request.user)
    dash.delete()
    return JsonResponse({'success': True})


@login_required
def list_dashboards(request):
    """List user's saved dashboards"""
    dashboards = Dashboard.objects.filter(user=request.user).order_by('-is_default', '-updated_at')
    data = [{
        'id': d.id,
        'name': d.name,
        'description': d.description,
        'default_format': d.default_format,
        'is_default': d.is_default,
        'auto_refresh_interval': d.auto_refresh_interval,
        'updated_at': d.updated_at.isoformat(),
    } for d in dashboards]
    return JsonResponse({'dashboards': data})


# ================= USER SESSIONS PAGE =================

@login_required
@admin_only
def user_sessions(request):
    
    sessions = UserSession.objects.all().select_related('user').order_by('-date', '-login_time')
    
    # Apply filters
    date_filter = request.GET.get('date')
    user_filter = request.GET.get('user')
    
    active_sessions = sessions.filter(logout_time__isnull=True)
    
    if date_filter:
        try:
            from datetime import datetime
            date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
            sessions = sessions.filter(date=date_obj)
        except ValueError:
            pass
    
    if user_filter:
        sessions = sessions.filter(user_id=user_filter)
    
    # Get all users for the filter dropdown
    users = CustomUser.objects.all().order_by('username')
    
    context = {
        'sessions': sessions,
        'users': users,
        'selected_date': date_filter or '',
        'selected_user': user_filter or '',
        'active_sessions_count': active_sessions.count(),
    }
    
    return render(request, 'user_sessions.html', context)

# ================= LOGS PAGE (ADMIN ONLY) =================

@admin_only
def logs_view(request):
    """Render the admin logs page with Audit, Business, and Error tabs"""
    return render(request, 'logs.html')


@admin_only
def logs_api(request):
    """API endpoint that reads log files and returns structured JSON"""
    log_dir = settings.LOGGING_DIR
    
    def parse_log_line(line, log_type):
        """Parse a single log line into structured data"""
        line = line.strip()
        if not line:
            return None
        
        try:
            # Format: LEVEL YYYY-MM-DD HH:MM:SS module message
            parts = line.split(' ', 4)
            if len(parts) < 5:
                return None
            
            level = parts[0]
            date_str = parts[1]
            time_str = parts[2]
            module = parts[3]
            message = parts[4]
            
            # Convert UTC timestamp to local timezone (Asia/Calcutta)
            try:
                naive_dt = datetime.datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                aware_utc = timezone.make_aware(naive_dt, timezone.utc)
                local_dt = timezone.localtime(aware_utc)
                date_str = local_dt.strftime("%Y-%m-%d")
                time_str = local_dt.strftime("%H:%M:%S")
            except Exception:
                # If conversion fails, use original values
                pass
            
            result = {
                'date': date_str,
                'time': time_str,
                'action': message,
                'level': level,
            }
            
            # Extract user info from message for audit logs
            if log_type == 'audit':
                user_prefix = ''
                username = ''
                user_extra = ''
                
                # Pattern: "by user Username" or "by user Username, extra data"
                if 'by user ' in message:
                    user_prefix = 'by user'
                    after = message.split('by user ', 1)[1].strip()
                    # Username ends at comma, period, space-dash, or end of string
                    import re
                    match = re.match(r'^([A-Za-z0-9_]+)(.*)', after)
                    if match:
                        username = match.group(1)
                        user_extra = match.group(2).strip()
                
                # Pattern: "approved by Username" or "updated by Username"
                elif 'approved by ' in message:
                    user_prefix = 'approved by'
                    after = message.split('approved by ', 1)[1].strip()
                    match = re.match(r'^([A-Za-z0-9_]+)(.*)', after)
                    if match:
                        username = match.group(1)
                        user_extra = match.group(2).strip()
                
                elif 'updated by ' in message:
                    user_prefix = 'updated by'
                    after = message.split('updated by ', 1)[1].strip()
                    match = re.match(r'^([A-Za-z0-9_]+)(.*)', after)
                    if match:
                        username = match.group(1)
                        user_extra = match.group(2).strip()
                
                result['user_prefix'] = user_prefix
                result['user'] = username
                result['user_extra'] = user_extra
            else:
                result['user_prefix'] = ''
                result['user'] = ''
                result['user_extra'] = ''
            
            return result
        except Exception:
            return None
    
    def read_log_file(filepath, log_type, max_lines=500):
        """Read a log file and return parsed entries in reverse chronological order"""
        entries = []
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                # Take last max_lines and reverse for newest first
                for line in reversed(lines[-max_lines:]):
                    entry = parse_log_line(line, log_type)
                    if entry:
                        entries.append(entry)
        except FileNotFoundError:
            pass
        except Exception:
            pass
        return entries
    
    audit_path = Path(log_dir) / 'audit' / 'audit.log'
    business_path = Path(log_dir) / 'business' / 'business.log'
    error_path = Path(log_dir) / 'errors' / 'error.log'
    
    audit_logs = read_log_file(audit_path, 'audit')
    business_logs = read_log_file(business_path, 'business')
    error_logs = read_log_file(error_path, 'error')
    
    return JsonResponse({
        'audit': audit_logs,
        'business': business_logs,
        'error': error_logs,
    })
